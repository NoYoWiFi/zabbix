#!/usr/bin/python3
# coding:utf-8

import argparse
import base64
from concurrent.futures import ThreadPoolExecutor
import datetime
from datetime import datetime
import hashlib
import hmac
import inspect
import json
import openpyxl
import re
import sys
import time
from time import mktime
import requests
from requests.adapters import HTTPAdapter
from urllib.parse import urlencode
from wsgiref.handlers import format_date_time
from zabbix_api import CusZabbixApi
from zabbix_api import CusExcelOp
from zabbix_api import CusMyThreadSendDir
from zabbix_api import CusMyThreadCfgZabbixAgent
from zabbix_api import CusMyThreadCfgSj
from zabbix_api import GV_CPU_COUNT


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='zabbix  api ', usage='%(prog)s [options]')
    # ![14]
    parser.add_argument('-get_item_history', nargs='?', metavar='无参数', dest='get_item_history', default='get_item_history',
                        help=u"按主机名批量计算历史最小值、平均值、最大值")
    # ![]
    parser.add_argument('-get_all_history', nargs='?', metavar='无参数', dest='get_all_history', default='get_all_history',
                        help=u"按主机名批量导出历史数据")
    # ![]
    parser.add_argument('-get_hostgroup_host', nargs='?', metavar='无参数', dest='get_hostgroup_host', default='get_hostgroup_host',
                        help=u"为主机组批量获取主机名")
    # ![]
    parser.add_argument('-get_hostgroup_item', nargs='?', metavar='无参数', dest='get_hostgroup_item', default='get_hostgroup_item',
                        help=u"为主机组批量获取主机监控项")
    # ![]
    parser.add_argument('-stop_all_priority_trigger', nargs='?', metavar='无参数', dest='stop_all_priority_trigger', default='stop_all_priority_trigger',
                        help=u"批量停止已启用触发器")
    parser.add_argument('-start_all_priority_trigger', nargs='?', metavar='无参数', dest='start_all_priority_trigger', default='start_all_priority_trigger',
                        help=u"批量启用已停止触发器")
    # ![]
    parser.add_argument('-stop_all_unsupport_item', nargs='?', metavar='无参数', dest='stop_all_unsupport_item', default='stop_all_unsupport_item',
                        help=u"批量停止不支持的监控项")
    parser.add_argument('-start_all_unsupport_item', nargs='?', metavar='无参数', dest='start_all_unsupport_item', default='start_all_unsupport_item',
                        help=u"批量启用不支持的监控项")
    parser.add_argument('-massadd_host_template_base_20221003', nargs='?', metavar='无参数', dest='massadd_host_template_base_20221003', default='massadd_host_template_base_20221003',
                        help=u"主机组下所有主机附加模板")
    parser.add_argument('-massupdate_host_template_base_20221003', nargs='?', metavar='无参数', dest='massupdate_host_template_base_20221003', default='massupdate_host_template_base_20221003',
                        help=u"主机组下所有主机更新模板")
    parser.add_argument('-def_massremove_host_templateids_clear_base_20221003', nargs='?', metavar='无参数', dest='def_massremove_host_templateids_clear_base_20221003', default='def_massremove_host_templateids_clear_base_20221003',
                        help=u"主机组下所有主机脱离模板清理监控项")
    parser.add_argument('-get_all_alert', nargs='?', metavar='无参数', dest='get_all_alert', default='get_all_alert',
                        help=u"获取所有告警信息")
    parser.add_argument('-get_all_event', nargs='?', metavar='无参数', dest='get_all_event', default='get_all_event',
                        help=u"获取所有事件信息")
    parser.add_argument('-createfile', nargs='?', metavar='无参数', dest='createfile', default='createfile',
                        help=u"生成配置文件")
    parser.add_argument('-senddir', nargs='?', metavar='无参数', dest='senddir', default='senddir',
                        help=u"下发文件")
    parser.add_argument('-sendcfg', nargs='?', metavar='无参数', dest='sendcfg', default='sendcfg',
                        help=u"配置代理")
    parser.add_argument('-sendsj', nargs='?', metavar='无参数', dest='sendsj', default='sendsj',
                        help=u"配置审计")
    parser.add_argument('-v', '--version', action='version', version='%(prog)s 如有问题请联系作者QQ1284524409',
                        help=u"如有问题请联系作者QQ1284524409")
    if len(sys.argv) == 1:
        print(parser.print_help())
    else:
        args = parser.parse_args()
        cus_zabbix_api = CusZabbixApi()
        cus_excel_op = CusExcelOp()
        # ![14_按主机批量计算历史最小值、平均值、最大值]
        if args.get_item_history != 'get_item_history':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=14)

            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            column_2_list = []
            column_3_list = []
            column_5_list = []
            column_6_list = []
            column_4_list = []
            for i in range(len(column_1_list)):
                column_2_list.append(cus_excel_op.def_get_cell_value(i + 2, 2))
                column_3_list.append(cus_excel_op.def_get_cell_value(i + 2, 3))
                column_5_list.append(cus_zabbix_api.def_timecovert(str(cus_excel_op.def_get_cell_value(i + 2, 5))))
                column_6_list.append(cus_zabbix_api.def_timecovert(str(cus_excel_op.def_get_cell_value(i + 2, 6))))
                column_4_list.append(cus_excel_op.def_get_cell_value(i + 2, 4))

            title_name = ['主机名', '监控项键值', '数据类型', '单位', '开始时间', '结束时间', '最小值', '平均值', '最大值']
            cus_excel_op.def_creat_excel()
            cus_excel_op.def_create_sheet(cus_excel_op.sheet_name)
            [cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]

            executor = ThreadPoolExecutor(GV_CPU_COUNT)
            lv_list_get_all_host_itemid = []
            lv_result = None
            for lv_result in executor.map(cus_zabbix_api.def_get_host_item, column_1_list, column_2_list):
                if lv_result['tag'] is True:
                    lv_list_get_all_host_itemid.append(lv_result['result'][0]['itemid'])
                else:
                    lv_list_get_all_host_itemid.append('')

            executor = ThreadPoolExecutor(GV_CPU_COUNT)
            lv_list_get_all_item_history = []
            lv_result = None
            lv_list_01 = []
            i = 1
            for lv_result in executor.map(cus_zabbix_api.def_get_item_history, lv_list_get_all_host_itemid, column_3_list, column_5_list, column_6_list):
                if lv_result['tag'] is True:
                    for lv_int_01 in range(len(lv_result['result'])):
                        lv_list_get_all_item_history.append(float(lv_result['result'][lv_int_01]['value']))
                    if column_4_list[i - 1] is None:
                        lv_list_01.append({'value': lv_list_get_all_item_history, 'type': 'N'})
                        # print(dict4)
                    else:
                        lv_list_01.append({'value': lv_list_get_all_item_history, 'type': column_4_list[i - 1]})
                    lv_list_get_all_item_history = []
                else:
                    lv_list_01.append({'value': [], 'type': ''})
                i = i + 1
            i = 0
            for value_list in lv_list_01:
                min_value = None
                avge_value = None
                max_value = None
                if value_list['type'] == 'Mb':
                    min_value = float('%.2f' % (min(value_list['value']) / 1024 / 1024))
                    avge_value = float('%.2f' % (sum(value_list['value']) / len(value_list['value']) / 1024 / 1024))
                    max_value = float('%.2f' % (max(value_list['value']) / 1024 / 1024))
                elif value_list['type'] == 'N':
                    min_value = float('%.2f' % (min(value_list['value'])))
                    avge_value = float('%.2f' % (sum(value_list['value']) / len(value_list['value'])))
                    max_value = float('%.2f' % (max(value_list['value'])))
                print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 按主机批量计算历史值: \033[;32m%s\033[0m 成功 返回值为: \033[;32m%s\033[0m \033[;32m%s\033[0m \033[;32m%s\033[0m'
                      % (len(lv_list_01), i + 1, column_1_list[i], min_value, avge_value, max_value))
                for x in range(1, cus_excel_op.def_get_row_clo_num()['columns'] + 1):
                    if x <= 6:
                        cus_excel_op.def_set_cell_value(i + 2, x, cus_excel_op.def_get_cell_value(i + 2, x))
                    if x == 7:
                        cus_excel_op.def_set_cell_value(i + 2, x, min_value)
                    if x == 8:
                        cus_excel_op.def_set_cell_value(i + 2, x, avge_value)
                    if x == 9:
                        cus_excel_op.def_set_cell_value(i + 2, x, max_value)
                i = i + 1
            cus_excel_op.def_save_create_xlsx(cus_excel_op.sheet_name + '.xlsx')
        # ![15_按主机批量导出历史数据]
        elif args.get_all_history != 'get_all_history':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=15)

            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            column_2_list = []
            column_3_list = []
            column_4_list = []
            column_5_list = []
            column_6_list = []
            for i in range(len(column_1_list)):
                column_2_list.append(cus_excel_op.def_get_cell_value(i + 2, 2))
                column_3_list.append(cus_excel_op.def_get_cell_value(i + 2, 3))
                column_4_list.append(cus_excel_op.def_get_cell_value(i + 2, 4))
                column_5_list.append(cus_zabbix_api.def_timecovert(str(cus_excel_op.def_get_cell_value(i + 2, 5))))
                column_6_list.append(cus_zabbix_api.def_timecovert(str(cus_excel_op.def_get_cell_value(i + 2, 6))))

            executor = ThreadPoolExecutor(GV_CPU_COUNT)
            lv_list_get_all_host_itemid = []
            lv_result = None
            for lv_result in executor.map(cus_zabbix_api.def_get_host_item, column_1_list, column_2_list):
                if lv_result['tag'] is True:
                    lv_list_get_all_host_itemid.append(lv_result['result'][0]['itemid'])
                else:
                    lv_list_get_all_host_itemid.append('')

            executor = ThreadPoolExecutor(GV_CPU_COUNT)
            lv_list_get_all_item_history = []
            lv_result = None
            lv_list_01 = []
            lv_list_02 = []
            i = 1
            for lv_result in executor.map(cus_zabbix_api.def_get_all_history, lv_list_get_all_host_itemid, column_3_list, column_5_list, column_6_list):
                if lv_result['tag'] is True:
                    if column_4_list[i - 1] == 'Mb':
                        for lv_int_01 in range(len(lv_result['result'])):
                            lv_list_get_all_item_history.append(float(lv_result['result'][lv_int_01]['value']) / 1024 / 1024)
                            lv_list_02.append(time.strftime("%Y-%m-%d %H:%M:%S",
                                                            time.localtime(float(lv_result['result'][lv_int_01]['clock']))))
                        lv_list_01.append({'value': lv_list_get_all_item_history, 'type': column_4_list[i - 1], 'clock': lv_list_02})
                    else:
                        for lv_int_01 in range(len(lv_result['result'])):
                            lv_list_get_all_item_history.append(lv_result['result'][lv_int_01]['value'])
                            lv_list_02.append(time.strftime("%Y-%m-%d %H:%M:%S",
                                                            time.localtime(float(lv_result['result'][lv_int_01]['clock']))))
                        lv_list_01.append({'value': lv_list_get_all_item_history, 'type': column_4_list[i - 1], 'clock': lv_list_02})
                    lv_list_get_all_item_history = []
                else:
                    lv_list_01.append({'value': [], 'type': '', "clock": []})
                i = i + 1
            i = None
            cus_excel_op.def_creat_excel()
            for lv_int_01 in range(len(lv_list_01)):
                item_name = re.sub(r'\W', "", column_2_list[lv_int_01])
                pattern = re.compile(r'^(.{1,29}).*$')
                item_name = re.search(pattern, item_name).group(1)
                cus_excel_op.def_create_sheet(str(lv_int_01 + 1) + '_' + item_name)
                title_name = ['主机名', '监控项键值', '数据类型', '单位', '开始时间', '结束时间', '收到值的时间', '收到的值']
                [cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                for y in range(len(title_name)):
                    for lv_int_03 in range(len(lv_list_01[lv_int_01]['value'])):
                        if y <= 5:
                            cus_excel_op.def_set_cell_value(lv_int_03 + 2, y + 1, cus_excel_op.def_get_cell_value(lv_int_01 + 2, y + 1))
                        elif y == 6:
                            cus_excel_op.def_set_cell_value(lv_int_03 + 2, y + 1, lv_list_01[lv_int_01]['clock'][lv_int_03])
                        elif y == 7:
                            cus_excel_op.def_set_cell_value(lv_int_03 + 2, y + 1, lv_list_01[lv_int_01]['value'][lv_int_03])
                print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 按主机批量获取历史值: \033[;32m%s\033[0m 成功'
                      % (len(column_1_list), lv_int_01 + 1, column_1_list[lv_int_01]))
            cus_excel_op.def_save_create_xlsx(cus_excel_op.sheet_name + '.xlsx')
        # ![16_按主机组导出主机名]
        elif args.get_hostgroup_host != 'get_hostgroup_host':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=16)
            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]

            executor = ThreadPoolExecutor(GV_CPU_COUNT)
            lv_list_get_all_host_groupid = []
            lv_result = None
            for lv_result in executor.map(cus_zabbix_api.def_get_hostgroup_6_0, column_1_list):
                if lv_result['tag'] is True:
                    lv_list_get_all_host_groupid.append([lv_result['result'][0]['groupid']])
                else:
                    lv_list_get_all_host_groupid.append([])

            executor = ThreadPoolExecutor(GV_CPU_COUNT)
            lv_list_get_all_hostgroup_host = []
            lv_dic_01 = {}
            lv_result = None
            lv_int_05 = 1
            for lv_result in executor.map(cus_zabbix_api.def_get_hostgroup_host, lv_list_get_all_host_groupid):
                if lv_result['tag'] is True:
                    for lv_int_04 in range(len(lv_result['result'])):
                        lv_list_get_all_hostgroup_host.append({'host': lv_result['result'][lv_int_04]['host'], 'name': lv_result['result'][lv_int_04]['name']})
                else:
                    lv_list_get_all_hostgroup_host.append({'host': '', 'name': ''})
                lv_dic_01.update({lv_int_05: lv_list_get_all_hostgroup_host})
                lv_list_get_all_hostgroup_host = []
                lv_int_05 = lv_int_05 + 1
            lv_int_05 = None

            cus_excel_op.def_creat_excel()
            for lv_int_03 in range(len(column_1_list)):
                item_name = re.sub(r'\W', "", column_1_list[lv_int_03])
                pattern = re.compile(r'^(.{1,29}).*$')
                item_name = re.search(pattern, item_name).group(1)
                cus_excel_op.def_create_sheet(str(lv_int_03 + 1) + '_' + item_name)
                title_name = ['主机组', '主机名', '可见名']
                [cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                lv_int_06 = 0
                for lv_int_07 in list(lv_dic_01.values())[lv_int_03]:
                    for lv_int_02 in range(len(title_name)):
                        if lv_int_02 == 0:
                            cus_excel_op.def_set_cell_value(lv_int_06 + 2, lv_int_02 + 1, column_1_list[lv_int_03])
                        if lv_int_02 == 1:
                            cus_excel_op.def_set_cell_value(lv_int_06 + 2, lv_int_02 + 1, lv_int_07['host'])
                        elif lv_int_02 == 2:
                            cus_excel_op.def_set_cell_value(lv_int_06 + 2, lv_int_02 + 1, lv_int_07['name'])
                    lv_int_06 = lv_int_06 + 1
                lv_int_06 = None
                print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 按主机组导出主机: \033[;32m%s\033[0m 成功'
                      % (len(column_1_list), lv_int_03 + 1, column_1_list[lv_int_03]))
            cus_excel_op.def_save_create_xlsx(cus_excel_op.sheet_name + '.xlsx')
        # ![17_批量停止已启用触发器]
        elif args.stop_all_priority_trigger != 'stop_all_priority_trigger':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=17)
            priority = []
            for i in range(2, cus_excel_op.def_get_row_clo_num()['rows'] + 1):
                try:
                    priority.append(cus_excel_op.def_get_cell_value(i, 1))
                except Exception as e:
                    print(u"%s 表第\033[;31m%s\033[0m行数据异常 共\033[;31m%s\033[0m行\n\033[;31m%s\033[0m" % (cus_excel_op.sheet_name, i, cus_excel_op.def_get_row_clo_num()['rows'] + 1, e))
                    sys.exit(1)
            cus_zabbix_api.def_stop_all_priority_trigger(priority)
        elif args.start_all_priority_trigger != 'start_all_priority_trigger':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=17)
            priority = []
            for i in range(2, cus_excel_op.def_get_row_clo_num()['rows'] + 1):
                try:
                    priority.append(cus_excel_op.def_get_cell_value(i, 1))
                except Exception as e:
                    print(u"%s 表第\033[;31m%s\033[0m行数据异常 共\033[;31m%s\033[0m行\n\033[;31m%s\033[0m" % (cus_excel_op.sheet_name, i, cus_excel_op.def_get_row_clo_num()['rows'] + 1, e))
                    sys.exit(1)
            cus_zabbix_api.def_start_all_priority_trigger(priority)
        # ![批量停止不支持监控项]
        elif args.stop_all_unsupport_item != 'stop_all_unsupport_item':
            cus_zabbix_api.def_stop_all_unsupport_item()
        elif args.start_all_unsupport_item != 'start_all_unsupport_item':
            cus_zabbix_api.def_start_all_unsupport_item()

        # ![18_根据监控项历史值关联模板]
        elif args.massadd_host_template_base_20221003 != 'massadd_host_template_base_20221003':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=18)
            column_1_list = cus_excel_op.def_get_col_value_base_20221003(1)
            del column_1_list[0]
            sheet_name = u'附加模板'
            title_name = ['序号', '主机名', 'IP', '模板名称', '结果', '原因']
            cus_excel_op.def_creat_excel()
            cus_excel_op.def_create_sheet(cus_excel_op.sheet_name)
            [cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
            [
                [
                    [
                        (
                            (print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> ' % (len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])), x + 1), end=''), cus_zabbix_api.def_massadd_host_template_base_20221003(cus_excel_op, x + 1, cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], ''.join([f_i['ip'] for f_i in cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['interfaces']]), cus_excel_op.def_get_cell_value(2, 5), [{'hostid': cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['hostid']}], cus_zabbix_api.def_get_template(cus_excel_op.def_get_cell_value(2, 5))))
                            if cus_zabbix_api.def_get_template(cus_excel_op.def_get_cell_value(2, 5)) != [] else
                            [print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 主机: \033[;31m%s\033[0m 关联模板 \033[;31m%s\033[0m 失败! 原因: \033[;31m%s\033[0m"' % (len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])), x + 1, cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], cus_excel_op.def_get_cell_value(2, 5), '未找到定义的模板名')), cus_excel_op.def_set_cell_value(x + 2, 1, x + 1), cus_excel_op.def_set_cell_value(x + 2, 2, cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host']), cus_excel_op.def_set_cell_value(x + 2, 3, ''.join([f_i['ip'] for f_i in cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['interfaces']])), cus_excel_op.def_set_cell_value(x + 2, 4, cus_excel_op.def_get_cell_value(2, 5)), cus_excel_op.def_set_cell_value(x + 2, 5, '失败'), cus_excel_op.def_set_cell_value(x + 2, 6, '未找到定义的模板名')]
                        ) for x in range(len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])))
                    ] for o in cus_zabbix_api.def_get_hostgroup_6_0(column_1_list[i])
                ] for i in range(len(column_1_list))
            ]
            cus_excel_op.def_save_create_xlsx('01_' + sheet_name + '.xlsx')
        elif args.massupdate_host_template_base_20221003 != 'massupdate_host_template_base_20221003':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=18)
            column_1_list = cus_excel_op.def_get_col_value_base_20221003(1)
            del column_1_list[0]
            column_2_list = cus_excel_op.def_get_col_value_base_20221003(2)
            del column_2_list[0]
            column_3_list = cus_excel_op.def_get_col_value_base_20221003(3)
            del column_3_list[0]
            dic_key_list = cus_excel_op.def_get_col_value(6)
            del dic_key_list[0]
            dic_value_list = cus_excel_op.def_get_col_value(7)
            del dic_value_list[0]
            dic_res = {}
            [dic_res.update({dic_key_list[f_i]: dic_value_list[f_i]}) for f_i in range(len(dic_key_list))]
            sheet_name = u'更新模板'
            title_name = ['序号', '主机名', 'IP', '历史值', '模板名称', '结果', '原因']
            cus_excel_op.def_creat_excel()
            cus_excel_op.def_create_sheet(cus_excel_op.sheet_name)
            [cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
            [
                [
                    [
                        [
                            (print('(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> ' % (len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])), x + 1), end=''),
                             cus_zabbix_api.def_massadd_host_template_base_item_20221003(cus_excel_op,
                                                                                         x + 1,
                                                                                         cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'],
                                                                                         ''.join([f_i['ip'] for f_i in cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['interfaces']]),
                                                                                         ''.join([''.join([''.join([a['value'] for a in cus_zabbix_api.def_get_item_history_base_20221003(z['itemid'], column_3_list[y])]) for z in cus_zabbix_api.def_get_host_item(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], column_2_list[y])]) for y in range(len(column_2_list))]),
                                                                                         dic_res.get(''.join([''.join([''.join([a['value'] for a in cus_zabbix_api.def_get_item_history_base_20221003(z['itemid'], column_3_list[y])]) for z in cus_zabbix_api.def_get_host_item(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], column_2_list[y])]) for y in range(len(column_2_list))])),
                                                                                         [{'hostid': c['hostid']}],
                                                                                         cus_zabbix_api.def_get_template(dic_res.get(''.join([''.join([''.join([a['value'] for a in cus_zabbix_api.def_get_item_history_base_20221003(z['itemid'], column_3_list[y])]) for z in cus_zabbix_api.def_get_host_item(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], column_2_list[y])]) for y in range(len(column_2_list))]))))
                             ) for c in cus_zabbix_api.def_get_host(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'])

                        ]
                        if dic_res.get(''.join([''.join([''.join([a['value'] for a in cus_zabbix_api.def_get_item_history_base_20221003(z['itemid'], column_3_list[y])]) for z in cus_zabbix_api.def_get_host_item(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], column_2_list[y])]) for y in range(len(column_2_list))])) is not None else
                        [
                            print('(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 主机: \033[;31m%s\033[0m 获取监控项值为: \033[;31m%s\033[0m 关联模板失败! 原因: 历史值对应关系异常' % (len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])), x + 1, cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], ''.join([''.join([''.join([a['value'] for a in cus_zabbix_api.def_get_item_history_base_20221003(z['itemid'], column_3_list[y])]) for z in cus_zabbix_api.def_get_host_item(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], column_2_list[y])]) for y in range(len(column_2_list))]))),
                            cus_excel_op.def_set_cell_value(x + 2, 1, x + 1),
                            cus_excel_op.def_set_cell_value(x + 2, 2, cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host']),
                            cus_excel_op.def_set_cell_value(x + 2, 3, ''.join([f_i['ip'] for f_i in cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['interfaces']])),
                            cus_excel_op.def_set_cell_value(x + 2, 4, ''.join([''.join([''.join([a['value'] for a in cus_zabbix_api.def_get_item_history_base_20221003(z['itemid'], column_3_list[y])]) for z in cus_zabbix_api.def_get_host_item(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], column_2_list[y])]) for y in range(len(column_2_list))])),
                            cus_excel_op.def_set_cell_value(x + 2, 5, ''),
                            cus_excel_op.def_set_cell_value(x + 2, 6, '失败'),
                            cus_excel_op.def_set_cell_value(x + 2, 7, '历史值对应关系异常')
                        ] for x in range(len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])))
                    ] for o in cus_zabbix_api.def_get_hostgroup_6_0(column_1_list[i])
                ] for i in range(len(column_1_list))
            ]
            cus_excel_op.def_save_create_xlsx('02_' + sheet_name + '.xlsx')
        elif args.def_massremove_host_templateids_clear_base_20221003 != 'def_massremove_host_templateids_clear_base_20221003':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=18)
            column_1_list = cus_excel_op.def_get_col_value_base_20221003(1)
            del column_1_list[0]
            sheet_name = u'脱离模板'
            title_name = ['序号', '主机名', 'IP', '模板名称', '结果', '原因']
            cus_excel_op.def_creat_excel()
            cus_excel_op.def_create_sheet(cus_excel_op.sheet_name)
            [cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
            [
                [
                    [
                        (
                            (
                                print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> ' % (len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])), x + 1), end=''),
                                cus_zabbix_api.def_massremove_host_templateids_clear_base_20221003(cus_excel_op,
                                                                                                   x + 1,
                                                                                                   cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'],
                                                                                                   ''.join([f_i['ip'] for f_i in cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['interfaces']]),
                                                                                                   cus_excel_op.def_get_cell_value(2, 5),
                                                                                                   [cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['hostid']],
                                                                                                   [u['templateid'] for u in cus_zabbix_api.def_get_template(cus_excel_op.def_get_cell_value(2, 5))]
                                                                                                   )
                            )
                            if cus_zabbix_api.def_get_template(cus_excel_op.def_get_cell_value(2, 5)) != [] else
                            [print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 主机: \033[;31m%s\033[0m 脱离模板 \033[;31m%s\033[0m 失败! 原因: \033[;31m%s\033[0m"' % (len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])), x + 1, cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host'], cus_excel_op.def_get_cell_value(2, 5), '未找到定义的模板名')), cus_excel_op.def_set_cell_value(x + 2, 1, x + 1), cus_excel_op.def_set_cell_value(x + 2, 2, cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host']), cus_excel_op.def_set_cell_value(x + 2, 3, ''.join([f_i['ip'] for f_i in cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['interfaces']])), cus_excel_op.def_set_cell_value(x + 2, 4, cus_excel_op.def_get_cell_value(2, 5)), cus_excel_op.def_set_cell_value(x + 2, 5, '失败'), cus_excel_op.def_set_cell_value(x + 2, 6, '未找到定义的模板名')]
                        ) for x in range(len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])))
                    ] for o in cus_zabbix_api.def_get_hostgroup_6_0(column_1_list[i])
                ] for i in range(len(column_1_list))]
            cus_excel_op.def_save_create_xlsx('03_' + sheet_name + '.xlsx')
        # ![34_获取所有告警信息]
        elif args.get_all_alert != 'get_all_alert':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=34)
            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            cus_excel_op.def_creat_excel()
            # def_get_all_alert | def_get_all_alert_custom
            executor = ThreadPoolExecutor(GV_CPU_COUNT)
            lv_list_get_action = []
            lv_result = None
            for lv_result in executor.map(cus_zabbix_api.def_get_action, column_1_list):
                if lv_result['tag'] is True:
                    lv_list_get_action.append(lv_result['result'][0]['actionid'])
                else:
                    lv_list_get_action.append('')

            res = [cus_zabbix_api.def_get_all_alert("".join(lv_list_get_action),
                                                    cus_zabbix_api.def_timecovert(str(cus_excel_op.def_get_cell_value(i_01 + 2, 2))),
                                                    cus_zabbix_api.def_timecovert(str(cus_excel_op.def_get_cell_value(i_01 + 2, 3))))
                   for i_01 in range(len(column_1_list))]
            for i_01 in range(len(res)):
                cus_excel_op.def_create_sheet(str(i_01 + 1))
                lv_list_02 = []
                if not res[i_01]:
                    print('告警列表为空，请检查告警动作与excel导出日期')
                    exit(1)
                [lv_list_02.append(i_04) for i_04 in res[i_01][0].keys()]
                for i in range(len(lv_list_02)):
                    cus_excel_op.def_set_cell_value(1, i + 1, lv_list_02[i])
                for i_02 in range(len(res[i_01])):
                    lv_list_01 = []
                    [lv_list_01.append(i_03) for i_03 in res[i_01][i_02].values()]
                    for i_04 in range(len(lv_list_02)):
                        try:
                            if lv_list_01[i_04] == '':
                                pass
                            # elif lv_list_02[i_04] == u'告警日期':
                            #     excel_op.def_set_cell_value(i_02 + 2, i_04 + 1, time.strftime("%Y-%m-%d", time.localtime(float(res[i_01][i_02][u'告警日期']))))
                            # elif lv_list_02[i_04] == u'告警时间':
                            #     excel_op.def_set_cell_value(i_02 + 2, i_04 + 1, time.strftime("%H:%M:%S", time.localtime(float(res[i_01][i_02][u'告警时间']))))
                            elif lv_list_02[i_04] == u'执行状态' or lv_list_02[i_04] == u'告警级别':
                                cus_excel_op.def_set_cell_value(i_02 + 2, i_04 + 1,
                                                                cus_zabbix_api.def_convert_numeric_to_text(
                                                                    lv_list_01[i_04]))
                            elif lv_list_02[i_04] == u'触发器状态':
                                cus_excel_op.def_set_cell_value(i_02 + 2, i_04 + 1,
                                                                cus_zabbix_api.def_convert_numeric_to_text_problem(
                                                                    lv_list_01[i_04]))
                            else:
                                cus_excel_op.def_set_cell_value(i_02 + 2, i_04 + 1, lv_list_01[i_04])
                        except:
                            pass

                    print(u'(进度: -> \033[;34m%s\033[0m/\033[;34m%s\033[0m)\n' % (len(res[i_01]), i_02 + 1), end='')
            cus_excel_op.def_save_create_xlsx(cus_excel_op.sheet_name + '.xlsx')
        # ![35_获取所有事件信息]
        elif args.get_all_event != 'get_all_event':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=35)
            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            list_group_host = []
            [[[list_group_host.append(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])[x]['host']) for x in range(len(cus_zabbix_api.def_get_hostgroup_host(o['groupid'])))] for o in cus_zabbix_api.def_get_hostgroup_6_0(column_1_list[i])] for i in range(len(column_1_list))]
            column_2_list = cus_excel_op.def_get_col_value(2)
            del column_2_list[0]
            list_all_event_host = []
            for i_01 in range(len(column_2_list)):
                for str_01 in cus_zabbix_api.def_get_all_event(cus_excel_op.def_get_cell_value(i_01 + 2, 2),
                                                               cus_zabbix_api.def_timecovert(str(cus_excel_op.def_get_cell_value(i_01 + 2, 3))),
                                                               cus_zabbix_api.def_timecovert(str(cus_excel_op.def_get_cell_value(i_01 + 2, 4)))):
                    for str_02 in str_01['hosts']:
                        list_all_event_host.append(str_02['host'])
            res = [val for val in list_group_host if val in list_all_event_host]
            list_ip = []
            str_ip = ""
            list_ip_tmp = []
            for i in range(len(res)):
                for o in cus_zabbix_api.def_get_host(res[i]):
                    for u in cus_zabbix_api.def_get_host_ip([(o["hostid"])]):
                        for x in u['interfaces']:
                            list_ip_tmp.append(x['ip'])
                        str_ip = ",".join(list_ip_tmp)
                        list_ip.append(str_ip)
                        list_ip_tmp = []
                        str_ip = ""
            title_name = ['主机名', 'IP地址']
            cus_excel_op.def_creat_excel()
            cus_excel_op.def_create_sheet(cus_excel_op.sheet_name)
            [cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
            for i in range(len(res)):
                for o in range(len(title_name)):
                    if o == 0:
                        cus_excel_op.def_set_cell_value(i + 2, 1, res[i])
                    if o == 1:
                        cus_excel_op.def_set_cell_value(i + 2, 2, list_ip[i])
            cus_excel_op.def_save_create_xlsx(cus_excel_op.sheet_name + '.xlsx')
            [[print(u'(进度: -> \033[;34m%s\033[0m/\033[;34m%s\033[0m)' % (len(res), i + 1), end=''), cus_zabbix_api.def_delete_host([(o["hostid"]) for o in cus_zabbix_api.def_get_host(res[i])])] for i in range(len(res))]

        # ![36_下发文件]
        elif args.createfile != 'createfile':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=36)
            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            # 根据需要生成配置文件
            with open('senddir/ip/hosts', 'w') as s_hosts:
                for n_1 in range(len(column_1_list)):
                    s_hosts.write(cus_excel_op.def_get_cell_value(n_1 + 2, 6) +
                                  ' ' + cus_excel_op.def_get_cell_value(n_1 + 2, 11) +
                                  ' ' + cus_excel_op.def_get_cell_value(n_1 + 2, 12) + '\n')
            s_hosts.close()
            with open('senddir/ip/ip.txt', 'w') as s_hosts:
                for n_1 in range(len(column_1_list)):
                    s_hosts.write(cus_excel_op.def_get_cell_value(n_1 + 2, 6) + " " + cus_excel_op.def_get_cell_value(n_1 + 2, 3) + '\n')
            s_hosts.close()
            with open('senddir/ip/mysql.txt', 'w') as s_hosts:
                s_hosts.write(cus_excel_op.def_get_cell_value(11, 6) + '\n')
            s_hosts.close()
            with open('senddir/ip/zookeeper.txt', 'w') as s_hosts:
                s_hosts.write(cus_excel_op.def_get_cell_value(5, 11) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(6, 11) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(7, 11) + ' \n')
            s_hosts.close()
            with open('senddir/ip/kafka.txt', 'w') as s_hosts:
                s_hosts.write(cus_excel_op.def_get_cell_value(5, 11) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(6, 11) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(7, 11) + ' \n')
            s_hosts.close()
            with open('senddir/ip/elasticsearch.txt', 'w') as s_hosts:
                s_hosts.write(cus_excel_op.def_get_cell_value(8, 6) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(9, 6) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(10, 6) + ' \n')
                s_hosts.write(cus_excel_op.def_get_cell_value(8, 11) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(9, 11) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(10, 11) + ' \n')
            s_hosts.close()
            with open('senddir/ip/kibana.txt', 'w') as s_hosts:
                s_hosts.write(cus_excel_op.def_get_cell_value(8, 6) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(9, 6) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(10, 6) + ' \n')
                s_hosts.write(cus_excel_op.def_get_cell_value(8, 11) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(9, 11) + ' ')
                s_hosts.write(cus_excel_op.def_get_cell_value(10, 11) + ' \n')
            s_hosts.close()
            with open('senddir/ip/nacos.txt', 'w') as s_hosts:
                s_hosts.write('mysql ' + cus_excel_op.def_get_cell_value(11, 6) + ' \n')
                s_hosts.write('redis ' + cus_excel_op.def_get_cell_value(11, 6) + ' \n')
                s_hosts.write('kafka ' + cus_excel_op.def_get_cell_value(5, 6) + ':9092,')
                s_hosts.write(cus_excel_op.def_get_cell_value(6, 6) + ':9092,')
                s_hosts.write(cus_excel_op.def_get_cell_value(7, 6) + ':9092 \n')
                s_hosts.write('elasticsearch ' + cus_excel_op.def_get_cell_value(8, 6) + ':9200,')
                s_hosts.write(cus_excel_op.def_get_cell_value(9, 6) + ':9200,')
                s_hosts.write(cus_excel_op.def_get_cell_value(10, 6) + ':9200 \n')
            s_hosts.close()
            with open('senddir/ip/spark.txt', 'w') as s_hosts:
                s_hosts.write(cus_excel_op.def_get_cell_value(3, 11) + ' \n')
                s_hosts.write(cus_excel_op.def_get_cell_value(4, 11) + ' \n')
                s_hosts.write(cus_excel_op.def_get_cell_value(5, 11) + ' \n')
            s_hosts.close()
            # 根据需要生成配置文件
            print('配置文件生成完毕')

        # ![36_下发文件]
        elif args.senddir != 'senddir':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=36)
            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            for n_1 in range(len(column_1_list)):
                total_num = len(column_1_list)
                cur_num = n_1 + 1
                ssh_ip = cus_excel_op.def_get_cell_value(n_1 + 2, 1)
                port = int(cus_excel_op.def_get_cell_value(n_1 + 2, 2))
                pwd = cus_excel_op.def_get_cell_value(n_1 + 2, 3)
                # print(ssh_ip, port, pwd)
                try:
                    m = CusMyThreadSendDir(cur_num, total_num, ssh_ip, port, pwd)
                    m.start()
                    print(u'(进度: -> \033[;34m%s\033[0m/\033[;34m%s\033[0m \033[;34m%s\033[0m)' % (total_num, cur_num, ssh_ip), end='')
                except Exception as e:
                    print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], e))
                    sys.exit(1)

        elif args.sendcfg != 'sendcfg':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=36)
            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            for n_1 in range(len(column_1_list)):
                total_num = len(column_1_list)
                cur_num = n_1 + 1
                ssh_ip = cus_excel_op.def_get_cell_value(n_1 + 2, 1)
                port = int(cus_excel_op.def_get_cell_value(n_1 + 2, 2))
                pwd = cus_excel_op.def_get_cell_value(n_1 + 2, 3)
                zabbix_server_ip = cus_excel_op.def_get_cell_value(n_1 + 2, 4)
                # print(ssh_ip, port, pwd)
                try:
                    m = CusMyThreadCfgZabbixAgent(cur_num, total_num, ssh_ip, port, pwd, zabbix_server_ip)
                    m.start()
                except Exception as e:
                    print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], e))
                    sys.exit(1)

        elif args.sendsj != 'sendsj':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=36)
            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            for n_1 in range(len(column_1_list)):
                total_num = len(column_1_list)
                cur_num = n_1 + 1
                s_dic = {'ssh_ip': "", 'port': "", 'pwd': "", 's_eth': "", 's_ip': "", 's_newmask': "", 's_gateway': "",
                         's_dns1': "", 's_dns2': "", 's_host': "", 's_app': ""}
                s_dic['ssh_ip'] = cus_excel_op.def_get_cell_value(n_1 + 2, 1)
                s_dic['port'] = int(cus_excel_op.def_get_cell_value(n_1 + 2, 2))
                s_dic['pwd'] = cus_excel_op.def_get_cell_value(n_1 + 2, 3)
                s_dic['s_eth'] = cus_excel_op.def_get_cell_value(n_1 + 2, 5)
                s_dic['s_ip'] = cus_excel_op.def_get_cell_value(n_1 + 2, 6)
                s_dic['s_newmask'] = cus_excel_op.def_get_cell_value(n_1 + 2, 7)
                s_dic['s_gateway'] = cus_excel_op.def_get_cell_value(n_1 + 2, 8)
                s_dic['s_dns1'] = cus_excel_op.def_get_cell_value(n_1 + 2, 9)
                s_dic['s_dns2'] = cus_excel_op.def_get_cell_value(n_1 + 2, 10)
                s_dic['s_host'] = cus_excel_op.def_get_cell_value(n_1 + 2, 11)
                s_dic['s_app'] = cus_excel_op.def_get_cell_value(n_1 + 2, 12)
                # print(ssh_ip, port, pwd)
                try:
                    m = CusMyThreadCfgSj(cur_num, total_num, s_dic)
                    m.start()
                except Exception as e:
                    print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], e))
                    sys.exit(1)
        # ![16_按主机组导出主机名]
        elif args.get_hostgroup_item != 'get_hostgroup_item':
            cus_excel_op.def_load_excel(file='zabbix_api.xlsx', index=48)
            column_1_list = cus_excel_op.def_get_col_value(1)
            del column_1_list[0]
            column_2_list = []
            for i in range(len(column_1_list)):
                column_2_list.append(cus_excel_op.def_get_cell_value(i + 2, 2))

            executor = ThreadPoolExecutor(GV_CPU_COUNT)
            lv_list_itemid = []
            lv_list_hostid = []
            lv_list_name = []
            lv_list_key_ = []
            lv_list_lastclock = []
            lv_list_lastvalue = []
            lv_list_prevvalue = []
            lv_result = None
            for lv_result in executor.map(cus_zabbix_api.def_get_group_item, column_1_list, column_2_list):
                if lv_result['tag'] is True:
                    for lv_int01 in range(len(lv_result['result'])):
                        lv_list_itemid.append(lv_result['result'][lv_int01]['itemid'])
                        lv_list_hostid.append(lv_result['result'][lv_int01]['hostid'])
                        lv_list_name.append(lv_result['result'][lv_int01]['name'])
                        lv_list_key_.append(lv_result['result'][lv_int01]['key_'])
                        lv_list_lastclock.append(cus_zabbix_api.def_timeCovertIntToYMD(int(lv_result['result'][lv_int01]['lastclock'])))
                        lv_list_lastvalue.append(int(lv_result['result'][lv_int01]['lastvalue']))
                        lv_list_prevvalue.append(int(lv_result['result'][lv_int01]['prevvalue']))

                else:
                    lv_list_itemid.append('')
                    lv_list_hostid.append('')
                    lv_list_name.append('')
                    lv_list_key_.append('')
                    lv_list_lastclock.append('')
                    lv_list_lastvalue.append('')
                    lv_list_prevvalue.append('')

            cus_excel_op.def_creat_excel()
            for lv_int_03 in range(len(column_1_list)):
                cus_excel_op.def_create_sheet('目录积压情况')
                title_name = ['IP', '归属', '来源', '协议', '目录', '队列积压情况', '队列消费情况', '时间']
                [cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                for lv_int02 in range(len(lv_list_itemid)):
                    try:
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 1, lv_list_name[lv_int02].split('__')[0].split('_')[0])
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 2, lv_list_name[lv_int02].split('__')[0].split('_')[1])
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 3, lv_list_name[lv_int02].split('__')[0].split('_')[2])
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 4, lv_list_name[lv_int02].split('__')[0].split('_')[3])
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 5, lv_list_name[lv_int02].split('__')[1].replace('_', '/').replace('通道队列深度', ''))
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 6, lv_list_lastvalue[lv_int02])
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 7, lv_list_lastvalue[lv_int02] - lv_list_prevvalue[lv_int02])
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 8, lv_list_lastclock[lv_int02])
                    except:
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 6, lv_list_lastvalue[lv_int02])
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 7, lv_list_lastvalue[lv_int02] - lv_list_prevvalue[lv_int02])
                        cus_excel_op.def_set_cell_value(lv_int02 + 2, 8, lv_list_lastclock[lv_int02])
                        pass
                    print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 按主机组导出主机: \033[;32m%s\033[0m 成功'
                          % (len(lv_list_itemid), lv_int02 + 1, lv_list_name[lv_int02]))
            cus_excel_op.def_save_create_xlsx(cus_excel_op.sheet_name + '.xlsx')
