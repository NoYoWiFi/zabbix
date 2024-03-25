#!/etc/zabbix/venv_centos8/bin/python3
### !/etc/zabbix/venv_centos7/bin/python3
# -*- coding: utf-8 -*-

import argparse
import inspect
import logging
import logging.handlers
import json
import socket

import openpyxl
import os
from ping3 import ping
import paramiko
import re
import subprocess
import sys
import telnetlib
import time
from stat import S_ISDIR as isdir
from concurrent.futures import ThreadPoolExecutor

import requests
from requests.adapters import HTTPAdapter
from urllib.parse import urlencode

# python3 -m pip install --upgrade pip
# pip3 install openpyxl paramiko requests
# pip3 install  cryptography==3.4.8
# telnet ./zabbix_sender.py --type='telnet' --ip='172.169.10.10' --port='23' --user='admin' --pwd='admin123' --host="test" --discovery
# ssh ./zabbix_sender.py  --type='ssh' --ip='127.0.0.1' --port='22' --user='root' --pwd='123.com' --host="test" --discovery
# web ./zabbix_sender.py  --type='web' --url_schoolid='2c9f83fd6cc8d788016ccc71706b01de' --url_userid='000000082702' --url_pwd='1f82c942befda29b6ed487a51da199f78fce7f05' --url_http='http://testwww.qlzhy.com/trade-web/login/login.do' --host="test" --discovery
# web ./zabbix_sender.py  --type='web' --url_schoolid='' --url_userid='Admin' --url_pwd='zabbix' --url_http='http://172.169.10.4/api_jsonrpc.php' --host="test" --discovery
# linux ./zabbix_sender.py  --type='linux' --host="test" --discovery
# Create log-object


LOG_FILENAME = "/tmp/stateTmpFile.log"
# sys.argv[5] contain this string "--storage_name=<storage_name_in_zabbix>". List slicing delete this part "--storage_name="
STORAGE_NAME = sys.argv[2][9:]
logger = logging.getLogger("logger")
logger.setLevel(logging.INFO)
XLSX_FILENAME = '/tmp/zabbix_sender.xlsx'
SERVER_IP = '172.169.10.3'
SERVER_PORT = '10051'

GV_CPU_COUNT = os.cpu_count()

TIMESTAMPNOW = None

# Set handler
handler = logging.handlers.RotatingFileHandler(LOG_FILENAME, maxBytes=(1024 ** 2) * 10, backupCount=5)
formatter = logging.Formatter('{0} - %(asctime)s - %(name)s - %(levelname)s - %(message)s'.format(STORAGE_NAME))

# Set formatter for handler
handler.setFormatter(formatter)

# Add handler to log-object
logger.addHandler(handler)


class CusExcelOp(object):
    def __init__(self):
        """
        excel_op = ExcelOp(file='zabbix_api.xlsx', index=16)
        column_1_list = excel_op.def_get_col_value(1)
        del column_1_list[0]
        title_name = ['主机名']
        excel_op.def_create_sheet(excel_op.sheet_name)
        [excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
        [excel_op.def_set_cell_value(i + 2, 1, column_1_list[i]) for i in range(len(column_1_list))]
        excel_op.def_save_xlsx(excel_op.sheet_name)
        """
        self.file = None
        self.wb_object = None
        self.sheet_name = None
        self.ws_object = None

        self.create_ws_object = None
        self.create_wb_object = None

    def def_load_excel(self, file, index):
        if index == 0:
            print("index must at least 1")
            exit(1)
        self.file = file
        self.wb_object = openpyxl.load_workbook(self.file, data_only=True)
        sheet_name_list = self.wb_object.sheetnames
        self.sheet_name = sheet_name_list[index - 1]
        self.ws_object = self.wb_object[self.sheet_name]

    def def_creat_excel(self):
        self.create_wb_object = openpyxl.Workbook()

    def def_update_sheet_name(self, sheet_name):
        self.sheet_name = sheet_name
        try:
            self.ws_object = self.wb_object[sheet_name]
        except Exception as e:
            print(u"数据异常当前Zabbix版本不支持汉化: %s" % e)
        return None

    # 获取表格的总行数和总列数
    def def_get_row_clo_num(self):
        """
        :return: {"rows": rows, "columns": columns}
        """
        rows = self.ws_object.max_row
        columns = self.ws_object.max_column
        return {"rows": rows, "columns": columns}

    # 获取某个单元格的值
    def def_get_cell_value(self, row, column):
        """
        获取单元格的值

        :param row: 行号
        :param column: 列号
        :return: 返回字符串
        """
        cell_value = self.ws_object.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def def_get_col_value(self, column):
        """
        根据列号返回一列数据

        :param column: 列号: 3
        :return: 表格数据: ['a','b','c']
        """
        # rows = self.ws_object.max_row
        # cell_value = [self.ws_object.cell(row=i, column=column).value if self.ws_object.cell(row=i, column=column).value is not None else (
        #     print('第 %s 行错误! 原因: %s' % (i, u'错误! 列表存在空行')), exit(1)) for i in range(1, rows + 1)]
        v_dic_num = {"row": 1}
        v_flag = True
        v_list_row = []
        while v_flag:
            v_cell_value = self.ws_object.cell(row=v_dic_num['row'], column=column).value
            if v_cell_value is not None:
                v_list_row.append(v_cell_value)
                v_dic_num.update(row=v_dic_num['row'] + 1)
            else:
                v_flag = False
        return v_list_row

        # return cell_value

    def def_get_creat_col_value(self, column):
        """
        根据列号返回一列数据

        :param column: 列号: 3
        :return: 表格数据: ['a','b','c']
        """
        # rows = self.ws_object.max_row
        # cell_value = [self.ws_object.cell(row=i, column=column).value if self.ws_object.cell(row=i, column=column).value is not None else (
        #     print('第 %s 行错误! 原因: %s' % (i, u'错误! 列表存在空行')), exit(1)) for i in range(1, rows + 1)]
        v_dic_num = {"row": 1}
        v_flag = True
        v_list_row = []
        while v_flag:
            v_cell_value = self.create_ws_object.cell(row=v_dic_num['row'], column=column).value
            if v_cell_value is not None:
                v_list_row.append(v_cell_value)
                v_dic_num.update(row=v_dic_num['row'] + 1)
            else:
                v_flag = False
        return v_list_row

        # return cell_value

    def def_get_col_value_base_20221003(self, column):
        """
        根据列号返回一列数据

        :param column: 列号: 3
        :return: 表格数据: ['a','b','c']
        """
        rows = self.ws_object.max_row
        cell_value = [self.ws_object.cell(row=i, column=column).value for i in range(1, rows + 1) if self.ws_object.cell(row=i, column=column).value is not None]
        return cell_value

    # 获取某行所有值
    def def_get_row_value(self, row):
        """
        根据行号返回一行数据

        :param row: 行号: 1
        :return: 表格数据: ['a','b','c']
        """
        columns = self.ws_object.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws_object.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def def_create_sheet(self, sheet_name):
        """
        创建sheet页

        :param sheet_name: sheet名称: '页面1'
        :return: None
        """
        self.create_ws_object = self.create_wb_object.create_sheet(sheet_name)

    def def_load_create_sheet(self, sheet_name):
        """
        加载sheet页

        :param sheet_name: sheet名称: '页面1'
        :return: None
        """
        self.create_wb_object = openpyxl.load_workbook(XLSX_FILENAME, data_only=True)
        self.create_ws_object = self.create_wb_object[sheet_name]

    def def_save_xlsx(self, xlsx_name):
        """
        保存为:excel名称_时间戳.xlsx

        :param xlsx_name: excel名称: 表格1
        :return: None
        """
        del self.create_wb_object['Sheet']
        self.create_wb_object.save(xlsx_name + "_" + time.strftime('%Y%m%d_%H%M%S', time.gmtime()) + ".xlsx")

    def def_save_create_xlsx(self, xlsx_name):
        """
        保存为:excel名称_时间戳.xlsx

        :param xlsx_name: excel名称: 表格1
        :return: None
        """
        del self.create_wb_object['Sheet']
        self.create_wb_object.save(xlsx_name)

    # 设置某个单元格的值
    def def_set_cell_value(self, row, colunm, cellvalue):
        """
        设置某个单元格的值

        :param row: 行号: 1
        :param colunm: 列号: 3
        :param cellvalue: 单元格值: '成功'
        :return: None
        """
        try:
            self.create_ws_object.cell(row=row, column=colunm).value = cellvalue
        except:
            self.create_ws_object.cell(row=row, column=colunm).value = "writefail"


class CusTelnetClient(object):
    def __init__(self, ):
        self.zabbix_sender = CusZabbixSender()
        self.tn = telnetlib.Telnet()
        self.cus_excel_op = CusExcelOp()
        self.cus_excel_op.def_creat_excel()
        self.value = None
        self.local_methord = CusLocalMethod()

    # 此函数实现telnet登录主机
    def def_connect(self, host_ip, username, password, port):
        try:
            # self.tn = telnetlib.Telnet(host_ip,port=23)
            self.tn.open(host_ip, port=port)
            time.sleep(0.1)
        except:
            logger.error(u"错误: {0}".format(inspect.stack()[0][2]))
            return False
        # 等待login出现后输入用户名，最多等待10秒
        self.tn.read_until(b'Username:', timeout=10)
        self.tn.write(username.encode('ascii') + b'\n')
        time.sleep(0.3)
        # 等待Password出现后输入用户名，最多等待10秒
        self.tn.read_until(b'Password:', timeout=10)
        self.tn.write(password.encode('ascii') + b'\n')
        time.sleep(0.3)
        # 延时两秒再收取返回结果，给服务端足够响应时间
        # 获取登录结果
        # read_very_eager()获取到的是的是上次获取之后本次获取之前的所有输出
        command_result = self.tn.read_very_eager().decode('ascii')
        if 'Login incorrect' not in command_result:
            logger.info('%s登录成功' % host_ip)
            return True
        else:
            logger.error('%s登录失败，用户名或密码错误' % host_ip)
            return False

    # 此函数实现执行传过来的命令，并输出其执行结果
    def def_execute_command_with_more(self, command):
        # 执行命令
        self.tn.write(command.encode('ascii'))
        rackreply = self.tn.expect([], timeout=1)[2].decode().strip()
        # print(rackreply)
        while True:
            self.tn.read_until(b'\r\n[', timeout=0.1)
            self.tn.write(b'\r\n')
            time.sleep(0.1)
            stdout = self.tn.expect([], timeout=1)[2]
            if stdout.find(b'\r\n[') != -1:
                break
            else:
                self.tn.read_until(b'  ---- More ----', timeout=0.1)
                self.tn.write(b' ')
                time.sleep(0.1)
                stdout = self.tn.expect([], timeout=1)[2].decode().strip()
                rackreply = rackreply + stdout
                if stdout.find(u'\r\n[') != -1:
                    break
        # 获取命令结果
        return rackreply

    # 退出telnet
    def def_logout(self):
        try:
            self.tn.close()
            logger.info("Connection Closed Successfully")
            print(0)
        except Exception as oops:
            logger.info(u"错误: {0} {1}".format(inspect.stack()[0][2], oops))
            sys.exit("{0}".format(inspect.currentframe().f_lineno))

    def def_discovering_resources(self, user, password, ip, port, host, list_resources):
        try:
            self.def_connect(ip, user, password, port)
            for resource in list_resources:
                if ['disname'].count(resource) == 1:
                    logger.info("Starting discovering resource - {0}".format(resource))
                    one_object_list = {}
                    discovered_resource = []
                    one_object_list["{#INDEX}"] = 1
                    discovered_resource.append(one_object_list)
                    logger.info("Succes get resource - {0}".format(resource))
                    converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                    timestampnow = int(time.time())
                    TIMESTAMPNOW = timestampnow
                    xer = []
                    xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                elif ['discpu'].count(resource) == 1:
                    logger.info("Starting discovering resource - {0}".format(resource))
                    one_object_list = {}
                    discovered_resource = []
                    one_object_list["{#INDEX}"] = 1
                    discovered_resource.append(one_object_list)
                    logger.info("Succes get resource - {0}".format(resource))
                    converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                    timestampnow = int(time.time())
                    TIMESTAMPNOW = timestampnow
                    xer = []
                    xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                elif ['discpu_20230104'].count(resource) == 1:
                    logger.info("Starting discovering resource - {0}".format(resource))
                    one_object_list = {}
                    discovered_resource = []
                    one_object_list["{#INDEX}"] = 1
                    discovered_resource.append(one_object_list)
                    logger.info("Succes get resource - {0}".format(resource))
                    converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                    timestampnow = int(time.time())
                    TIMESTAMPNOW = timestampnow
                    xer = []
                    xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                elif ['dismem_20230104'].count(resource) == 1:
                    logger.info("Starting discovering resource - {0}".format(resource))
                    one_object_list = {}
                    discovered_resource = []
                    one_object_list["{#INDEX}"] = 1
                    discovered_resource.append(one_object_list)
                    logger.info("Succes get resource - {0}".format(resource))
                    converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                    timestampnow = int(time.time())
                    TIMESTAMPNOW = timestampnow
                    xer = []
                    xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                elif ['disinterface'].count(resource) == 1:
                    discovered_resource = []
                    stdout = self.def_execute_command_with_more("""sys\ndis int brief\n""")
                    # stdout = self.execute_command_with_more("""sys\ndis cu | include sysname\n""")
                    # stdout = self.execute_command_with_more("""sys\n dis cpu-usage\n""")
                    if len(stdout) > 0:
                        logger.info("Starting discovering resource - {0}".format(resource))

                        title_name = ['index', 'Interface', 'IP Address/Mask', 'Physical', 'Protocol']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        pattern = "(\w+\d+/*\d*/*\d*)\s*(\w*\.?\d*\.?\d*\.?\d*/?\d*)\s*(up|down)\s*(up|down)"
                        result = re.findall(pattern, stdout)
                        for i_01 in range(0, len(result)):
                            self.cus_excel_op.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                            for i_02 in range(0, len(result[i_01])):
                                self.cus_excel_op.def_set_cell_value(i_01 + 2, i_02 + 2, result[i_01][i_02].replace("42D", ""))

                    col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                    del col_02_list[0]
                    for len_col_02 in range(len(col_02_list)):
                        one_object_list = {}
                        one_object_list["{#INTERFACE}"] = col_02_list[len_col_02]
                        discovered_resource.append(one_object_list)
                    logger.info("Succes get resource - {0}".format(resource))
                    converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                    timestampnow = int(time.time())
                    TIMESTAMPNOW = timestampnow
                    xer = []
                    xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                else:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
            self.def_logout()
            # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)
        except Exception as pizdec:
            logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], pizdec))
            self.def_logout()
            sys.exit("{0}".format(inspect.currentframe().f_lineno))

    def def_get_status_resources(self, user, password, ip, port, host, list_resources):
        try:
            self.def_connect(ip, user, password, port)
            for resource in list_resources:
                if ['disname'].count(resource) == 1:
                    stdout = self.def_execute_command_with_more("""sys\ndis cu | include sysname\n""")
                    if len(stdout) > 0:
                        logger.info("Starting collecting status of resource - {0}".format(resource))

                        title_name = ['index', 'name']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        pattern = "sysname\s(\w.*)"
                        result = re.search(pattern, stdout).group(1).replace('\r', '')
                        self.cus_excel_op.def_set_cell_value(2, 1, 1)
                        self.cus_excel_op.def_set_cell_value(2, 2, result)

                        col_01_list = self.cus_excel_op.def_get_creat_col_value(1)
                        del col_01_list[0]
                        col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                        del col_02_list[0]
                        TIMESTAMPNOW = int(time.time())
                        state_resources = []
                        for len_col_01 in range(len(col_01_list)):
                            key_name = "{0}[{1}]".format(resource, col_01_list[len_col_01])
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_name, TIMESTAMPNOW, col_02_list[len_col_01]))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                elif ['discpu'].count(resource) == 1:
                    stdout = self.def_execute_command_with_more("""sys\n dis cpu-usage\n""")
                    if len(stdout) > 0:
                        logger.info("Starting collecting status of resource - {0}".format(resource))

                        title_name = ['index', 'name']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        pattern = "CPU\sUsage\s+:\s(\d.*%)\sMax"
                        result = re.search(pattern, stdout).group(1).replace('%', '')
                        self.cus_excel_op.def_set_cell_value(2, 1, 1)
                        self.cus_excel_op.def_set_cell_value(2, 2, result)

                        col_01_list = self.cus_excel_op.def_get_creat_col_value(1)
                        del col_01_list[0]
                        col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                        del col_02_list[0]
                        TIMESTAMPNOW = int(time.time())
                        state_resources = []
                        for len_col_01 in range(len(col_01_list)):
                            key_used = "{0}.[{1}]".format(resource, col_01_list[len_col_01])
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_used, TIMESTAMPNOW, col_02_list[len_col_01]))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                elif ['discpu_20230104'].count(resource) == 1:
                    stdout = self.def_execute_command_with_more("""sys\n dis cpu\n""")
                    if len(stdout) > 0:
                        logger.info("Starting collecting status of resource - {0}".format(resource))

                        title_name = ['index', 'name']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        pattern = re.compile(r'(\d*%)\sin')
                        result = re.search(pattern, stdout).group(1).replace('%', '')
                        self.cus_excel_op.def_set_cell_value(2, 1, 1)
                        self.cus_excel_op.def_set_cell_value(2, 2, result)

                        col_01_list = self.cus_excel_op.def_get_creat_col_value(1)
                        del col_01_list[0]
                        col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                        del col_02_list[0]
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        state_resources = []
                        for len_col_01 in range(len(col_01_list)):
                            key_used = "{0}.[{1}]".format(resource, col_01_list[len_col_01])
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_used, timestampnow, col_02_list[len_col_01]))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                elif ['dismem_20230104'].count(resource) == 1:
                    stdout = self.def_execute_command_with_more("""sys\n dis mem\n""")
                    if len(stdout) > 0:
                        logger.info("Starting collecting status of resource - {0}".format(resource))

                        title_name = ['index', 'name']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        pattern = re.compile(r'Rate:\s(\d*%)')
                        result = re.search(pattern, stdout).group(1).replace('%', '')
                        self.cus_excel_op.def_set_cell_value(2, 1, 1)
                        self.cus_excel_op.def_set_cell_value(2, 2, result)

                        col_01_list = self.cus_excel_op.def_get_creat_col_value(1)
                        del col_01_list[0]
                        col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                        del col_02_list[0]
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        state_resources = []
                        for len_col_01 in range(len(col_01_list)):
                            key_used = "{0}.[{1}]".format(resource, col_01_list[len_col_01])
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_used, timestampnow, col_02_list[len_col_01]))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                elif ['disinterface'].count(resource) == 1:
                    stdout = self.def_execute_command_with_more("""sys\ndis int brief\n""")
                    # stdout = self.execute_command_with_more("""sys\ndis cu | include sysname\n""")
                    # stdout = self.execute_command_with_more("""sys\n dis cpu-usage\n""")
                    if len(stdout) > 0:
                        logger.info("Starting collecting status of resource - {0}".format(resource))

                        title_name = ['index', 'Interface', 'IP Address/Mask', 'Physical', 'Protocol']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        pattern = "(\w+\d+/*\d*/*\d*)\s*(\w*\.?\d*\.?\d*\.?\d*/?\d*)\s*(up|down)\s*(up|down)"
                        result = re.findall(pattern, stdout)
                        for i_01 in range(0, len(result)):
                            self.cus_excel_op.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                            for i_02 in range(0, len(result[i_01])):
                                self.cus_excel_op.def_set_cell_value(i_01 + 2, i_02 + 2, result[i_01][i_02].replace("42D", ""))

                        col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                        del col_02_list[0]
                        col_04_list = self.cus_excel_op.def_get_creat_col_value(4)
                        del col_04_list[0]
                        col_05_list = self.cus_excel_op.def_get_creat_col_value(5)
                        del col_05_list[0]
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        state_resources = []
                        for len_col_02 in range(len(col_02_list)):
                            key_phy = "physical.{0}[{1}]".format(resource, col_02_list[len_col_02])
                            key_pro = "protocol.{0}[{1}]".format(resource, col_02_list[len_col_02])
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_phy, timestampnow, self.local_methord.def_convert_text_to_numeric(col_04_list[len_col_02])))
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_pro, timestampnow, self.local_methord.def_convert_text_to_numeric(col_05_list[len_col_02])))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                else:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
            self.def_logout()
            # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)
        except Exception as pizdec:
            logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], pizdec))
            self.def_logout()
            sys.exit("{0}".format(inspect.currentframe().f_lineno))


class CusSSHClient(object):
    def __init__(self, ):
        self.cus_sshClient = None
        self.cus_zabbixSender = CusZabbixSender()
        self.cus_excelOp = CusExcelOp()
        self.cus_excelOp.def_creat_excel()
        self.cus_localMethord = CusLocalMethod()
        self.lv_dic01 = {}

    def def_connect_fileCount(self, funcIp):
        try:
            # print(funcIp)
            ssh_client = paramiko.SSHClient()

            ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            timeout = 5
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout)
            sock.connect((funcIp, self.lv_dic01[funcIp]['port']))
            ssh_client = paramiko.Transport(sock)
            # print(self.dic01[funcIp]['file'])
            ssh_client.connect(username=self.lv_dic01[funcIp]['usr'], password=self.lv_dic01[funcIp]['pwd'])

            for lvInt01 in range(len(self.lv_dic01[funcIp]['file'])):
                print("""find {v00} {v01} start""".format(v01=self.lv_dic01[funcIp]['file'][lvInt01][1]['fileName'], v00=funcIp))
                ssh_channel = ssh_client.open_channel(kind='session')
                ssh_channel.exec_command("""find {v01} -type f -name "*.*" | head -n {v02} | wc -l""".format(v01=self.lv_dic01[funcIp]['file'][lvInt01][1]['fileName'], v02=self.lv_dic01[funcIp]['file'][lvInt01][0]['fileDepth']))
                stdout = ssh_channel.recv(1024).decode()
                self.lv_dic01[funcIp]['file'][lvInt01][2]['fileCount'] = stdout.replace('\n', '')
                stderr = ssh_channel.recv_stderr(1024).decode()
                # print(stdout, stderr)
                if len(stderr) != 0 and stdout.replace('\n', '') == 0:
                    pattern = re.compile(r'.*No such file.*')
                    result = re.findall(pattern, str(stderr))
                    if len(result) != 0:
                        self.lv_dic01[funcIp]['file'][lvInt01][2]['fileCount'] = 1000002  # [Errno 2] No such file
                    pattern = re.compile(r'.*Permission denied')
                    result = re.findall(pattern, str(stderr))
                    if len(result) != 0:
                        self.lv_dic01[funcIp]['file'][lvInt01][2]['fileCount'] = 1000005  # Permission denied
                ssh_channel.close()
                print("""find {v00} {v01} done""".format(v01=self.lv_dic01[funcIp]['file'][lvInt01][1]['fileName'], v00=funcIp))
            ssh_client.close()
            # pattern = re.compile(r'.*')
            # result = re.findall(pattern, stdout)
            # self.fileCount = stdout
        except Exception as e:
            print(e)
            for lvInt01 in range(len(self.lv_dic01[funcIp]['file'])):
                pattern = re.compile(r'Authentication failed.*')
                result = re.findall(pattern, str(e))
                if len(result) != 0:
                    self.lv_dic01[funcIp]['file'][lvInt01][2]['fileCount'] = 1000001  # Authentication failed.
                pattern = re.compile(r'.*Connection refused')
                result = re.findall(pattern, str(e))
                if len(result) != 0:
                    self.lv_dic01[funcIp]['file'][lvInt01][2]['fileCount'] = 1000003  # [Errno 111] Connection refused
                pattern = re.compile(r'timed out.*')
                result = re.findall(pattern, str(e))
                if len(result) != 0:
                    self.lv_dic01[funcIp]['file'][lvInt01][2]['fileCount'] = 1000004  # timed out


    def def_discovering_resources(self, host, list_resources):
        for resource in list_resources:
            if ['df'].count(resource) == 1:
                self.cus_excelOp.def_load_excel(file='/etc/zabbix/scripts/checkLinuxDisk.xlsx', index=1)

                column_1_list = self.cus_excelOp.def_get_col_value(1)
                del column_1_list[0]
                column_2_list = []
                column_3_list = []
                column_4_list = []
                column_6_list = []
                column_7_list = []

                for i in range(len(column_1_list)):
                    column_1_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 1))  # 主机名
                    column_2_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 2))  # IP地址
                    column_3_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 3))  # 协议
                    column_4_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 4))  # 用户名
                    column_6_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 6))  # 密码
                    column_7_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 7))  # 端口
                    connection = self.def_connect(self.cus_excelOp.def_get_cell_value(i + 2, 4),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 6),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 2),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 7))
                    stdin, stdout, stderr = connection.exec_command("""df | awk -v OFS=',' '{{print $6,$2,$5}}'""")
                    time.sleep(0.1)
                    if len(stdout.read().decode().strip()) > 0:
                        logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], stderr.read()))
                        connection.close()
                        sys.exit("{0}".format(inspect.currentframe().f_lineno))
                    else:
                        logger.info("Starting discovering resource - {0}".format(resource))

                        title_name = ['index', 'name', 'size', 'used']
                        self.cus_excelOp.def_create_sheet(resource)
                        [self.cus_excelOp.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        res_01 = stdout.read().decode().strip().split('\n')
                        del res_01[0]
                        for i_01 in range(len(res_01)):
                            res_02 = res_01[i_01].split(',')
                            self.cus_excelOp.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                            for i_02 in range(len(res_02)):
                                self.cus_excelOp.def_set_cell_value(i_01 + 2, i_02 + 2, res_02[i_02])

                        discovered_resource = []
                        col_01_list = self.cus_excelOp.def_get_creat_col_value(2)
                        del col_01_list[0]
                        for col_01 in col_01_list:
                            one_object_list = {}
                            one_object_list["{#MOUNTED}"] = col_01.replace("/", "_")
                            discovered_resource.append(one_object_list)
                        logger.info("Succes get resource - {0}".format(resource))
                        converted_resource = self.cus_zabbixSender.def_convert_to_zabbix_json(discovered_resource)
                        xer = []
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                        # print(xer)
                        self.cus_zabbixSender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
            elif ['netstat'].count(resource) == 1:
                self.cus_excelOp.def_load_excel(file='/etc/zabbix/scripts/checkLinuxPort.xlsx', index=1)

                column_1_list = self.cus_excelOp.def_get_col_value(1)
                del column_1_list[0]
                column_2_list = []
                column_3_list = []
                column_4_list = []
                column_6_list = []
                column_7_list = []

                for i in range(len(column_1_list)):
                    column_1_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 1))  # 主机名
                    column_2_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 2))  # IP地址
                    column_3_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 3))  # 协议
                    column_4_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 4))  # 用户名
                    column_6_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 6))  # 密码
                    column_7_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 7))  # 端口
                    connection = self.def_connect(self.cus_excelOp.def_get_cell_value(i + 2, 4),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 6),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 2),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 7))
                    stdin, stdout, stderr = connection.exec_command("""netstat -ntlp | awk -v OFS=',' '{{print $7,$4,$6}}'""")
                    time.sleep(0.1)
                    if len(stderr.read().decode().strip()) > 0:
                        logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], stderr.read()))
                        connection.close()
                        sys.exit("{0}".format(inspect.currentframe().f_lineno))
                    else:
                        logger.info("Starting discovering resource - {0}".format(resource))

                        title_name = ['index', 'name', 'port', 'state']
                        self.cus_excelOp.def_create_sheet(resource)
                        [self.cus_excelOp.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        pattern = re.compile(r'(\d*/\w*).*,.*:(\d*),(\w*)n*')
                        result = re.findall(pattern, stdout.read().decode().strip())
                        # 去重
                        unit_column_1_list = list(set(result))
                        # 使用index保持不乱序
                        unit_column_1_list.sort(key=result.index)
                        for i_01 in range(0, len(unit_column_1_list)):
                            self.cus_excelOp.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                            for i_02 in range(0, len(unit_column_1_list[i_01])):
                                self.cus_excelOp.def_set_cell_value(i_01 + 2, i_02 + 2, unit_column_1_list[i_01][i_02].replace("/", "-"))
                        discovered_resource = []
                        col_02_list = self.cus_excelOp.def_get_creat_col_value(2)
                        del col_02_list[0]
                        col_03_list = self.cus_excelOp.def_get_creat_col_value(3)
                        del col_03_list[0]
                        for len_col_01 in range(len(col_02_list)):
                            one_object_list = {}
                            one_object_list["{#NAME}"] = col_02_list[len_col_01].replace("/", "-") + '_' + col_03_list[len_col_01]
                            discovered_resource.append(one_object_list)
                        logger.info("Succes get resource - {0}".format(resource))
                        converted_resource = self.cus_zabbixSender.def_convert_to_zabbix_json(discovered_resource)
                        xer = []
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                        # print(xer)
                        self.cus_zabbixSender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
            elif ['fileCount'].count(resource) == 1:
                self.cus_excelOp.def_load_excel(file='/etc/zabbix/scripts/checkFileCount.xlsx', index=1)

                column_1_list = self.cus_excelOp.def_get_col_value(1)
                del column_1_list[0]
                column_2_list = []
                column_3_list = []
                column_4_list = []
                column_6_list = []
                column_7_list = []
                column_8_list = []
                column_9_list = []
                column_10_list = []

                for i in range(len(column_1_list)):
                    column_1_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 1))  # 通道ID
                    column_2_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 2))  # 通道从哪
                    column_3_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 3))  # 通道到哪
                    column_4_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 4))  # 属地通道IP地址
                    column_6_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 6))  # 用户名
                    column_7_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 7))  # 密码
                    column_8_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 8))  # 端口
                    column_9_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 9))  # 告警文件数
                    column_10_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 10))  # 监控目录

                discovered_resource = []
                for col_01 in range(len(column_4_list)):
                    one_object_list = {}
                    one_object_list["{#IPLIST}"] = str(column_4_list[col_01].replace(".", "_")) + "_" + \
                                                   column_10_list[col_01].replace("/", "_")
                    one_object_list["{#ADDRESSLIST}"] = str(column_1_list[col_01]) + "_" + \
                                                        column_2_list[col_01] + "_" + \
                                                        column_3_list[col_01] + "_" + \
                                                        column_10_list[col_01].replace("/", "_")
                    discovered_resource.append(one_object_list)
                converted_resource = self.cus_zabbixSender.def_convert_to_zabbix_json(discovered_resource)
                xer = []
                timestampnow = int(time.time())
                TIMESTAMPNOW = timestampnow
                xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                # print(xer)
                self.cus_zabbixSender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)

    def def_get_status_resources(self, host, list_resources):
        for resource in list_resources:
            if ['df'].count(resource) == 1:
                self.cus_excelOp.def_load_excel(file='/etc/zabbix/scripts/checkLinuxDisk.xlsx', index=1)

                column_1_list = self.cus_excelOp.def_get_col_value(1)
                del column_1_list[0]
                column_2_list = []
                column_3_list = []
                column_4_list = []
                column_6_list = []
                column_7_list = []

                for i in range(len(column_1_list)):
                    column_1_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 1))  # 主机名
                    column_2_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 2))  # IP地址
                    column_3_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 3))  # 协议
                    column_4_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 4))  # 用户名
                    column_6_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 6))  # 密码
                    column_7_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 7))  # 端口
                    connection = self.def_connect(self.cus_excelOp.def_get_cell_value(i + 2, 4),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 6),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 2),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 7))
                    stdin, stdout, stderr = connection.exec_command("""df | awk -v OFS=',' '{{print $6,$2,$5}}'""")
                    time.sleep(0.1)
                    if len(stderr.read()) > 0:
                        logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], stderr.read()))
                        connection.close()
                        sys.exit("{0}".format(inspect.currentframe().f_lineno))
                    else:
                        logger.info("Starting collecting status of resource - {0}".format(resource))

                        title_name = ['index', 'name', 'size', 'used']
                        self.cus_excelOp.def_create_sheet(resource)
                        [self.cus_excelOp.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        res_01 = stdout.read().decode().strip().split('\n')
                        del res_01[0]
                        for i_01 in range(len(res_01)):
                            res_02 = res_01[i_01].split(',')
                            self.cus_excelOp.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                            for i_02 in range(len(res_02)):
                                self.cus_excelOp.def_set_cell_value(i_01 + 2, i_02 + 2, res_02[i_02])

                        col_02_list = self.cus_excelOp.def_get_creat_col_value(2)
                        del col_02_list[0]
                        col_03_list = self.cus_excelOp.def_get_creat_col_value(3)
                        del col_03_list[0]
                        col_04_list = self.cus_excelOp.def_get_creat_col_value(4)
                        del col_04_list[0]
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        state_resources = []
                        for len_col_01 in range(len(col_02_list)):
                            key_size = "size.{0}[{1}]".format(resource, col_02_list[len_col_01].replace("/", "_"))
                            key_used = "used.{0}[{1}]".format(resource, col_02_list[len_col_01].replace("/", "_"))
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + key_size + "\"", timestampnow, col_03_list[len_col_01]))
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + key_used + "\"", timestampnow, col_04_list[len_col_01].replace("%", "")))
                        self.cus_zabbixSender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                    connection.close()
            elif ['netstat'].count(resource) == 1:
                self.cus_excelOp.def_load_excel(file='/etc/zabbix/scripts/checkLinuxPort.xlsx', index=1)

                column_1_list = self.cus_excelOp.def_get_col_value(1)
                del column_1_list[0]
                column_2_list = []
                column_3_list = []
                column_4_list = []
                column_6_list = []
                column_7_list = []

                for i in range(len(column_1_list)):
                    column_1_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 1))  # 主机名
                    column_2_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 2))  # IP地址
                    column_3_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 3))  # 协议
                    column_4_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 4))  # 用户名
                    column_6_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 6))  # 密码
                    column_7_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 7))  # 端口
                    connection = self.def_connect(self.cus_excelOp.def_get_cell_value(i + 2, 4),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 6),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 2),
                                                  self.cus_excelOp.def_get_cell_value(i + 2, 7))
                    stdin, stdout, stderr = connection.exec_command("""netstat -ntlp | awk -v OFS=',' '{{print $7,$4,$6}}'""")
                    time.sleep(0.1)
                    if len(stderr.read()) > 0:
                        logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], stderr.read()))
                        connection.close()
                        sys.exit("{0}".format(inspect.currentframe().f_lineno))
                    else:
                        logger.info("Starting collecting status of resource - {0}".format(resource))

                        title_name = ['index', 'name', 'port', 'state']
                        self.cus_excelOp.def_create_sheet(resource)
                        [self.cus_excelOp.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        pattern = re.compile(r'(\d*/\w*).*,.*:(\d*),(\w*)n*')
                        result = re.findall(pattern, stdout.read().decode().strip())
                        # 去重
                        unit_column_1_list = list(set(result))
                        # 使用index保持不乱序
                        unit_column_1_list.sort(key=result.index)
                        for i_01 in range(0, len(unit_column_1_list)):
                            self.cus_excelOp.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                            for i_02 in range(0, len(unit_column_1_list[i_01])):
                                self.cus_excelOp.def_set_cell_value(i_01 + 2, i_02 + 2, unit_column_1_list[i_01][i_02].replace("/", "-"))
                                if i_02 == 2:
                                    stdin, stdout, stderr = connection.exec_command("""tcping {0} {1}""".format(self.cus_excelOp.def_get_cell_value(i + 2, 2), unit_column_1_list[i_01][i_02 - 1]))
                                    time.sleep(0.1)
                                    pattern = re.compile(r'.*(open|closed).*')
                                    result = re.search(pattern, stdout.read().decode().strip())
                                    if self.cus_localMethord.def_convert_text_to_numeric(result.group(1)) == 0 and \
                                            self.cus_localMethord.def_convert_text_to_numeric(unit_column_1_list[i_01][i_02]) == 0:
                                        self.cus_excelOp.def_set_cell_value(i_01 + 2, i_02 + 2, 0)
                                    else:
                                        self.cus_excelOp.def_set_cell_value(i_01 + 2, i_02 + 2, 1)
                        col_02_list = self.cus_excelOp.def_get_creat_col_value(2)
                        del col_02_list[0]
                        col_03_list = self.cus_excelOp.def_get_creat_col_value(3)
                        del col_03_list[0]
                        col_04_list = self.cus_excelOp.def_get_creat_col_value(4)
                        del col_04_list[0]
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        state_resources = []
                        for len_col_01 in range(len(col_02_list)):
                            key_state = "state.{0}[{1}]".format(resource, col_02_list[len_col_01] + '_' + col_03_list[len_col_01])
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + key_state + "\"", timestampnow, col_04_list[len_col_01]))
                        self.cus_zabbixSender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                    connection.close()
            elif ['fileCount'].count(resource) == 1:
                self.cus_excelOp.def_load_excel(file='/etc/zabbix/scripts/checkFileCount.xlsx', index=1)

                column_1_list = self.cus_excelOp.def_get_col_value(1)
                del column_1_list[0]
                column_2_list = []
                column_3_list = []
                column_4_list = []
                column_6_list = []
                column_7_list = []
                column_8_list = []
                column_9_list = []
                column_10_list = []
                self.lv_dic01 = {}
                lvList03 = []
                for i in range(len(column_1_list)):
                    column_1_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 1))  # 通道ID
                    column_2_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 2))  # 通道从哪
                    column_3_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 3))  # 通道到哪
                    column_4_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 4))  # 属地通道IP地址
                    column_6_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 6))  # 用户名
                    column_7_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 7))  # 密码
                    column_8_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 8))  # 端口
                    column_9_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 9))  # 告警文件数
                    column_10_list.append(self.cus_excelOp.def_get_cell_value(i + 2, 10))  # 监控目录
                self.lv_dic01 = dict.fromkeys(column_4_list, [])
                for lvInt01 in range(len(column_4_list)):
                    # print(str(lvInt01+1) + 'AAA',end='')
                    # print(self.dic01.get(column_4_list[lvInt01]))
                    if self.lv_dic01.get(column_4_list[lvInt01], []):
                        lvList03 = self.lv_dic01[column_4_list[lvInt01]]['file']
                        lvList03.append([{'fileDepth': column_9_list[lvInt01]},
                                         {'fileName': column_10_list[lvInt01]},
                                         {'fileCount': 0}])
                        # print(lvList03)
                        self.lv_dic01.update({column_4_list[lvInt01]: {'file': lvList03, 'usr': column_6_list[lvInt01],
                                                                    'pwd': column_7_list[lvInt01], 'port': column_8_list[lvInt01]}})
                    else:
                        lvList02 = [[{'fileDepth': column_9_list[lvInt01]},
                                    {'fileName': column_10_list[lvInt01]},
                                    {'fileCount': 0}]]
                        # print(lvList02)
                        self.lv_dic01.update({column_4_list[lvInt01]: {'file': lvList02, 'usr': column_6_list[lvInt01],
                                                                    'pwd': column_7_list[lvInt01], 'port': column_8_list[lvInt01]}})

                    # print(self.dic01)
                # exit(1)
                # [{key:[[6,7,8],[6,7,8]]}, {key:[[6,7,8],[6,7,8]]]
                executor = ThreadPoolExecutor(GV_CPU_COUNT)
                [lv_dic02 for lv_dic02 in executor.map(self.def_connect_fileCount, list(self.lv_dic01.keys()))]
                timestampnow = int(time.time())
                TIMESTAMPNOW = timestampnow
                state_resources = []
                for lv_key01, lv_value01 in self.lv_dic01.items():
                    for lv_value02 in range(len(lv_value01['file'])):
                        keyQueueDepth = "queueDepth.[{0}_{1}]".format(lv_key01.replace(".", "_"), lv_value01['file'][lv_value02][1]['fileName'].replace("/", "_"))
                        state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + keyQueueDepth + "\"", timestampnow, lv_value01['file'][lv_value02][2]['fileCount']))
                # print(state_resources)
                # exit(1)
                self.cus_zabbixSender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)

            else:
                logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))


class CusSftpClient(object):
    def __init__(self, ):
        self.sftpClient = None
        self.sftp = None
        self.sftpCount = 0
        self.zabbix_sender = CusZabbixSender()
        self.cus_excel_op = CusExcelOp()
        self.cus_excel_op.def_creat_excel()
        self.local_methord = CusLocalMethod()

    def def_connect(self, ip, user, password, port, queueDepth, remote):
        try:
            timeout = 5
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout)
            sock.connect((ip, port))
            self.sftpClient = paramiko.Transport(sock)
            self.sftpClient.connect(username=user, password=password)
            self.sftp = paramiko.SFTPClient.from_transport(self.sftpClient)
            self.sftpCount = 0
            self.def_get(int(queueDepth), str(remote))
            time.sleep(0.1)
            self.sftpClient.close()
            return self.sftpCount
        except Exception as e:
            print(e)
            pattern = re.compile(r'Authentication failed.*')
            result = re.findall(pattern, str(e))
            if len(result) != 0:
                self.sftpCount = 100001  # Authentication failed.
                return self.sftpCount
            pattern = re.compile(r'.*No such file')
            result = re.findall(pattern, str(e))
            if len(result) != 0:
                self.sftpCount = 100002  # [Errno 2] No such file
                return self.sftpCount
            pattern = re.compile(r'.*Connection refused')
            result = re.findall(pattern, str(e))
            if len(result) != 0:
                self.sftpCount = 100003  # [Errno 111] Connection refused
                return self.sftpCount
            pattern = re.compile(r'timed out.*')
            result = re.findall(pattern, str(e))
            if len(result) != 0:
                self.sftpCount = 100004  # timed out
                return self.sftpCount

    def def_get(self, queueDepth, remote):
        # 检查远程文件是否存在
        result = self.sftp.stat(remote)
        if isdir(result.st_mode):
            for file in self.sftp.listdir(remote):
                sub_remote = os.path.join(remote, file)
                sub_remote = sub_remote.replace('\\', '/')
                self.def_get(queueDepth, sub_remote)
        else:
            # 拷贝文件
            pattern = re.compile(r'^/\..*')
            result = re.findall(pattern, str(remote))
            if len(result) == 0:
                self.sftpCount = self.sftpCount + 1
                if self.sftpCount > queueDepth:
                    self.sftpCount = 999999
                    return self.sftpCount

    def def_discovering_resources(self, host, list_resources):
        for resource in list_resources:
            if ['queueDepth'].count(resource) == 1:
                self.cus_excel_op.def_load_excel(file='/etc/zabbix/scripts/checkFileCount.xlsx', index=1)

                column_1_list = self.cus_excel_op.def_get_col_value(1)
                del column_1_list[0]
                column_2_list = []
                column_3_list = []
                column_4_list = []
                column_6_list = []
                column_7_list = []
                column_8_list = []
                column_9_list = []
                column_10_list = []

                for i in range(len(column_1_list)):
                    column_1_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 1))  # 通道ID
                    column_2_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 2))  # 通道从哪
                    column_3_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 3))  # 通道到哪
                    column_4_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 4))  # 属地通道IP地址
                    column_6_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 6))  # 用户名
                    column_7_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 7))  # 密码
                    column_8_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 8))  # 端口
                    column_9_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 9))  # 告警文件数
                    column_10_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 10))  # 监控目录

                discovered_resource = []
                for col_01 in range(len(column_4_list)):
                    one_object_list = {}
                    one_object_list["{#IPLIST}"] = str(column_4_list[col_01].replace(".", "_")) + "_" + \
                                                   column_10_list[col_01].replace("/", "_")
                    one_object_list["{#ADDRESSLIST}"] = str(column_1_list[col_01]) + "_" + \
                                                        column_2_list[col_01] + "_" + \
                                                        column_3_list[col_01] + "_" + \
                                                        column_10_list[col_01].replace("/", "_")
                    discovered_resource.append(one_object_list)
                converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                xer = []
                timestampnow = int(time.time())
                TIMESTAMPNOW = timestampnow
                xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                # print(xer)
                self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)

    def def_get_status_resources(self, host, list_resources):
        for resource in list_resources:
            if ['queueDepth'].count(resource) == 1:
                self.cus_excel_op.def_load_excel(file='/etc/zabbix/scripts/checkFileCount.xlsx', index=1)

                column_1_list = self.cus_excel_op.def_get_col_value(1)
                del column_1_list[0]
                column_2_list = []
                column_3_list = []
                column_4_list = []
                column_6_list = []
                column_7_list = []
                column_8_list = []
                column_9_list = []
                column_10_list = []

                for i in range(len(column_1_list)):
                    column_1_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 1))  # 通道ID
                    column_2_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 2))  # 通道从哪
                    column_3_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 3))  # 通道到哪
                    column_4_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 4))  # 属地通道IP地址
                    column_6_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 6))  # 用户名
                    column_7_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 7))  # 密码
                    column_8_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 8))  # 端口
                    column_9_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 9))  # 告警文件数
                    column_10_list.append(self.cus_excel_op.def_get_cell_value(i + 2, 10))  # 监控目录

                executor = ThreadPoolExecutor(GV_CPU_COUNT)
                lv_listGetAllSftCount = []
                lv_result = None
                # def_connect(self, ip, user, password, port, queueDepth, remote):
                for lv_result in executor.map(self.def_connect, column_4_list, column_6_list, column_7_list, column_8_list, column_9_list, column_10_list):
                    lv_listGetAllSftCount.append(lv_result)
                timestampnow = int(time.time())
                TIMESTAMPNOW = timestampnow
                state_resources = []
                for len_col_01 in range(len(column_4_list)):
                    keyQueueDepth = "queueDepth.[{0}_{1}]".format(column_4_list[len_col_01].replace(".", "_"), column_10_list[len_col_01].replace("/", "_"))
                    state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + keyQueueDepth + "\"", timestampnow, lv_listGetAllSftCount[len_col_01]))
                self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)


class CusLinuxClient(object):
    def __init__(self, ):
        self.zabbix_sender = CusZabbixSender()
        self.cus_excel_op = CusExcelOp()
        self.cus_excel_op.def_creat_excel()
        self.local_methord = CusLocalMethod()

    def def_cmd(self, command):
        result = subprocess.Popen(command, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, encoding="utf-8")
        return result

    def def_discovering_resources(self, host, list_resources):
        for resource in list_resources:
            if ['linux_df'].count(resource) == 1:
                cmdline = self.def_cmd("""df | awk -v OFS=',' '{{print $6,$2,$5}}'""")
                time.sleep(0.1)
                if len(cmdline.stderr.read().strip()) > 0:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], cmdline.stderr.read().strip()))
                    sys.exit("{0}".format(inspect.currentframe().f_lineno))
                else:
                    logger.info("Starting discovering resource - {0}".format(resource))
                    title_name = ['index', 'name', 'size', 'used']
                    self.cus_excel_op.def_create_sheet(resource)
                    [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                    res_01 = cmdline.stdout.read().strip().split('\n')
                    del res_01[0]
                    for i_01 in range(len(res_01)):
                        res_02 = res_01[i_01].split(',')
                        self.cus_excel_op.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                        for i_02 in range(len(res_02)):
                            self.cus_excel_op.def_set_cell_value(i_01 + 2, i_02 + 2, res_02[i_02])

                    discovered_resource = []
                    col_01_list = self.cus_excel_op.def_get_creat_col_value(2)
                    del col_01_list[0]
                    for col_01 in col_01_list:
                        one_object_list = {}
                        one_object_list["{#MOUNTED}"] = col_01.replace("/", "_")
                        discovered_resource.append(one_object_list)
                    logger.info("Succes get resource - {0}".format(resource))
                    converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                    xer = []
                    TIMESTAMPNOW = int(time.time())
                    xer.append("%s %s %s %s" % (host, resource, TIMESTAMPNOW, converted_resource))
                    # print(xer)
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                    print("df done")
            elif ['linux_netstat'].count(resource) == 1:
                cmdline = self.def_cmd("""netstat -ntlp | awk -v OFS=',' '{{print $7,$4,$6}}'""")
                time.sleep(0.1)
                # print(cmdline.stdout.read().strip())
                if len(cmdline.stderr.read().strip()) > 0:
                    # print(cmdline.stdout.read().strip())
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], cmdline.stderr.read().strip()))
                    sys.exit("{0}".format(inspect.currentframe().f_lineno))
                else:
                    logger.info("Starting discovering resource - {0}".format(resource))
                    title_name = ['index', 'name', 'ip', 'port', 'state']
                    self.cus_excel_op.def_create_sheet(resource)
                    [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                    pattern = re.compile(r'\d*/(\w*).*,(.*):(\d*),(\w*)n*')
                    result = re.findall(pattern, cmdline.stdout.read().strip())
                    # 去重
                    unit_column_1_list = list(set(result))
                    # 使用index保持不乱序
                    unit_column_1_list.sort(key=result.index)
                    for i_01 in range(0, len(unit_column_1_list)):
                        self.cus_excel_op.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                        for i_02 in range(0, len(unit_column_1_list[i_01])):
                            self.cus_excel_op.def_set_cell_value(i_01 + 2, i_02 + 2, unit_column_1_list[i_01][i_02].replace("/", "-"))
                    discovered_resource = []
                    col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                    del col_02_list[0]
                    col_03_list = self.cus_excel_op.def_get_creat_col_value(4)
                    del col_03_list[0]
                    col_04_list = self.cus_excel_op.def_get_creat_col_value(3)
                    del col_04_list[0]
                    for len_col_01 in range(len(col_02_list)):
                        one_object_list = {}
                        one_object_list["{#NAME}"] = col_02_list[len_col_01].replace("/", "-") + '_' + col_03_list[len_col_01]
                        discovered_resource.append(one_object_list)
                    logger.info("Succes get resource - {0}".format(resource))
                    converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                    xer = []
                    TIMESTAMPNOW = int(time.time())
                    xer.append("%s %s %s %s" % (host, resource, TIMESTAMPNOW, converted_resource))
                    # print(xer)
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                    print("netstat done")
            elif ['linux_checkping'].count(resource) == 1:
                logger.info("Starting discovering resource - {0}".format(resource))
                self.cus_excel_op.def_load_excel(file='/etc/zabbix/scripts/checkping.xlsx', index=1)
                col_02_list = self.cus_excel_op.def_get_col_value(1)
                del col_02_list[0]
                col_01_list = self.cus_excel_op.def_get_col_value(2)
                del col_01_list[0]
                col_03_list = self.cus_excel_op.def_get_col_value(3)
                del col_03_list[0]
                discovered_resource = []
                for len_col_01 in range(len(col_02_list)):
                    one_object_list = {}
                    one_object_list["{#NAME}"] = col_01_list[len_col_01]
                    one_object_list["{#INDEX}"] = col_02_list[len_col_01]
                    discovered_resource.append(one_object_list)
                logger.info("Succes get resource - {0}".format(resource))
                converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                xer = []
                TIMESTAMPNOW = int(time.time())
                xer.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + resource + "\"", TIMESTAMPNOW, converted_resource))
                # print(xer)
                self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                print("checkping done")
            else:
                logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
                print(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
        # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)

    def def_get_status_resources(self, host, list_resources):
        # try:
        for resource in list_resources:
            if ['linux_df'].count(resource) == 1:
                cmdline = self.def_cmd("""df | awk -v OFS=',' '{{print $6,$2,$5}}'""")
                time.sleep(0.1)
                if len(cmdline.stderr.read().strip()) > 0:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], cmdline.stderr.read().strip()))
                    sys.exit("{0}".format(inspect.currentframe().f_lineno))
                else:
                    logger.info("Starting collecting status of resource - {0}".format(resource))

                    title_name = ['index', 'name', 'size', 'used']
                    self.cus_excel_op.def_create_sheet(resource)
                    [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                    res_01 = cmdline.stdout.read().strip().split('\n')
                    del res_01[0]
                    for i_01 in range(len(res_01)):
                        res_02 = res_01[i_01].split(',')
                        self.cus_excel_op.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                        for i_02 in range(len(res_02)):
                            self.cus_excel_op.def_set_cell_value(i_01 + 2, i_02 + 2, res_02[i_02])

                    col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                    del col_02_list[0]
                    col_03_list = self.cus_excel_op.def_get_creat_col_value(3)
                    del col_03_list[0]
                    col_04_list = self.cus_excel_op.def_get_creat_col_value(4)
                    del col_04_list[0]
                    TIMESTAMPNOW = int(time.time())
                    state_resources = []
                    for len_col_01 in range(len(col_02_list)):
                        key_size = "size.{0}[{1}]".format(resource, col_02_list[len_col_01].replace("/", "_"))
                        key_used = "used.{0}[{1}]".format(resource, col_02_list[len_col_01].replace("/", "_"))
                        state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + key_size + "\"", TIMESTAMPNOW, col_03_list[len_col_01]))
                        state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + key_used + "\"", TIMESTAMPNOW, col_04_list[len_col_01].replace("%", "")))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                    print("df done")
            elif ['linux_netstat'].count(resource) == 1:
                cmdline = self.def_cmd("""netstat -ntlp | awk -v OFS=',' '{{print $7,$4,$6}}'""")
                time.sleep(0.1)
                if len(cmdline.stderr.read().strip()) > 0:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], cmdline.stderr.read().strip()))
                    sys.exit("{0}".format(inspect.currentframe().f_lineno))
                else:
                    logger.info("Starting collecting status of resource - {0}".format(resource))

                    title_name = ['index', 'name', 'ip', 'port', 'state']
                    self.cus_excel_op.def_create_sheet(resource)
                    [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                    cmd_stdout = cmdline.stdout.read().strip()
                    pattern = re.compile(r'\d*/(\w*).*,(.*):(\d*),(\w*)n*')
                    result = re.findall(pattern, cmd_stdout)
                    # 去重
                    unit_column_1_list = list(set(result))
                    # 使用index保持不乱序
                    unit_column_1_list.sort(key=result.index)
                    for i_01 in range(0, len(unit_column_1_list)):
                        self.cus_excel_op.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                        for i_02 in range(0, len(unit_column_1_list[i_01])):
                            self.cus_excel_op.def_set_cell_value(i_01 + 2, i_02 + 2, unit_column_1_list[i_01][i_02].replace("/", "-"))
                            if i_02 == 3:
                                cmdline = self.def_cmd("""tcping {0} {1}""".format(self.local_methord.def_convert_text_to_text(unit_column_1_list[i_01][i_02 - 2]), unit_column_1_list[i_01][i_02 - 1]))
                                time.sleep(0.1)
                                pattern = re.compile(r'.*(open|closed).*')
                                cmd_stdout = cmdline.stdout.read().strip()
                                result = re.search(pattern, cmd_stdout)
                                if self.local_methord.def_convert_text_to_numeric(result.group(1)) == 0 and \
                                        self.local_methord.def_convert_text_to_numeric(unit_column_1_list[i_01][i_02]) == 0:
                                    self.cus_excel_op.def_set_cell_value(i_01 + 2, i_02 + 2, 0)
                                else:
                                    self.cus_excel_op.def_set_cell_value(i_01 + 2, i_02 + 2, 1)
                    col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                    del col_02_list[0]
                    col_04_list = self.cus_excel_op.def_get_creat_col_value(4)
                    del col_04_list[0]
                    col_05_list = self.cus_excel_op.def_get_creat_col_value(5)
                    del col_05_list[0]
                    TIMESTAMPNOW = int(time.time())
                    state_resources = []
                    for len_col_01 in range(len(col_02_list)):
                        key_state = "state.{0}[{1}]".format(resource, col_02_list[len_col_01] + '_' + col_04_list[len_col_01])
                        state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + key_state + "\"", TIMESTAMPNOW, col_05_list[len_col_01]))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                    print("netstat done")
            elif ['linux_checkping'].count(resource) == 1:
                logger.info("Starting collecting status of resource - {0}".format(resource))
                self.cus_excel_op.def_load_excel(file='/etc/zabbix/scripts/checkping.xlsx', index=1)
                col_02_list = self.cus_excel_op.def_get_col_value(1)
                del col_02_list[0]
                col_03_list = self.cus_excel_op.def_get_col_value(3)
                del col_03_list[0]
                TIMESTAMPNOW = int(time.time())
                state_resources = []
                v_result = None
                for len_col_01 in range(len(col_02_list)):
                    key_state = "{0}.status[{1}]".format(resource, col_02_list[len_col_01])
                    for i_03 in range(3):
                        response = ping(col_03_list[len_col_01], timeout=1, size=56)
                        if response is not None:
                            v_result = int(response * 1000)
                            break
                        else:
                            v_result = 999
                            continue
                    state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + key_state + "\"", TIMESTAMPNOW, v_result))
                v_result = None
                self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                print("checkping done")
            elif ['linux_test'].count(resource) == 1:
                logger.info("Starting collecting status of resource - {0}".format(resource))
                col_02_list = [{'EventCount': 1, 'Domain': None, "User": "slt1-admin@takenaka.cn", "IPAddress": "192.168.100.1"},
                               {'EventCount': 2, 'Domain': None, "User": "slt2-admin@takenaka.cn", "IPAddress": "192.168.100.2"}]
                TIMESTAMPNOW = int(time.time())
                state_resources = []
                key_Event = "{0}.[1]".format(resource)
                for len_col_01 in range(len(col_02_list)):
                    state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "\"" + key_Event + "\"", TIMESTAMPNOW, col_02_list[len_col_01]))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                print("test done")

            else:
                logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
                print(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
        # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)


class CusDirectClient(object):
    def __init__(self, ):
        self.zabbix_sender = CusZabbixSender()
        self.local_methord = CusLocalMethod()

    def def_get_status_resources(self, host, list_resources, value):
        # try:
        for resource in list_resources:
            if ['ssh_login_log'].count(resource) == 1:
                logger.info("Starting collecting status of resource - {0}".format(resource))
                timestampnow = int(time.time())
                TIMESTAMPNOW = timestampnow
                state_resources = []
                state_resources.append("%s %s %s %s" % ("\"" + host + "\"", "ssh_login_log", timestampnow, "\"" + value + "\""))
                self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                print(0)
            else:
                logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
                print(1)
        # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)
    # except Exception as pizdec:
    #     logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], pizdec))
    #     sys.exit("{0}".format(inspect.currentframe().f_lineno))


class CusWebClinet(object):
    def __init__(self, ):
        self.zabbix_sender = CusZabbixSender()
        self.cus_excel_op = CusExcelOp()
        self.cus_excel_op.def_creat_excel()
        self.local_methord = CusLocalMethod()

        self.header = {"Content-Type": "application/json"}
        self.session = requests.Session()

    def def_discovering_resources_webcode_schoolid(self, host, list_resources):
        try:
            for resource in list_resources:
                if ['webcode_schoolid'].count(resource) == 1:
                    logger.info("Starting discovering resource - {0}".format(resource))
                    one_object_list = {}
                    discovered_resource = []
                    one_object_list["{#INDEX}"] = 1
                    discovered_resource.append(one_object_list)
                    logger.info("Succes get resource - {0}".format(resource))
                    converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                    timestampnow = int(time.time())
                    TIMESTAMPNOW = timestampnow
                    xer = []
                    xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                    self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                else:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
            # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)
            print(0)
        except Exception as oops:
            logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], oops))
            sys.exit("{0}".format(inspect.currentframe().f_lineno))

    def def_discovering_resources_web_service(self, host, list_resources, url_http, serviceid_str, marketid_int, stockcode_list):
        try:
            for resource in list_resources:
                if ['serviceid'].count(resource) == 1:
                    try:
                        logger.info("Starting discovering resource - {0}".format(resource))
                        self.session.mount(url_http, requests.adapters.HTTPAdapter(max_retries=3))

                        json_data_1 = {
                            "serviceid": serviceid_str,
                            "body": {
                                "marketid": marketid_int,
                                "stockcode": stockcode_list,
                            }
                        }
                        request = self.session.post(url=url_http, headers=self.header, json=json_data_1)
                        response = request.json()
                        if response.get('result', '') != '':
                            self.authID = response['result']
                        elif response.get('error', '') != '':
                            sys.exit("{0}".format(inspect.currentframe().f_lineno))
                        time.sleep(0.1)

                        title_name = ['index', 'name', 'now', 'volume', 'amount']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]

                        for i_01 in range(len(response['data'])):
                            self.cus_excel_op.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                            for i_02 in range(len(title_name)):
                                if i_02 == 1:
                                    self.cus_excel_op.def_set_cell_value(i_01 + 2, 2, response['data'][i_01]['code'] + '_' + response['data'][i_01]['name'])
                        discovered_resource = []
                        col_01_list = self.cus_excel_op.def_get_creat_col_value(2)
                        del col_01_list[0]
                        for col_01 in col_01_list:
                            one_object_list = {}
                            one_object_list["{#NAME}"] = col_01
                            discovered_resource.append(one_object_list)
                        logger.info("Succes get resource - {0}".format(resource))
                        converted_resource = self.zabbix_sender.def_convert_to_zabbix_json(discovered_resource)
                        xer = []
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        xer.append("%s %s %s %s" % (host, resource, timestampnow, converted_resource))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, xer, host)
                    except Exception as ee:
                        logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], ee))
                        sys.exit("{0}".format(inspect.currentframe().f_lineno))
                else:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
            # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)
            print(0)
        except Exception as oops:
            logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], oops))
            sys.exit("{0}".format(inspect.currentframe().f_lineno))

    def def_get_status_resources_webcode_schoolid(self, host, list_resources, url_http, url_userid, url_pwd, url_schoolid):
        try:
            for resource in list_resources:
                if ['webcode_schoolid'].count(resource) == 1:
                    logger.info("Starting discovering resource - {0}".format(resource))
                    self.session.mount(url_http, requests.adapters.HTTPAdapter(max_retries=3))
                    try:
                        json_data_1 = {
                            'schoolid': url_schoolid,
                            'userid': url_userid,
                            'password': url_pwd,
                        }
                        json_data_2 = json.dumps({
                            "jsonrpc": "2.0",
                            "method": "user.login",
                            "params": {
                                "user": url_userid,  # web页面登录用户名
                                "password": url_pwd  # web页面登录密码
                            },
                            "id": 0
                        })
                        request = self.session.post(url=url_http, headers=self.header, json=json_data_1)
                        # request = self.session.post(url=url_http, headers=self.header, data=json_data_2)
                        response = request.json()
                        if response.get('result', '') != '':
                            self.authID = response['result']
                        elif response.get('error', '') != '':
                            sys.exit("{0}".format(inspect.currentframe().f_lineno))
                        time.sleep(0.1)
                        title_name = ['index', 'status']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
                        self.cus_excel_op.def_set_cell_value(2, 1, 1)
                        self.cus_excel_op.def_set_cell_value(2, 2, response['status'])

                        col_01_list = self.cus_excel_op.def_get_creat_col_value(1)
                        del col_01_list[0]
                        col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                        del col_02_list[0]
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        state_resources = []
                        for len_col_01 in range(len(col_01_list)):
                            key_name = "{0}[{1}]".format(resource, col_01_list[len_col_01])
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_name, timestampnow, col_02_list[len_col_01]))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                    except Exception as ee:
                        state_resources = []
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        key_name = "{0}[{1}]".format(resource, 1)
                        state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_name, timestampnow, 1))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                        logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], ee))
                        sys.exit("{0}".format(inspect.currentframe().f_lineno))
                else:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
            # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)
            print(0)
        except Exception as pizdec:
            logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], pizdec))
            sys.exit("{0}".format(inspect.currentframe().f_lineno))

    def def_get_status_resources_web_service(self, host, list_resources, url_http, serviceid_str, marketid_int, stockcode_list):
        try:
            for resource in list_resources:
                if ['serviceid'].count(resource) == 1:
                    try:
                        logger.info("Starting discovering resource - {0}".format(resource))
                        self.session.mount(url_http, requests.adapters.HTTPAdapter(max_retries=3))

                        json_data_1 = {
                            "serviceid": serviceid_str,
                            "body": {
                                "marketid": marketid_int,
                                "stockcode": stockcode_list,
                            }
                        }
                        request = self.session.post(url=url_http, headers=self.header, json=json_data_1)
                        response = request.json()
                        if response.get('result', '') != '':
                            self.authID = response['result']
                        elif response.get('error', '') != '':
                            sys.exit("{0}".format(inspect.currentframe().f_lineno))
                        time.sleep(0.1)

                        title_name = ['index', 'name', 'now', 'volume', 'amount']
                        self.cus_excel_op.def_create_sheet(resource)
                        [self.cus_excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]

                        for i_01 in range(len(response['data'])):
                            self.cus_excel_op.def_set_cell_value(i_01 + 2, 1, i_01 + 1)
                            for i_02 in range(len(title_name)):
                                if i_02 == 1:
                                    self.cus_excel_op.def_set_cell_value(i_01 + 2, 2, response['data'][i_01]['code'] + '_' + response['data'][i_01]['name'])
                                elif i_02 == 2:
                                    self.cus_excel_op.def_set_cell_value(i_01 + 2, 3, response['data'][i_01]['now'])
                                elif i_02 == 3:
                                    self.cus_excel_op.def_set_cell_value(i_01 + 2, 4, response['data'][i_01]['volume'])
                                elif i_02 == 4:
                                    self.cus_excel_op.def_set_cell_value(i_01 + 2, 5, response['data'][i_01]['amount'])
                        col_02_list = self.cus_excel_op.def_get_creat_col_value(2)
                        del col_02_list[0]
                        col_03_list = self.cus_excel_op.def_get_creat_col_value(3)
                        del col_03_list[0]
                        col_04_list = self.cus_excel_op.def_get_creat_col_value(4)
                        del col_04_list[0]
                        col_05_list = self.cus_excel_op.def_get_creat_col_value(5)
                        del col_05_list[0]
                        timestampnow = int(time.time())
                        TIMESTAMPNOW = timestampnow
                        state_resources = []
                        for len_col_01 in range(len(col_02_list)):
                            key_name = "{0}[{1}]".format(resource, col_02_list[len_col_01])
                            state_resources.append("%s %s %s %s" % ("\"" + host + "\"", key_name, timestampnow, col_02_list[len_col_01]))
                        self.zabbix_sender.def_send_data_to_zabbix(TIMESTAMPNOW, state_resources, host)
                    except Exception as ee:
                        logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], ee))
                        sys.exit("{0}".format(inspect.currentframe().f_lineno))
                else:
                    logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], u'没有匹配的自动发现规则'))
            # self.cus_excel_op.def_save_create_xlsx(XLSX_FILENAME)
            print(0)
        except Exception as pizdec:
            logger.error(u"错误: {0} {1}".format(inspect.stack()[0][2], pizdec))
            sys.exit("{0}".format(inspect.currentframe().f_lineno))


class CusZabbixSender(object):
    def __init__(self, ):
        self.send_code = None
        self.output = None

    def def_convert_to_zabbix_json(self, data):
        self.output = json.dumps({"data": data}, indent=None, separators=(',', ': '))
        return self.output

    def def_send_data_to_zabbix(self, timestampnow, zabbix_data, storage_name):
        sender_command = "/usr/bin/zabbix_sender"
        time_of_create_file = timestampnow
        temp_file = "/tmp/{0}_{1}.tmp".format(storage_name, time_of_create_file)
        with open(temp_file, "w") as f:
            f.write("")
            f.write(u"\n".join(zabbix_data))
        self.send_code = subprocess.call([sender_command, "-vv", "-z", SERVER_IP, "-p", SERVER_PORT, "-s", storage_name, "-T", "-i", temp_file], stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30)
        time.sleep(0.1)
        os.remove(temp_file)
        if os.path.isfile(LOG_FILENAME):
            os.remove(LOG_FILENAME)
            pass


class CusLocalMethod(object):
    def __init__(self, ):
        self.value = None
        self.numericValue = None
        self.textreplace = None

    def def_hum_convert(self, value):
        # units = ["B", "KB", "MB", "GB", "TB", "PB"]
        units = ["KB", "MB", "GB", "TB", "PB"]
        size = 1024.0
        for i in range(len(units)):
            if (float(value) / size) < 1:
                self.value = "%.2f%s" % (value, units[i])
                return value
            value = float(value) / size

    def def_convert_text_to_numeric(self, value):
        if value == 'online':
            self.numericValue = 0
        elif value == 'offline':
            self.numericValue = 1
        elif value == 'degraded':
            self.numericValue = 2
        elif value == 'active':
            self.numericValue = 3
        elif value == 'inactive_configured':
            self.numericValue = 4
        elif value == 'inactive_unconfigured':
            self.numericValue = 5
        elif value == 'offline_unconfigured':
            self.numericValue = 6
        elif value == 'excluded':
            self.numericValue = 7
        elif value == 'on':
            self.numericValue = 8
        elif value == 'off':
            self.numericValue = 9
        elif value == 'slow_flashing':
            self.numericValue = 10
        elif value == 'degraded_paths':
            self.numericValue = 11
        elif value == 'degraded_ports':
            self.numericValue = 12
        elif value == 'up':
            self.numericValue = 0
        elif value == 'down':
            self.numericValue = 1
        elif value == 'LISTEN':
            self.numericValue = 0
        elif value == 'open':
            self.numericValue = 0
        elif value == 'closed':
            self.numericValue = 1
        else:
            self.numericValue = 100
        return self.numericValue

    def def_convert_text_to_text(self, value):
        if value == "::1" or value == "::" or value == "0.0.0.0":
            self.textreplace = "127.0.0.1"
        else:
            self.textreplace = value
        return self.textreplace

    def def_ping(self, ip):
        """
        获取节点的延迟的作用
        :param node:
        :return:
        """
        ip_address = ip
        response = ping(ip_address)
        if response is not None:
            delay = int(response * 1000)
            return delay


def main():
    cus_telnet_client = CusTelnetClient()
    cus_ssh_client = CusSSHClient()
    cus_web_client = CusWebClinet()
    cus_linux_client = CusLinuxClient()
    cus_direct_client = CusDirectClient()
    cus_sftp_client = CusSftpClient()
    ###########################################################
    parser = argparse.ArgumentParser()
    parser.add_argument('--type')
    parser.add_argument('--ip', help="Where to connect")
    parser.add_argument('--port')
    parser.add_argument('--user')
    parser.add_argument('--pwd')
    parser.add_argument('--host')
    parser.add_argument('--url_schoolid')
    parser.add_argument('--url_userid')
    parser.add_argument('--url_pwd')
    parser.add_argument('--url_http')
    parser.add_argument('--serviceid_str')
    parser.add_argument('--marketid_int')
    parser.add_argument('--stockcode_list')
    parser.add_argument('--value')

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--discovery', action='store_true')
    group.add_argument('--status', action='store_true')
    arguments = parser.parse_args()

    list_resources = []
    if ['telnet'].count(arguments.type) == 1:
        list_resources = ['disname', 'discpu', 'disinterface']
        list_resources = ['disname', 'discpu_20230104', 'dismem_20230104']
    elif ['ssh'].count(arguments.type) == 1:
        list_resources = ['fileCount']
    elif ['sftp'].count(arguments.type) == 1:
        list_resources = ['queueDepth']
    elif ['linux'].count(arguments.type) == 1:
        # list_resources = ['linux_df', 'linux_netstat', 'linux_checkping']
        list_resources = ['linux_test']
    elif ['web'].count(arguments.type) == 1:
        list_resources = ['webcode_schoolid']
    elif ['web_service'].count(arguments.type) == 1:
        list_resources = ['serviceid']
    elif ['direct'].count(arguments.type) == 1:
        list_resources = ['ssh_login_log']

    if arguments.discovery:
        logger.info("********************************* Starting Discovering *********************************")
        if ['telnet'].count(arguments.type) == 1:
            cus_telnet_client.def_discovering_resources(arguments.user, arguments.pwd, arguments.ip, arguments.port, arguments.host, list_resources)
        elif ['ssh'].count(arguments.type) == 1:
            cus_ssh_client.def_discovering_resources(arguments.host, list_resources)
        elif ['sftp'].count(arguments.type) == 1:
            cus_sftp_client.def_discovering_resources(arguments.host, list_resources)
        elif ['linux'].count(arguments.type) == 1:
            cus_linux_client.def_discovering_resources(arguments.host, list_resources)
        elif ['web'].count(arguments.type) == 1:
            cus_web_client.def_discovering_resources_webcode_schoolid(arguments.host, list_resources)
        elif ['web_service'].count(arguments.type) == 1:
            cus_web_client.def_discovering_resources_web_service(arguments.host, list_resources, arguments.url_http, arguments.serviceid_str, arguments.marketid_int, arguments.stockcode_list)

    elif arguments.status:
        logger.info("********************************* Starting Get Status *********************************")
        if ['telnet'].count(arguments.type) == 1:
            cus_telnet_client.def_get_status_resources(arguments.user, arguments.pwd, arguments.ip, arguments.port, arguments.host, list_resources)
        elif ['ssh'].count(arguments.type) == 1:
            cus_ssh_client.def_get_status_resources(arguments.host, list_resources)
        elif ['sftp'].count(arguments.type) == 1:
            cus_sftp_client.def_get_status_resources(arguments.host, list_resources)
        elif ['linux'].count(arguments.type) == 1:
            cus_linux_client.def_get_status_resources(arguments.host, list_resources)
        elif ['direct'].count(arguments.type) == 1:
            cus_direct_client.def_get_status_resources(arguments.host, list_resources, arguments.value)
        elif ['web'].count(arguments.type) == 1:
            cus_web_client.def_get_status_resources_webcode_schoolid(arguments.host, list_resources, arguments.url_http, arguments.url_userid, arguments.url_pwd, arguments.url_schoolid)
        elif ['web_service'].count(arguments.type) == 1:
            cus_web_client.def_get_status_resources_web_service(arguments.host, list_resources, arguments.url_http, arguments.serviceid_str, arguments.marketid_int, arguments.stockcode_list)


# result_status = get_status_resources(arguments.user, arguments.password, arguments.ip, arguments.port, arguments.storage_name, list_resources)
# print(result_status)


if __name__ == "__main__":
    #     a="""System Total Memory(bytes): 34590336
    # Total Used Memory(bytes): 14141896
    # Used Rate: 40%
    #
    #     """
    #     b = re.compile(r'Rate:\s(\d*%)')
    #     r = re.search(b, a)
    #     print(r.group(1))
    #     exit(1)
    main()
