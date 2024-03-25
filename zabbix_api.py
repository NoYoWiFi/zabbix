#!/usr/bin/python3
# coding:utf-8

__author__ = 'NoYoWiFi'
__date__ = '2023-9-11 16:53:39'

import argparse
import base64
import datetime
import os
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
import hashlib
import hmac
import inspect
import json
import openpyxl
import paramiko
import polib
import re
import sys
import queue
import threading
import time
from time import mktime
import requests
from requests.adapters import HTTPAdapter
from urllib.parse import urlencode
from wsgiref.handlers import format_date_time
import urllib3

urllib3.disable_warnings()

GV_CPU_COUNT = os.cpu_count()
GV_ERROR_MESS = {"error": ""}


class CusZabbixApi:
    def __init__(self):
        self.authID = None
        # self.url = 'https://172.16.50.50:8443/api_jsonrpc.php'  # 修改URL
        self.url = 'https://172.169.10.3:8443/api_jsonrpc.php'  # 修改URL
        self.header = {"Content-Type": "application/json"}
        self.session = requests.Session()
        self.session.mount(self.url, requests.adapters.HTTPAdapter(max_retries=3))
        self.def_login()
        self.gv_apiVersion = None
        self.textValue = None

    def def_login(self):
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "user.login",
            "params": {
                # "user": "Admin",  # 5.0.x版本web页面登录用户名
                "username": "Admin",  # 6.x.x版本web页面登录用户名
                "password": "zabbix"  # web页面登录密码
            },
            "id": 0
        })
        try:
            request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
            response = request.json()
            if response.get('result', '') != '':
                self.authID = response['result']
            elif response.get('error', '') != '':
                print(u"用户认证失败请检查! 原因: \033[;31m%s\033[0m" % (response['error']['data']))
                sys.exit(1)
        except Exception as ee:
            print(u"地址请求失败请检查! 原因: \033[;31m%s\033[0m" % ee)
            sys.exit(1)

    # ![01_创建主机组]
    def def_create_hostgroup(self, hostgroup_name):
        """
        通过该方式可以创建新的主机组。

        :param hostgroup_name: 主机组名称: "Linux servers"
        :return: "groupids": ["107819"]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "hostgroup.create",
            "params": {
                "name": "%s" % hostgroup_name,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_hostgroup_6_0(self, hostgroup_name_list):
        """
        该方法允许根据给定的参数检索主机组。

        :param hostgroup_name_list: 主机组列表: ["Zabbixservers","Linuxservers"]
        :return: [{'groupid': '21', 'name': 'Zabbixservers'}, {'groupid': '22', 'name': 'Linuxservers'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "hostgroup.get",
            "params": {
                "selectTemplates": [],
                "selectHosts": [],
                "filter": {
                    "name": hostgroup_name_list,
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_hostgroup_6_4(self, hostgroup_name_list):
        """
        该方法允许根据给定的参数检索主机组。

        :param hostgroup_name_list: 主机组列表: ["Zabbixservers","Linuxservers"]
        :return: [{'groupid': '21', 'name': 'Zabbixservers'}, {'groupid': '22', 'name': 'Linuxservers'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "hostgroup.get",
            "params": {
                "selectHosts": [],
                "filter": {
                    "name": hostgroup_name_list,
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}


    def def_get_all_hostgroup(self):
        """
        该方法允许根据给定的参数检索主机组。

        :param hostgroup_name_list: 主机组列表: ["Zabbixservers","Linuxservers"]
        :return: [{'groupid': '21', 'name': 'Zabbixservers'}, {'groupid': '22', 'name': 'Linuxservers'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "hostgroup.get",
            "params": {
                "output": "extend",
                "sortfield": "groupid",
                "sortorder": "ASC",
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_all_hostgroup_name(self, hostgroup_name_list):
        """
        此方法允许删除主机组。

        :param hostgroup_id_list: 主机组ID列表["107824"]
        :return: {'groupids': ['34']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "hostgroup.update",
            "params": hostgroup_name_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_delete_hostgroup(self, hostgroup_id_list):
        """
        此方法允许删除主机组。

        :param hostgroup_id_list: 主机组ID列表["107824"]
        :return: {'groupids': ['34']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "hostgroup.delete",
            "params": hostgroup_id_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    # ![02_创建模板]
    def def_create_template(self, template_name, groupids_list):
        """
        此方法允许创建新模板。

        :param template_name: 模板名称: "Linux template"
        :param groupids_list: 主机组ID列表: [{"groupid":1},{"groupid":2}]
        :return: {'templateids': ['10537']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.create",
            "params": {
                "host": "%s" % template_name,
                "groups": groupids_list,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template(self, template_name):
        """
        该方法允许根据给定的参数检索模板。

        :param template_name: 模板名称 "Linux"
        :return: [{'templateid': '10558'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.get",
            "params": {
                "output": "templateid",
                "filter": {
                    "host": template_name
                }
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_all_templategroup(self):
        """
        该方法允许根据给定的参数检索主机组。

        :param hostgroup_name_list: 主机组列表: ["Zabbixservers","Linuxservers"]
        :return: [{'groupid': '21', 'name': 'Zabbixservers'}, {'groupid': '22', 'name': 'Linuxservers'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "templategroup.get",
            "params": {
                "output": "extend",
                "sortfield": "groupid",
                "sortorder": "ASC",
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_all_templategroup_name(self, templategroup_name_list):
        """
        此方法允许删除主机组。

        :param hostgroup_id_list: 主机组ID列表["107824"]
        :return: {'groupids': ['34']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "templategroup.update",
            "params": templategroup_name_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}


    def def_get_all_templateid(self):
        """
        该方法允许根据给定的参数检索模板。

        :param template_name: 模板名称 "Linux"
        :return: [{'templateid': '10558'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.get",
            "params": {
                "output": ["host", "templateid"],
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_delete_template(self, template_id_list):
        """
        此方法允许删除模板。

        :param template_id_list: 模板ID列表: ["13"]
        :return:  {"templateids":["13"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.delete",
            "params": template_id_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massadd_template_groups(self, template_id_list, group_id_list):
        """
        此方法允许同时向给定模板添加多个相关对象。

        :param template_id_list: 模板ID列表: [{"templateid": "10085"}]
        :param group_id_list: 主机群组ID列表: [{"groupid": "2"},{"groupid": "3"}]
        :return: {'templateids': ['10560']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.massadd",
            "params": {
                "templates": template_id_list,
                "groups": group_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massremove_template_groups(self, template_id_list, group_id_list):
        """
        此方法允许从多个模板中删除相关对象。

        :param template_id_list: 模板ID列表: ["10085"]
        :param group_id_list: 主机群组ID列表: ["2","3"]
        :return: {"templateids":["10085"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.massremove",
            "params": {
                "templateids": template_id_list,
                "groupids": group_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massadd_template_macros(self, template_id_list, macros_list):
        """
        此方法允许同时向给定模板添加多个相关对象。

        :param template_id_list: 模板ID列表: [{"templateid": "10085"}]
        :param macros_list: 用户宏列表: [{"macro":"{$cacro1}","value":"1"},{"macro":"{$cacro2}","value":"2"}]
        :return: {'templateids': ['10560']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.massadd",
            "params": {
                "templates": template_id_list,
                "macros": macros_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massremove_template_macros(self, template_id_list, macros_id_list):
        """
        此方法允许从多个模板中删除相关对象。

        :param template_id_list: 模板ID列表: ["10085"]
        :param macros_id_list: 用户宏名称列表: ["{$E2}","{$E4}"]
        :return: {"templateids":["10085"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.massremove",
            "params": {
                "templateids": template_id_list,
                "macros": macros_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massadd_template_templates_link(self, template_id_list, templates_link_id_list):
        """
        此方法允许同时向给定模板添加多个相关对象。

        :param template_id_list: 模板ID列表: [{"templateid": "10085"}]
        :param templates_link_id_list: 模板ID列表: [{"templateid":"10106"},{"templateid":"10104"}]
        :return: {'templateids': ['10560']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.massadd",
            "params": {
                "templates": template_id_list,
                "templates_link": templates_link_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massremove_templateids_clear(self, template_id_list, template_clear_id_list):
        """
        此方法允许从多个模板中删除相关对象。

        :param template_id_list: 模板ID列表: ["10085"]
        :param template_clear_id_list: 模板ID列表: ["10090","10091"]
        :return: {"templateids":["10085"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.massremove",
            "params": {
                "templateids": template_id_list,
                "templateids_clear": template_clear_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massremove_templateids_link(self, template_id_list, template_link_id_list):
        """
        此方法允许从多个模板中删除相关对象。

        :param template_id_list: 模板ID列表: ["10085"]
        :param template_link_id_list: 模板ID列表: ["10090","10091"]
        :return: {"templateids":["10085"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.massremove",
            "params": {
                "templateids": template_id_list,
                "templateids_link": template_link_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_tags(self, template_id_list, tags_list):
        """
        此方法允许更新现有模板。

        :param template_id_list: 模板ID: "10085"
        :param tags_list: 标签列表: [{"tag":"Hostname1","value":"{HOST.NAME1}"},{"tag":"Hostname2","value":"{HOST.NAME2}"}]
        :return: {"hostids":["10086"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.update",
            "params": {
                "templateid": template_id_list,
                "tags": tags_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_tags_bytemplateid(self, tags_list):
        """
        此方法允许更新现有模板。

        :param template_id_list: 模板ID: "10085"
        :param tags_list: 标签列表: [{"tag":"Hostname1","value":"{HOST.NAME1}"},{"tag":"Hostname2","value":"{HOST.NAME2}"}]
        :return: {"hostids":["10086"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.update",
            "params": tags_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_tags(self, template_name):
        """
        该方法允许根据给定的参数检索模板。

        :param template_name: 模板名称 "Linux"
        :return: [{'tag': '1', 'value': '1'}, {'tag': '2', 'value': '2'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.get",
            "params": {
                "output": ["tags"],
                "selectTags": "extend",
                "evaltype": 0,
                "filter": {
                    "host": [
                        "%s" % template_name
                    ]
                }
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_tags_bytemplateid(self, template_id):
        """
        该方法允许根据给定的参数检索模板。

        :param template_name: 模板名称 "Linux"
        :return: [{'tag': '1', 'value': '1'}, {'tag': '2', 'value': '2'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.get",
            "params": {
                "output": ["tags"],
                "selectTags": "extend",
                "evaltype": 0,
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_all_template_name(self):
        """
        此方法用于依据给定的参数检索监控项

        :return:  [{'templateid': '42187', 'host': 'Host name of Zabbix agent running'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "template.get",
            "params": {
                "output": ["host"],
                "sortfield": "host",
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    # ![06_模板创建监控项]
    def def_get_template_item(self, template_id, item_key):
        """
        此方法用于依据给定的参数检索监控项

        :param template_id: 模板ID: "10084"
        :param item_key: 监控项键值: "system.cpu.util[,idle]"
        :return: [{'itemid': '43855'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.get",
            "params": {
                "output": "itemids",
                "templateids": template_id,
                "filter": {
                    "key_": item_key
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_item_name(self, template_id):
        """
        此方法用于依据给定的参数检索监控项

        :param template_id: 模板ID: "10084"
        :return:  [{'itemid': '42187', 'name': 'Host name of Zabbix agent running'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.get",
            "params": {
                "output": ["name"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_itemprototype_name(self, template_id):
        """
        此方法用于依据给定的参数检索监控项

        :param template_id: 模板ID: "10084"
        :return:  [{'itemid': '42187', 'name': 'Host name of Zabbix agent running'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "itemprototype.get",
            "params": {
                "output": ["name"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_itemprototype_tags_bytemplateid(self, template_id):
        """
        此方法用于依据给定的参数检索监控项

        :param template_id: 模板ID: "10084"
        :return:  [{'itemid': '42187', 'name': 'Host name of Zabbix agent running'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "itemprototype.get",
            "params": {
                "output": ["tags"],
                "selectTags": "extend",
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_trigger_name(self, template_id):
        """
        此方法用于依据给定的参数检索触发器

        :param template_id: 模板ID: "10084"
        :return:  [{'triggerid': '42187', 'description': 'Remote Zabbix proxy: More than {$ZABBIX.PROXY.UTIL.MAX}% used in the history index cache'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.get",
            "params": {
                "output": ["description"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_event_name(self, template_id):
        """
        此方法用于依据给定的参数检索触发器

        :param template_id: 模板ID: "10084"
        :return:  [{'triggerid': '42187', 'description': 'Remote Zabbix proxy: More than {$ZABBIX.PROXY.UTIL.MAX}% used in the history index cache'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.get",
            "params": {
                "output": ["event_name"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_triggerprototype_name(self, template_id):
        """
        此方法用于依据给定的参数检索触发器

        :param template_id: 模板ID: "10084"
        :return:  [{'triggerid': '42187', 'description': 'Remote Zabbix proxy: More than {$ZABBIX.PROXY.UTIL.MAX}% used in the history index cache'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "triggerprototype.get",
            "params": {
                "output": ["description"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_triggerprototype_event_name(self, template_id):
        """
        此方法用于依据给定的参数检索触发器

        :param template_id: 模板ID: "10084"
        :return:  [{'triggerid': '42187', 'description': 'Remote Zabbix proxy: More than {$ZABBIX.PROXY.UTIL.MAX}% used in the history index cache'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "triggerprototype.get",
            "params": {
                "output": ["event_name"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_triggerprototype_tag_bytemplateid(self, template_id):
        """
        此方法用于依据给定的参数检索触发器

        :param template_id: 模板ID: "10084"
        :return:  [{'triggerid': '42187', 'description': 'Remote Zabbix proxy: More than {$ZABBIX.PROXY.UTIL.MAX}% used in the history index cache'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "triggerprototype.get",
            "params": {
                "output": ["tags"],
                "selectTags": "extend",
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_application_bytemplateid(self, template_id):
        """
        此方法用于依据给定的参数检索触发器

        :param template_id: 模板ID: "10084"
        :return:  [{'triggerid': '42187', 'description': 'Remote Zabbix proxy: More than {$ZABBIX.PROXY.UTIL.MAX}% used in the history index cache'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "application.get",
            "params": {
                "output": ["name"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_graph_name(self, template_id):
        """
        此方法用于依据给定的参数检索触发器

        :param template_id: 模板ID: "10084"
        :return:  [{'graphid': '42187', 'name': '"CPU jumps"'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "graph.get",
            "params": {
                "output": ["name"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_graphprototype_name(self, template_id):
        """
        此方法用于依据给定的参数检索触发器

        :param template_id: 模板ID: "10084"
        :return:  [{'graphid': '42187', 'name': '"CPU jumps"'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "graphprototype.get",
            "params": {
                "output": ["name"],
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_create_template_item(self, template_id, item_name, item_type, item_key, item_value_type, item_delay, snmp_oid):
        """
        此方法用于创建新监控项。

        :param template_id: 模板ID: "10085"
        :param item_name: 监控项名称: "Free disk space on /home/joe/"
        :param item_type: 监控项类型: 0
        :param item_key: 监控项关键字: "vfs.fs.size[/home/joe/,free]"
        :param item_value_type: 监控项数据类型: 3
        :param item_delay: 更新监控项的时间间隔: "30s"
        :param snmp_oid: snmp_oid: ".1.3.6.1.4.1.1"
        :return: {"itemids":["24758"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.create",
            "params": {
                "hostid": template_id,
                "name": item_name,
                "type": item_type,
                "key_": item_key,
                "value_type": item_value_type,
                "delay": item_delay,
                "snmp_oid": snmp_oid
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_delete_template_item(self, template_id_list):
        """
        此方法用于删除监控项。

        :param template_id_list: 模板监控项ID列表: ["22982","22986"]
        :return: {"itemids":["22982","22986"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.delete",
            "params": template_id_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_template_item_tags(self, item_id, tags_list):
        """
        此方法允许更新现有模板。

        :param item_id: 监控项ID: "10085"
        :param tags_list: 标签列表: [{"tag":"Hostname1","value":"{HOST.NAME1}"},{"tag":"Hostname2","value":"{HOST.NAME2}"}]
        :return: {"hostids":["10086"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.update",
            "params": {
                "itemid": item_id,
                "tags": tags_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_template_item_tags_bytemplateid(self, tags_list):
        """
        此方法允许更新现有模板。

        :param item_id: 监控项ID: "10085"
        :param tags_list: 标签列表: [{"tag":"Hostname1","value":"{HOST.NAME1}"},{"tag":"Hostname2","value":"{HOST.NAME2}"}]
        :return: {"hostids":["10086"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.update",
            "params": tags_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_item_tags(self, item_id):
        """
        该方法允许根据给定的参数检索监控项。

        :param item_id: 监控项ID: "23298"
        :return: [{'tag': '1', 'value': '1'}, {'tag': '2', 'value': '2'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.get",
            "params": {
                "output": ["tags"],
                "selectTags": "extend",
                "evaltype": 0,
                "filter": {
                    "itemid": [
                        "%s" % item_id
                    ]
                }
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_item_tags_bytemplateid(self, template_id):
        """
        该方法允许根据给定的参数检索监控项。

        :param item_id: 监控项ID: "23298"
        :return: [{'tag': '1', 'value': '1'}, {'tag': '2', 'value': '2'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.get",
            "params": {
                "output": ["tags"],
                "selectTags": "extend",
                "evaltype": 0,
                "templateids": template_id,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    # ![04_为模板创建触发器]
    def def_create_template_trigger(self, trigger_name, priority, expression_name, recovery_expression_name):
        """
        此方法允许创建新的触发器.

        :param trigger_name: 触发器名称: "Processor load is too high on {HOST.NAME}"
        :param priority: 触发器的严重性级别: 2
        :param expression_name: 简化的触发器表达式: "last(/Linux server/system.cpu.load[percpu,avg1])>5"
        :param recovery_expression_name: 生成的触发恢复表达式: "last(/Linux server/system.cpu.load[percpu,avg1])<=5"
        :return: {"triggerids":["17369","17370"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.create",
            "params": {
                "description": trigger_name,
                "priority": priority,
                "expression": expression_name,
                "recovery_expression": recovery_expression_name,
                "manual_close": 1,
                "recovery_mode": 1
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_trigger(self, template_id, trigger_name):
        """
        此方法允许根据指定的参数检索触发器.

        :param trigger_name: 触发器名称: "Processor load is too high on {HOST.NAME}"
        :return: "17369"
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.get",
            "params": {
                "output": "triggerid",
                "templateids": template_id,
                "filter": {
                    "description": trigger_name,
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_delete_template_trigger(self, trigger_id):
        """
             此方法用于删除触发器。

             :param trigger_id: 监控项ID列表: ["12002","12003"]
             :return: {"triggerids":["12002","12003"]}
             """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.delete",
            "params": trigger_id,
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_trigger_tags(self, trigger_id):
        """
        该方法允许根据给定的参数检索监控项。

        :param trigger_id: 触发器ID: "23298"
        :return: [{'tag': '1', 'value': '1'}, {'tag': '2', 'value': '2'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.get",
            "params": {
                "output": ["tags"],
                "selectTags": "extend",
                "evaltype": 0,
                "filter": {
                    "triggerid": trigger_id
                }
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_trigger_tags_bytemplateid(self, template_id):
        """
        该方法允许根据给定的参数检索监控项。

        :param trigger_id: 触发器ID: "23298"
        :return: [{'tag': '1', 'value': '1'}, {'tag': '2', 'value': '2'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.get",
            "params": {
                "output": ["tags"],
                "selectTags": "extend",
                "templateids": template_id,
                "evaltype": 0,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_template_trigger_tags(self, trigger_id, tags_list):
        """
        此方法允许更新现有模板。

        :param trigger_id: 触发器ID: "10085"
        :param tags_list: 标签列表: [{"tag":"Hostname1","value":"{HOST.NAME1}"},{"tag":"Hostname2","value":"{HOST.NAME2}"}]
        :return: {"triggerids":["13938"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.update",
            "params": {
                "triggerid": trigger_id,
                "tags": tags_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_template_trigger_tags_bytemplateid(self, tags_list):
        """
        此方法允许更新现有模板。

        :param trigger_id: 触发器ID: "10085"
        :param tags_list: 标签列表: [{"tag":"Hostname1","value":"{HOST.NAME1}"},{"tag":"Hostname2","value":"{HOST.NAME2}"}]
        :return: {"triggerids":["13938"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.update",
            "params": tags_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}


    def def_get_discoveryrule(self, hostid, name):
        """
        该方法允许根据给定的参数检索模板发现规则。

        :param hostid: 模板ID: "10534"
        :param name: LLD规则名称: "test"
        :return:  [{'itemid': '44081'}, {'itemid': '44083'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "discoveryrule.get",
            "params": {
                "output": ["itemid"],
                "hostids": hostid,
                "filter": {
                    "name": name
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_create_discoveryrule(self, hostid, name, type, key_):
        """
        此方法允许创建新的模板发现规则。

        :param hostid: LLD规则的主机ID: '10197'
        :param name: LLD规则名称: 'Mounted filesystem discovery'
        :param type: LLD规则类型: '0'
        :param key_: LLD规则键值: 'vfs.fs.discovery'
        :return: "itemids":["27665"]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "discoveryrule.create",
            "params": {
                "name": name,
                "key_": key_,
                "hostid": hostid,
                "type": type,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_delete_discoveryrule(self, discoveryrule_id_list):
        """
        此方法用于删除模板发现规则。

        :param discoveryrule_id_list: 模板发现规则ID列表: ["22982","22986"]
        :return: {"ruleids":["22982","22986"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "discoveryrule.delete",
            "params": discoveryrule_id_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_create_itemprototype(self, ruleid, template_id, item_name, item_type, item_key, item_units, item_value_type, item_delay, snmp_oid):
        """
        此方法允许创建新的模板发现规则。

        :param ruleid: 模板发现规则ID: "10065"
        :param template_id: 模板ID: "10085"
        :param item_name: 模板发现规则监控项名称: "Free disk space on /home/joe/"
        :param item_type: 模板发现规则监控项类型: 0
        :param item_key: 模板发现规则监控项关键字: "vfs.fs.size[/home/joe/,free]"
        :param item_units: 模板发现规则监控项单位: "%"
        :param item_value_type: 模板发现规则监控项数据类型: 3
        :param item_delay: 模板发现规则更新监控项的时间间隔: "30s"
        :param snmp_oid: 模板发现规则snmp_oid: ".1.3.6.1.4.1.1"
        :return: {"itemids":["24758"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "itemprototype.create",
            "params": {
                "ruleid": ruleid,
                "hostid": template_id,
                "name": item_name,
                "type": item_type,
                "key_": item_key,
                "units": item_units,
                "value_type": item_value_type,
                "delay": item_delay,
                "snmp_oid": snmp_oid
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_itemprototype_item(self, template_id, item_key):
        """
        此方法用于依据给定的参数检索模板发现规则监控项

        :param template_id: 模板ID: "10084"
        :param item_key: 监控项键值: "system.cpu.util[,idle]"
        :return: [{'itemid': '43855'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "itemprototype.get",
            "params": {
                "output": "itemid",
                "discoveryids": template_id,
                "filter": {
                    "key_": item_key
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_delete_itemprototype(self, discoveryrule_id_list):
        """
        此方法用于删除模板发现规则。

        :param discoveryrule_id_list: 模板发现规则ID列表: ["22982","22986"]
        :return: {'prototypeids': ['44237']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "itemprototype.delete",
            "params": discoveryrule_id_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_create_template_triggerprototype(self, trigger_name, priority, expression_name, recovery_expression_name):
        """
        此方法允许创建新的触发器.

        :param trigger_name: 触发器名称: "Processor load is too high on {HOST.NAME}"
        :param priority: 触发器的严重性级别: 2
        :param expression_name: 简化的触发器表达式: "last(/Linux server/system.cpu.load[percpu,avg1])>5"
        :param recovery_expression_name: 生成的触发恢复表达式: "last(/Linux server/system.cpu.load[percpu,avg1])<=5"
        :return: {"triggerids":["17369","17370"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "triggerprototype.create",
            "params": {
                "description": trigger_name,
                "priority": priority,
                "expression": expression_name,
                "recovery_expression": recovery_expression_name,
                "manual_close": 1,
                "recovery_mode": 1
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_template_triggerprototype(self, trigger_name):
        """
        此方法允许根据指定的参数检索触发器.

        :param trigger_name: 触发器名称: "Processor load is too high on {HOST.NAME}"
        :return: "17369"
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "triggerprototype.get",
            "params": {
                "output": "triggerid",
                "filter": {
                    "description": trigger_name
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_delete_template_triggerprototype(self, trigger_id):
        """
             此方法用于删除触发器。

             :param trigger_id: 监控项ID列表: ["12002","12003"]
             :return: {"triggerids":["12002","12003"]}
             """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "triggerprototype.delete",
            "params": trigger_id,
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    # ![10_创建主机]
    def def_create_host(self, host_name, template_id_list, group_id_list, type, ip, port, version,
                        securitylevel, authprotocol, privprotocol, ipmi_authtype, ipmi_privilege,
                        ipmi_username, ipmi_password, name, snmp_community, snmp_securityname, snmp_authpassphrase, snmp_privpassphrase):
        """
        这个方法可以用来创建主机。

        :param host_name: 主机名称: "Linux server"
        :param template_id_list: 模板ID列表: [{"templateid":"20045"}]
        :param group_id_list: 主机组ID列表: [{"groupid":"50"}]
        :param type: 接口类型: 1
        :param ip: IP地址: "192.168.3.1"
        :param port: 端口号: "10050"
        :param version: 版本: 3
        :param securitylevel: 安全级别: 2
        :param authprotocol: 身份认证协议: 0
        :param privprotocol: 隐私协议: 0
        :return: {"hostids":["10658"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.create",
            "params": {
                "host": host_name,
                "templates": template_id_list,
                "groups": group_id_list,
                "name": name,
                "interfaces": [
                    {
                        "type": type,
                        "ip": ip,
                        "port": port,
                        "dns": "",
                        "main": 1,
                        "useip": 1,
                        "details": {
                            "version": version,
                            "bulk": "1",
                            "community": snmp_community,
                            "securityname": snmp_securityname,
                            "securitylevel": securitylevel,
                            "authprotocol": authprotocol,
                            "authpassphrase": snmp_authpassphrase,
                            "privprotocol": privprotocol,
                            "privpassphrase": snmp_privpassphrase
                        }
                    }
                ],
                "ipmi_authtype": ipmi_authtype,
                "ipmi_privilege": ipmi_privilege,
                "ipmi_username": ipmi_username,
                "ipmi_password": ipmi_password
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_host(self, host_name):
        """
        此方法允许根据指定的参数获取主机。

        :param host_name: 主机名称: ["Zabbixserver"]
        :return: [{'hostid': '10600', 'groups': [{'groupid': '26', 'name': '000_LocalTemplates', 'internal': '0', 'flags': '0', 'uuid': 'dfcb57e7223e4446b8c27702e7c01ab8'}]}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.get",
            "params": {
                "output": ["hostid"],
                "selectGroups": "extend",
                "filter": {
                    "host": host_name
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_delete_host(self, host_id_list):
        """
        这个方法允许删除主机。

        :param host_id_list: 要删除主机的ID: ["13"]
        :return: {"hostids":["13","32"]}
        """

        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.delete",
            "params": host_id_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    # ![11_主机创建接口]
    def def_massadd_host_interface(self, host_id_list, type, ip, port, version, securitylevel, authprotocol, privprotocol,
                                   snmp_community, snmp_securityname, snmp_authpassphrase, snmp_privpassphrase):
        """
        此方法允许同时添加多个相关对象到所有给定的主机。

        :param host_id_list: 主机ID列表: [{"hostid":"10160"}]
        :param type: 接口类型: 1
        :param ip: IP地址: "192.168.3.1"
        :param port: 端口号: "10050"
        :param version: 版本: 3
        :param securitylevel: 安全级别: 2
        :param authprotocol: 身份认证协议: 0
        :param privprotocol: 隐私协议: 0
        :return: {'interfaceids': {'interfaceids': ['72']}}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "hostinterface.massadd",
            "params": {
                "hosts": host_id_list,
                "interfaces": [
                    {
                        "type": type,
                        "ip": ip,
                        "port": port,
                        "dns": "",
                        "main": 1,
                        "useip": 1,
                        "details": {
                            "version": version,
                            "bulk": "1",
                            "community": snmp_community,
                            "securityname": snmp_securityname,
                            "securitylevel": securitylevel,
                            "authprotocol": authprotocol,
                            "authpassphrase": snmp_authpassphrase,
                            "privprotocol": privprotocol,
                            "privpassphrase": snmp_privpassphrase
                        }
                    }
                ],
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massremove_host_interface(self, host_id_list, type, ip, port):
        """
        该方法允许从给定的主机列表中批量删除主机接口

        :param host_id_list: 主机ID列表: ["10160"]
        :param type: 接口类型: 1
        :param ip: IP地址: "192.168.3.1"
        :param port: 端口号: "10050"
        :return: {"interfaceids":["30069"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "hostinterface.massremove",
            "params": {
                "hostids": host_id_list,
                "interfaces": {
                    "dns": "",
                    "type": type,
                    "ip": ip,
                    "port": port,
                },
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    # ![12_主机关联模板]
    def def_massadd_host_template(self, host_id_list, template_id_list):
        """
        此方法允许同时添加多个相关对象到所有给定的主机。

        :param host_id_list: 主机ID列表: [{"hostid":"10160"}]
        :param template_id_list: 模板ID列表: [{"templateid":"1160"}]
        :return: {'hostids': ['10591']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.massadd",
            "params": {
                "hosts": host_id_list,
                "templates": template_id_list,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massadd_host_template_base_20221003(self, excel_op, index, host_name, host_ip, template_name, host_id_list, template_id_list):
        """
        此方法允许同时添加多个相关对象到所有给定的主机。

        :param index: 序号: "1"
        :param host_name: 主机名: "host1"
        :param host_ip: 主机IP: "172.169.10.2"
        :param template_name: 模板名称: "template1"
        :param host_id_list: 主机ID列表: [{"hostid":"10160"}]
        :param template_id_list: 模板ID列表: [{"templateid":"1160"}]
        :return: {'hostids': ['10591']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.massadd",
            "params": {
                "hosts": host_id_list,
                "templates": template_id_list,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', '') != '':
            print(u"主机: \033[;32m%s\033[0m 关联模板 \033[;32m%s\033[0m 成功! 返回值为: \033[;32m%s\033[0m" % (
                host_name, template_name, response['result']))
            excel_op.def_set_cell_value(index + 1, 1, index)
            excel_op.def_set_cell_value(index + 1, 2, host_name)
            excel_op.def_set_cell_value(index + 1, 3, host_ip)
            excel_op.def_set_cell_value(index + 1, 4, template_name)
            excel_op.def_set_cell_value(index + 1, 5, '成功')
        elif response.get('error', '') != '':
            print(u"主机: \033[;31m%s\033[0m 关联模板 \033[;32m%s\033[0m 失败! 原因: \033[;31m%s\033[0m" % (
                host_name, template_name, response['error']['data']))
            excel_op.def_set_cell_value(index + 1, 1, index)
            excel_op.def_set_cell_value(index + 1, 2, host_name)
            excel_op.def_set_cell_value(index + 1, 3, host_ip)
            excel_op.def_set_cell_value(index + 1, 4, template_name)
            excel_op.def_set_cell_value(index + 1, 5, '失败')
            excel_op.def_set_cell_value(index + 1, 6, response['error']['data'])

    def def_massadd_host_template_base_item_20221003(self, excel_op, index, host_name, host_ip, history_value, template_name, host_id_list, template_id_list):
        """
        此方法允许同时添加多个相关对象到所有给定的主机。

        :param index: 序号: "1"
        :param host_name: 主机名: "host1"
        :param host_ip: 主机IP: "192.168.0.1"
        :param history_value: 历史值: "1001"
        :param template_name: 模板名称: "template1"
        :param host_id_list: 主机ID列表: [{"hostid":"10160"}]
        :param template_id_list: 模板ID列表: [{"templateid":"1160"}]
        :return: {'hostids': ['10591']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.massadd",
            "params": {
                "hosts": host_id_list,
                "templates": template_id_list,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', '') != '' or not template_id_list:
            excel_op.def_set_cell_value(index + 1, 1, index)
            excel_op.def_set_cell_value(index + 1, 2, host_name)
            excel_op.def_set_cell_value(index + 1, 3, host_ip)
            excel_op.def_set_cell_value(index + 1, 4, history_value)
            excel_op.def_set_cell_value(index + 1, 5, template_name)
            if template_id_list:
                excel_op.def_set_cell_value(index + 1, 6, '成功')
                print(u"主机: \033[;32m%s\033[0m 获取监控项的值为: \033[;32m%s\033[0m 关联模板 \033[;32m%s\033[0m 成功! 返回值为: \033[;32m%s\033[0m"
                      % (host_name, history_value, template_name, response['result']))
            else:
                excel_op.def_set_cell_value(index + 1, 6, '失败')
                excel_op.def_set_cell_value(index + 1, 7, '未找到定义的模板名')
                print(u"主机: \033[;31m%s\033[0m 获取监控项的值为: \033[;31m%s\033[0m 关联模板 \033[;31m%s\033[0m 失败! 原因: \033[;31m%s\033[0m"
                      % (host_name, history_value, template_name, '未找到定义的模板名'))

        elif response.get('error', '') != '':
            print(u"主机: \033[;31m%s\033[0m 获取监控项的值为: \033[;32m%s\033[0m 关联模板 \033[;32m%s\033[0m 失败! 原因: \033[;31m%s\033[0m"
                  % (host_name, history_value, template_name, response['error']['data']))
            excel_op.def_set_cell_value(index + 1, 1, index)
            excel_op.def_set_cell_value(index + 1, 2, host_name)
            excel_op.def_set_cell_value(index + 1, 3, host_ip)
            excel_op.def_set_cell_value(index + 1, 4, history_value)
            excel_op.def_set_cell_value(index + 1, 5, template_name)
            excel_op.def_set_cell_value(index + 1, 6, '失败')
            excel_op.def_set_cell_value(index + 1, 7, response['error']['data'])

    def def_massremove_host_templateids(self, host_id_list, template_id_list):
        """
        该方法允许从多个主机中移除相关对象。

        :param host_id_list: 主机ID列表: ["10160"]
        :param template_id_list: 模板ID列表 ["325"]
        :return: {"hostids":["69665"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.massremove",
            "params": {
                "hostids": host_id_list,
                "templateids": template_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massremove_host_templateids_clear(self, host_id_list, template_id_list):
        """
        该方法允许从多个主机中移除相关对象。

        :param host_id_list: 主机ID列表: ["10160"]
        :param template_id_list: 模板ID列表 ["325"]
        :return: {"hostids":["69665"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.massremove",
            "params": {
                "hostids": host_id_list,
                "templateids_clear": template_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massremove_host_templateids_clear_base_20221003(self, excel_op, index, host_name, host_ip, template_name, host_id_list, template_id_list):
        """
        该方法允许从多个主机中移除相关对象。

        :param index: 序号: "1"
        :param host_name: 主机名: "host1"
        :param host_ip: 主机IP: "172.169.10.2"
        :param template_name: 模板名称: "template1"
        :param host_id_list: 主机ID列表: ["10160"]
        :param template_id_list: 模板ID列表 ["325"]
        :return: {"hostids":["69665"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.massremove",
            "params": {
                "hostids": host_id_list,
                "templateids_clear": template_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', '') != '':
            print(u"主机: \033[;32m%s\033[0m 脱离模板 \033[;32m%s\033[0m 清理监控项成功! 返回值为: \033[;32m%s\033[0m" % (
                host_name, template_name, response['result']))
            excel_op.def_set_cell_value(index + 1, 1, index)
            excel_op.def_set_cell_value(index + 1, 2, host_name)
            excel_op.def_set_cell_value(index + 1, 3, host_ip)
            excel_op.def_set_cell_value(index + 1, 4, template_name)
            excel_op.def_set_cell_value(index + 1, 5, '成功')
        elif response.get('error', '') != '':
            print(u"主机: \033[;31m%s\033[0m 脱离模板 \033[;32m%s\033[0m 清理监控项失败! 原因: \033[;31m%s\033[0m" % (
                host_name, template_name, response['error']['data']))
            excel_op.def_set_cell_value(index + 1, 1, index)
            excel_op.def_set_cell_value(index + 1, 2, host_name)
            excel_op.def_set_cell_value(index + 1, 3, host_ip)
            excel_op.def_set_cell_value(index + 1, 4, template_name)
            excel_op.def_set_cell_value(index + 1, 5, '失败')
            excel_op.def_set_cell_value(index + 1, 6, response['error']['data'])

    # ![13_主机关联主机组]
    def def_massadd_host_groups(self, host_id_list, group_id_list):
        """
        该方法允许从多个主机中移除相关对象。

        :param host_id_list: 主机组ID列表: [{"hostid":"10160"}]
        :param group_id_list: 主机组ID列表: [{"groupid":"1160"}]
        :return: {'hostids': ['10591']}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.massadd",
            "params": {
                "hosts": host_id_list,
                "groups": group_id_list,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_massremove_host_group(self, host_id_list, group_id_list):
        """
        该方法允许从多个主机中移除相关对象。

        :param host_id_list: 主机ID列表: ["10160"]
        :param group_id_list: 主机组ID列表 ["325"]
        :return: {"hostids":["69665"]}
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.massremove",
            "params": {
                "hostids": host_id_list,
                "groupids": group_id_list
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_host_item(self, host_name, item_key):
        """
        此方法用于依据给定的参数检索监控项

        :param host_name: 主机名称: "Linux"
        :param item_key: 监控项键值: "system.cpu.util[,idle]"
        :return: [{'itemid': '43855'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.get",
            "params": {
                "output": "extend",
                "host": host_name,
                "filter": {
                    "key_": item_key
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_group_item(self, group_name, item_key):
        """
        此方法用于依据给定的参数检索监控项

        :param host_name: 主机名称: "Linux"
        :param item_key: 监控项键值: "system.cpu.util[,idle]"
        :return: [{'itemid': '43855'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.get",
            "params": {
                "output": "extend",
                "group": group_name,
                "search": {
                    "key_": item_key
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}


    def def_get_item_history(self, item_id, history, time_from, time_till):
        """
        该方法允许根据给定的参数检索历史数据。

        :param item_id: 监控项ID: "23298"
        :param history: 数据类型: 0
        :param time_from: "1351090996"
        :param time_till: "1351091936"
        :return: [{'itemid': '47789', 'clock': '1664431589', 'value': '1', 'ns': '816892476'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "history.get",
            "params": {
                "output": "extend",
                "itemids": item_id,
                "history": history,
                "time_from": time_from,
                "time_till": time_till
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_item_history_base_20221003(self, item_id, history):
        """
        该方法允许根据给定的参数检索历史数据。

        :param item_id: 监控项ID: "23298"
        :param history: 数据类型: 0
        :return: [{'itemid': '47789', 'clock': '1664431589', 'value': '1', 'ns': '816892476'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "history.get",
            "params": {
                "output": "extend",
                "itemids": item_id,
                "history": history,
                "sortfield": "clock",
                "sortorder": "DESC",
                "limit": 1
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return [{'value': 'X'}]

    def def_get_all_history(self, item_id, history, time_from, time_till):
        """
        该方法允许根据给定的参数检索历史数据。

        :param item_id: 监控项ID: "23298"
        :param history: 数据类型: 0
        :param time_from: "1351090996"
        :param time_till: "1351091936"
        :return: [{'itemid': '47789', 'clock': '1664431589', 'value': '1', 'ns': '816892476'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "history.get",
            "params": {
                "output": "extend",
                "itemids": item_id,
                "history": history,
                "sortfield": "clock",
                "sortorder": "DESC",
                "time_from": time_from,
                "time_till": time_till
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_action(self, action_name):
        """
        该方法允许根据给定的参数检索动作。

        :param action_name: 动作名称 "Report problems to Zabbix administrators"
        :return: [{'actionid': '3'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "action.get",
            "params": {
                "output": "actionid",
                "filter": {
                    "name": action_name
                }
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_all_alert(self, actionids, time_from, time_till):
        """
        该方法允许根据给定的参数检索历史数据。

        :param actionids: "3"
        :param time_from: "1351090996"
        :param time_till: "1351091936"
        :return: [{'itemid': '47789', 'clock': '1664431589', 'value': '1', 'ns': '816892476'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "alert.get",
            "params": {
                "output": "extend",
                "actionids": actionids,
                "sortfield": "clock",
                "sortorder": "DESC",
                "time_from": time_from,
                "time_till": time_till
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        value_list_1 = []
        for i in response['result']:
            if i['message'] == '':
                continue
            dic = {}
            dic.update({'告警主题': i["subject"], '执行状态': i["status"]})
            vlist = i['message'].split('\r\n')
            for vl1 in range(len(list(vlist))):
                # print(vlist[vl1])
                pattern = re.compile(r'(.*?): (.*)$')
                result = None
                result = re.search(pattern, vlist[vl1])
                if result:
                    if result.group(2) == "":
                        dic.update({result.group(1): ''})
                    else:
                        dic.update({result.group(1): result.group(2)})
                else:
                    continue
            value_list_1.append(dic)
        if not value_list_1:
            return []
        return value_list_1

    def def_get_all_alert_custom(self, actionids, time_from, time_till):
        """
        该方法允许根据给定的参数检索历史数据。

        :param actionids: "3"
        :param time_from: "1351090996"
        :param time_till: "1351091936"
        :return: [{'itemid': '47789', 'clock': '1664431589', 'value': '1', 'ns': '816892476'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "alert.get",
            "params": {
                "output": "extend",
                "actionids": actionids,
                "sortfield": "clock",
                "sortorder": "DESC",
                "time_from": time_from,
                "time_till": time_till
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        value_list_1 = []
        for i in response['result']:
            if i['message'] == '':
                continue

            dic = {}
            dic.update({'clock': i["clock"], 'subject': i["subject"], 'status': i["status"]})

            pattern = re.compile(r'.*(\d{4}\.\d{2}\.\d{2}).*\r[^\r]')
            result = None
            result = re.search(pattern, i['message'])
            if result:
                dic.update({'message_date': result.group(1)})
            else:
                dic.update({'message_date': ''})

            pattern = re.compile(r'.*(\d{2}:\d{2}:\d{2}).*\r[^\r]')
            result = None
            result = re.search(pattern, i['message'])
            if result:
                dic.update({'message_clock': result.group(1)})
            else:
                dic.update({'message_clock': ''})

            pattern = re.compile(r'.*(故障).*\r[^\r]')
            result = None
            result = re.search(pattern, i['message'])
            if result:
                dic.update({'message_status': result.group(1)})
            else:
                dic.update({'message_status': ''})

            pattern = re.compile(r'.*(发生|,): (.*)\r[^\r]')
            result = None
            result = re.search(pattern, i['message'])
            if result:
                dic.update({'message_name': result.group(2)})
            else:
                dic.update({'message_name': ''})

            try:
                dic.update({'message_duration': result.group(1)})
            except:
                dic.update({'message_duration': ''})

            pattern = re.compile(r'.*告警主机:(.*_ROS_.*)\r[^\r]')
            result = None
            result = re.search(pattern, i['message'])
            if result:
                dic.update({'message_host': result.group(1)})
            else:
                dic.update({'message_host': ''})

            pattern = re.compile(r'.*告警等级:(.*)')
            result = None
            result = re.search(pattern, i['message'])
            if result:
                dic.update({'message_severity': result.group(1)})
            else:
                dic.update({'message_severity': ''})

            value_list_1.append(dic)

        if not value_list_1:
            return []
        return value_list_1

    def def_get_hostgroup_host(self, hostgroup_id_list):
        """
        获取主机组下所有主机名

        :param hostgroup_id_list: 主机组ID列表: ['10084']
        :return:  [{'hostid': '10084', 'host': 'Zabbix server'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.get",
            "params": {
                "output": ["host", "name"],
                "groupids": hostgroup_id_list,
                "selectInterfaces": ['ip'],
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_host_ip(self, host_id_list):
        """
        获取主机组下所有主机名

        :param hostgroup_id_list: 主机组ID列表: ['10084']
        :return: [{'ip': '127.0.0.1'}]}],
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "host.get",
            "params": {
                "output": ["interfaces"],
                "hostids": host_id_list,
                "selectInterfaces": ['ip'],
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_all_event(self, event_name, time_from, time_till):
        """
        该方法允许根据给定的参数检索事件。

        :param event_name: 动作名称 "Report problems to Zabbix administrators"
        :return:  [{'eventid': '15', 'hosts': [{'hostid': '10084', 'proxy_hostid': '0', 'host': 'Zabbix server', 'status': '0', 'lastaccess': '0', 'ipmi_authtype': '-1', 'ipmi_privi
lege': '2', 'ipmi_username': '', 'ipmi_password': '', 'maintenanceid': '0', 'maintenance_status': '0', 'maintenance_type': '0', 'maintenance_from': '0', 'name': 'Zabbix server', 'flags': '0',
'templateid': '0', 'description': '', 'tls_connect': '1', 'tls_accept': '1', 'tls_issuer': '', 'tls_subject': '', 'proxy_address': '', 'auto_compress': '1', 'custom_interfaces': '0', 'uuid': '
', 'inventory_mode': '-1'}]}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "event.get",
            "params": {
                "output": ['hosts'],
                "selectHosts": "extend",
                "sortfield": "clock",
                "sortorder": "DESC",
                "time_from": time_from,
                "time_till": time_till,
                "filter": {
                    "name": event_name
                },
            },
            "auth": self.authID,
            "id": 1,
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_all_priority_trigger(self, priority, status):
        """
        :param priority:
        :param status: 0已启用的触发器;1已禁用的触发器
        :return:
        """
        data = {
            "jsonrpc": "2.0",
            "method": "trigger.get",
            "params": {
                "output": [
                    "triggerid",
                    "description",
                    "priority"
                ],
                "filter": {
                    "status": status,
                    "priority": priority,
                },
                "active": '',
                "sortfield": "priority",
                "sortorder": "DESC"
            },
            "auth": self.authID,
            "id": 1
        }
        if status == 1:
            del data['params']['active']
        data = json.dumps(data)
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', '') != '':
            return response.get('result', '')
        else:
            return

    def def_stop_all_priority_trigger(self, priority):
        try:
            response = self.def_get_all_priority_trigger(priority, 0)
            for i in range(len(response)):
                self.def_update_all_priority_trigger(response[i]['triggerid'], 1)
                print(u"已成功禁用触发器: \033[;32m%s\033[0m 触发器ID为: \033[;32m%s\033[0m" %
                      (response[i]['description'], response[i]['triggerid']))
            return response
        except:
            return

    def def_start_all_priority_trigger(self, priority):
        try:
            response = self.def_get_all_priority_trigger(priority, 1)
            for i in range(len(response)):
                self.def_update_all_priority_trigger(response[i]['triggerid'], 0)
                print(u"已成功启动触发器: \033[;32m%s\033[0m 触发器ID为: \033[;32m%s\033[0m" %
                      (response[i]['description'], response[i]['triggerid']))
            return response
        except:
            return

    def def_update_all_priority_trigger(self, triggerid, status):
        """
        :param triggerid:
        :param status: 0启用触发器;1禁用触发器
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.update",
            "params": {
                "triggerid": triggerid,
                "status": status
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', '') != '':
            return response.get('result', '')
        else:
            return

    def def_get_all_unsupport_item(self):
        """
        :param state: 0标准的监控项;1不受支持的监控项
        :param status: 0已启用的监控项;1已禁用的监控项
        :return:
        """
        data = {
            "jsonrpc": "2.0",
            "method": "item.get",
            "params": {
                "output": [
                    "itemid",
                    "name",
                ],
                "filter": {
                    "state": 1,
                },
                "sortfield": "name",
                "sortorder": "DESC"
            },
            "auth": self.authID,
            "id": 1
        }
        data = json.dumps(data)
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', '') != '':
            return response.get('result', '')
        else:
            return

    def def_stop_all_unsupport_item(self):
        try:
            response = self.def_get_all_unsupport_item()
            for i in range(len(response)):
                self.def_update_all_unsupport_item(response[i]['itemid'], 1)
                print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 已成功禁用监控项: \033[;32m%s\033[0m 监控项ID为: \033[;32m%s\033[0m' % (len(response), i + 1, response[i]['name'], response[i]['itemid']))
            return response
        except:
            return

    def def_start_all_unsupport_item(self):
        try:
            response = self.def_get_all_unsupport_item()
            for i in range(len(response)):
                self.def_update_all_unsupport_item(response[i]['itemid'], 0)
                print(u'(\033[;34m%s\033[0m/\033[;34m%s\033[0m): -> 已成功启动监控项: \033[;32m%s\033[0m 监控项ID为: \033[;32m%s\033[0m' % (len(response), i + 1, response[i]['name'], response[i]['itemid']))
            return response
        except:
            return

    def def_update_all_unsupport_item(self, itemid, status):
        """
        :param itemid:
        :param status: 0启用触发器;1禁用触发器
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.update",
            "params": {
                "itemid": itemid,
                "status": status
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', '') != '':
            return response.get('result', '')
        else:
            return

    def def_check_zbx_version(self):
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "apiinfo.version",
            "params": [],
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_item_name(self, item_list):
        """
        更新监控项名称

        :param itemid: 监控项ID: ['10084']
        :param item_name: 监控项名称: 'Available memory'
        :return:
        """

        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "item.update",
            "params": item_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_itemprototype_name(self, item_list):
        """
        更新监控项名称

        :param itemid: 监控项ID: ['10084']
        :param item_name: 监控项名称: 'Available memory'
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "itemprototype.update",
            "params": item_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_itemprototype_tags_bytemplateid(self, tag_list):
        """
        更新监控项名称

        :param itemid: 监控项ID: ['10084']
        :param item_name: 监控项名称: 'Available memory'
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "itemprototype.update",
            "params": tag_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_trigger_name(self, item_list):
        """
        更新触发器名称

        :param triggerid: 触发器ID: ['10084']
        :param trigger_name: 触发器名称: 'Available memory'
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "trigger.update",
            "params": item_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_triggerprototype_name(self, item_list):
        """
        更新触发器名称

        :param triggerid: 触发器ID: ['10084']
        :param trigger_name: 触发器名称: 'Available memory'
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "triggerprototype.update",
            "params": item_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_triggerprototype_tag_bytemplateid(self, tag_list):
        """
        更新触发器名称

        :param triggerid: 触发器ID: ['10084']
        :param trigger_name: 触发器名称: 'Available memory'
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "triggerprototype.update",
            "params": tag_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_template_application_bytemplateid(self, application_list):
        """
        更新触发器名称

        :param triggerid: 触发器ID: ['10084']
        :param trigger_name: 触发器名称: 'Available memory'
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "application.update",
            "params": application_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_graph_name(self, item_list):
        """
        更新图表名称

        :param graphid: 图表ID: ['10084']
        :param graph_name: 图表名称: 'Available memory'
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "graph.update",
            "params": item_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_update_graphprototype_name(self, item_list):
        """
        更新图表名称

        :param graphid: 图表ID: ['10084']
        :param graph_name: 图表名称: 'Available memory'
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "graphprototype.update",
            "params": item_list,
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_get_en_list(self, excel_op, title_name, v_1, v_2, v_3, v_4):
        column_1_list = [f_1['host'] for f_1 in self.def_get_all_template_name()['result']]
        # 英译中
        translator = CusLanguageTrans()
        excel_op.def_creat_excel()
        excel_op.def_create_sheet(self.def_check_zbx_version()['result']),
        [excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))],
        def_get_name_list = []
        def_get_template_item_name_list = []
        en_list = []
        ch_list = []
        row_index = {"index": 2}
        executor = ThreadPoolExecutor(GV_CPU_COUNT)
        def_get_template_list = []

        def_get_template_list = []
        lv_result = None
        for lv_result in executor.map(self.def_get_template, column_1_list):
            if lv_result['tag'] is True:
                def_get_template_list.append(lv_result['result'][0]['templateid'])
            else:
                pass
        # [[def_get_template_list.append(f_2['templateid']) for f_2 in f_1['result']] for f_1 in executor.map(self.def_get_template, column_1_list)]

        def_get_template_item_name_list = []
        lv_result = None
        for lv_result in executor.map(v_3, def_get_template_list):
            if lv_result['tag'] is True:
                def_get_template_item_name_list.append(lv_result['result'])
            else:
                pass
        # [def_get_template_item_name_list.append(f_1['result']) for f_1 in executor.map(v_3, def_get_template_list)]
        [
            [
                [def_get_name_list.append(f_4) for f_4 in def_get_template_item_name_list[f_1]],
                [en_list.append(u'{0}'.format(def_get_name_list[f_3][v_1])) for f_3 in range(len(def_get_name_list))],
                [ch_list.append(u'{0}'.format(f_5).upper()) for f_5 in translator.def_trans("\n#| ".join(en_list)).split('#|')],
                [
                    [
                        # v_4(def_get_name_list[f_3][v_2], u'{0}'.format(ch_list[f_3])),
                        excel_op.def_set_cell_value(row_index['index'], 1, row_index['index'] - 1),
                        excel_op.def_set_cell_value(row_index['index'], 2, u'({0}/{1})'.format(len(def_get_template_list), f_1 + 1)),
                        excel_op.def_set_cell_value(row_index['index'], 3, u'({0}/{1})'.format(len(def_get_name_list), f_3 + 1)),
                        excel_op.def_set_cell_value(row_index['index'], 4, def_get_template_list[f_1]),
                        excel_op.def_set_cell_value(row_index['index'], 5, def_get_name_list[f_3][v_2]),
                        excel_op.def_set_cell_value(row_index['index'], 6, u'{0}'.format(def_get_name_list[f_3][v_1])),
                        excel_op.def_set_cell_value(row_index['index'], 7, u'{0}'.format(ch_list[f_3])),
                        excel_op.def_set_cell_value(row_index['index'], 8, u'{0}'.format(GV_ERROR_MESS['error'])),
                        GV_ERROR_MESS.update(error=''),
                        print(u'(\033[;34m{0}\033[0m/\033[;34m{1}\033[0m)|(\033[;34m{2}\033[0m/\033[;34m{3}\033[0m)\033[;32m{4}\033[0m: \033[;32m{5}\033[0m -> \033[;32m{6}\033[0m'.format(len(def_get_template_item_name_list), f_1 + 1, len(def_get_name_list), f_3 + 1, def_get_template_list[f_1], def_get_name_list[f_3][v_1], u'{0}'.format(ch_list[f_3]))),
                        row_index.update(index=row_index['index'] + 1),
                    ]
                    if ch_list != [''] and ch_list != ['[]'] and ch_list != ''
                    else
                    [
                        # excel_op.def_set_cell_value(row_index['index'], 1, row_index['index'] - 1),
                        # excel_op.def_set_cell_value(row_index['index'], 2, u'({0}/{1})'.format(len(def_get_template_list), f_1 + 1)),
                        # excel_op.def_set_cell_value(row_index['index'], 3, u'({0}/{1})'.format(len(def_get_name_list), f_3 + 1)),
                        # excel_op.def_set_cell_value(row_index['index'], 4, def_get_template_list[f_1]),
                        print(u'(\033[;34m{0}\033[0m/\033[;34m{1}\033[0m)|(\033[;34m{2}\033[0m/\033[;34m{3}\033[0m)\033[;32m{4}\033[0m'.format(len(def_get_template_item_name_list), f_1 + 1, len(def_get_name_list), f_3 + 1, def_get_template_list[f_1])),
                        # row_index.update(index=row_index['index'] + 1),
                    ]
                    for f_3 in range(len(ch_list))
                ],
                en_list.clear(),
                ch_list.clear(),
                def_get_name_list.clear()
            ] for f_1 in range(len(def_get_template_item_name_list))
        ]

        excel_op.def_save_create_xlsx(excel_op.sheet_name + '.xlsx')

    def def_set_zh_list(self, excel_op, v_2, v_1, v_4):
        excel_op.def_update_sheet_name(self.def_check_zbx_version()['result'])
        column_1_list = excel_op.def_get_col_value(5)
        del column_1_list[0]
        column_2_list = excel_op.def_get_col_value(7)
        del column_2_list[0]
        column_3_list = excel_op.def_get_col_value(2)
        del column_3_list[0]
        column_4_list = excel_op.def_get_col_value(3)
        del column_4_list[0]
        def_update_list = []
        executor = ThreadPoolExecutor(GV_CPU_COUNT)
        for f_1 in range(len(column_1_list)):
            if column_2_list[f_1].find("'tag':") != -1:
                column_2_list[f_1] = json.loads(column_2_list[f_1].replace("'", "\""))
            def_update_list.append({u"{v_2}".format(v_2=v_2): column_1_list[f_1], u"{v_1}".format(v_1=v_1): column_2_list[f_1]})
        # [
        #     [
        #         print(u'(\033[;34m{0}\033[0m/\033[;34m{1}\033[0m): \033[;32m{2}\033[0m \033[;32m{3}\033[0m \033[;32m{4}\033[0m'.format(len(column_1_list), f_1 + 1, column_3_list[f_1], column_4_list[f_1], def_update_list[f_1])),
        #         #for f_1 in range(len(v_4(def_update_list)))
        #
        #     ]
        #     for f_1 in range(len(def_update_list))
        # ]
        for number, prime in zip(def_update_list, executor.map(v_4, def_update_list)):
            pass
            # print('%s is prime: %s' % (number, prime))

    def def_get_dashboard(self):
        """
        获取主机组下所有主机名

        :return:  [{'hostid': '10084', 'name': 'Zabbix server'}]
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "dashboard.get",
            "params": {
                "output": "extend",
                "selectWidgets": "extend",
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        print(response)
        exit(1)
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_export_configuration(self, templateid):
        """

        :param templates_list: 模板列表 [10161, 10162]
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "configuration.export",
            "params": {
                "options": {
                    "templates": [templateid]
                },
                "format": "json"
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_import_configuration_5_0(self, source):
        """

        :param source: 模板数据源
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "configuration.import",
            "params": {
                "format": "json",
                "rules": {
                    "groups": {
                        "createMissing": True,
                    },
                    "templates": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "templateScreens": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "templateLinkage": {
                        "createMissing": True,
                    },
                    "applications": {
                        "createMissing": True,
                    },
                    "items": {
                        "createMissing": True,
                        "updateExisting": True,
                    },
                    "discoveryRules": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "triggers": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "graphs": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "httptests": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "valueMaps": {
                        "createMissing": True,
                        "updateExisting": False
                    },
                },
                "source": source,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_import_configuration_6_0(self, source):
        """

        :param source: 模板数据源
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "configuration.import",
            "params": {
                "format": "json",
                "rules": {
                    "groups": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "templates": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "valueMaps": {
                        "createMissing": True,
                        "updateExisting": False
                    },
                    "templateDashboards": {
                        "createMissing": True,
                        "updateExisting": True,
                    },
                    "templateLinkage": {
                        "createMissing": True,
                    },
                    "items": {
                        "createMissing": True,
                        "updateExisting": True,
                    },
                    "discoveryRules": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "triggers": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "graphs": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "httptests": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                },
                "source": source,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_import_configuration_6_4(self, source):
        """

        :param source: 模板数据源
        :return:
        """
        data = json.dumps({
            "jsonrpc": "2.0",
            "method": "configuration.import",
            "params": {
                "format": "json",
                "rules": {
                    "template_groups": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "host_groups": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "templates": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "valueMaps": {
                        "createMissing": True,
                        "updateExisting": False
                    },
                    "templateDashboards": {
                        "createMissing": True,
                        "updateExisting": True,
                    },
                    "templateLinkage": {
                        "createMissing": True,
                    },
                    "items": {
                        "createMissing": True,
                        "updateExisting": True,
                    },
                    "discoveryRules": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "triggers": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "graphs": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                    "httptests": {
                        "createMissing": True,
                        "updateExisting": True
                    },
                },
                "source": source,
            },
            "auth": self.authID,
            "id": 1
        })
        request = self.session.post(url=self.url, headers=self.header, data=data, verify=False)
        response = request.json()
        if response.get('result', ''):
            return {'tag': True, 'result': response['result']}
        if response.get('error', ''):
            GV_ERROR_MESS['error'] = response['error']
            # print(u"错误: \033[;31m{0}\033[0m \033[;31m{1}\033[0m".format(inspect.stack()[0][2], response['error']))
            return {'tag': False, 'result': response['error']}
        else:
            return {'tag': False, 'result': response}

    def def_timecovert(self, stringtime):
        timeArray = time.strptime(stringtime, "%Y-%m-%d %H:%M:%S")
        timeStamp = int(time.mktime(timeArray))
        return timeStamp

    def def_timeCovertIntToYMD(self, inttime):
        timeArray = time.localtime(inttime)
        timeStamp = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        return timeStamp

    def def_convert_numeric_to_text(self, value):
        if value == '0':
            self.textValue = '消息未发送'
        elif value == '1':
            self.textValue = '消息已发送'
        elif value == '2':
            self.textValue = '经多次重试后失败'
        elif value == '3':
            self.textValue = '告警管理员尚未处理的新告警'
        elif value == 'Not classified':
            self.textValue = '未分类'
        elif value == 'Information':
            self.textValue = '信息'
        elif value == 'Warning':
            self.textValue = '警告'
        elif value == 'Average':
            self.textValue = '一般严重'
        elif value == 'High':
            self.textValue = '严重'
        elif value == 'Disaster':
            self.textValue = '灾难'
        else:
            self.textValue = value
        return self.textValue

    def def_convert_numeric_to_text_problem(self, value):
        if value == '0':
            self.textValue = '恢复'
        elif value == '1':
            self.textValue = '故障'
        else:
            self.textValue = value
        return self.textValue

class CusExcelOp(object):
    def __init__(self):
        """
        excel_op = ExcelOp(file='zabbix_api.xlsx', index=16)
        column_1_list = excel_op.def_get_col_value(1)
        del column_1_list[0]
        title_name = ['主机名']
        excel_op.def_create_sheet(excel_op.sheet_name)
        [excel_op.def_set_cell_value(1, i + 1, title_name[i]) for i in range(len(title_name))]
        [excel_op.def_set_cell_value(i + 2, 1, column_1_list[i]) for i in range(len(column_1_list))]
        excel_op.def_save_create_xlsx(excel_op.sheet_name)
        """
        self.file = None
        self.wb_object = None
        self.sheet_name = None
        self.ws_object = None

        self.create_file = None
        self.create_ws_object = None
        self.create_wb_object = None

    def def_load_excel(self, file, index):
        if index == 0:
            print("index must at least 1")
            exit(1)
        self.file = file
        self.wb_object = openpyxl.load_workbook(self.file)
        sheet_name_list = self.wb_object.sheetnames
        self.sheet_name = sheet_name_list[index - 1]
        self.ws_object = self.wb_object[self.sheet_name]

    def def_creat_excel(self):
        self.create_wb_object = openpyxl.Workbook()

    def def_create_sheet(self, sheet_name):
        """
        创建sheet页

        :param sheet_name: sheet名称: '页面1'
        :return: None
        """
        self.create_ws_object = self.create_wb_object.create_sheet(sheet_name)

    def def_update_sheet_name(self, sheet_name):
        self.sheet_name = sheet_name
        try:
            self.ws_object = self.wb_object[sheet_name]
        except Exception as e:
            print(u"数据异常当前Zabbix版本不支持汉化: %s" % e)
            exit(1)
        return None

    # 获取表格的总行数和总列数
    def def_get_row_clo_num(self):
        """
        :return: {"rows": rows, "columns": columns}
        """
        rows = self.ws_object.max_row
        columns = self.ws_object.max_column
        return {"rows": rows, "columns": columns}

    # 获取某个单元格的值
    def def_get_cell_value(self, row, column):
        """
        获取单元格的值

        :param row: 行号
        :param column: 列号
        :return: 返回字符串
        """
        cell_value = self.ws_object.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def def_get_col_value(self, column):
        """
        根据列号返回一列数据

        :param column: 列号: 3
        :return: 表格数据: ['a','b','c']
        """
        # rows = self.ws_object.max_row
        # cell_value = [self.ws_object.cell(row=i, column=column).value if self.ws_object.cell(row=i, column=column).value is not None else (
        #     print('第 %s 行错误! 原因: %s' % (i, u'错误! 列表存在空行')), exit(1)) for i in range(1, rows + 1)]
        v_dic_num = {"row": 1}
        v_flag = True
        v_list_row = []
        while v_flag:
            v_cell_value = self.ws_object.cell(row=v_dic_num['row'], column=column).value
            if v_cell_value is not None:
                v_list_row.append(v_cell_value)
                v_dic_num.update(row=v_dic_num['row'] + 1)
            else:
                v_flag = False
        return v_list_row

        # return cell_value

    def def_get_creat_col_value(self, column):
        """
        根据列号返回一列数据

        :param column: 列号: 3
        :return: 表格数据: ['a','b','c']
        """
        # rows = self.ws_object.max_row
        # cell_value = [self.ws_object.cell(row=i, column=column).value if self.ws_object.cell(row=i, column=column).value is not None else (
        #     print('第 %s 行错误! 原因: %s' % (i, u'错误! 列表存在空行')), exit(1)) for i in range(1, rows + 1)]
        v_dic_num = {"row": 1}
        v_flag = True
        v_list_row = []
        while v_flag:
            v_cell_value = self.create_ws_object.cell(row=v_dic_num['row'], column=column).value
            if v_cell_value is not None:
                v_list_row.append(v_cell_value)
                v_dic_num.update(row=v_dic_num['row'] + 1)
            else:
                v_flag = False
        return v_list_row

        # return cell_value

    def def_get_col_value_base_20221003(self, column):
        """
        根据列号返回一列数据

        :param column: 列号: 3
        :return: 表格数据: ['a','b','c']
        """
        rows = self.ws_object.max_row
        cell_value = [self.ws_object.cell(row=i, column=column).value for i in range(1, rows + 1) if self.ws_object.cell(row=i, column=column).value is not None]
        return cell_value

    # 获取某行所有值
    def def_get_row_value(self, row):
        """
        根据行号返回一行数据

        :param row: 行号: 1
        :return: 表格数据: ['a','b','c']
        """
        columns = self.ws_object.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws_object.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def def_load_create_sheet(self, sheet_name):
        """
        加载sheet页

        :param sheet_name: sheet名称: '页面1'
        :return: None
        """
        self.create_wb_object = openpyxl.load_workbook(self.create_file)
        self.create_ws_object = self.create_wb_object[sheet_name]

    def def_save_create_xlsx(self, xlsx_name):
        """
        保存为:excel名称_时间戳.xlsx

        :param xlsx_name: excel名称: 表格1
        :return: None
        """
        del self.create_wb_object['Sheet']
        self.create_wb_object.save(xlsx_name)

    # 设置某个单元格的值
    def def_set_cell_value(self, row, colunm, cellvalue):
        """
        设置某个单元格的值

        :param row: 行号: 1
        :param colunm: 列号: 3
        :param cellvalue: 单元格值: '成功'
        :return: None
        """
        try:
            self.create_ws_object.cell(row=row, column=colunm).value = cellvalue
        except Exception as ee:
            self.create_ws_object.cell(row=row, column=colunm).value = u'{0}'.format(ee)


class CusAssembleHeaderException(Exception):
    def __init__(self, msg):
        self.message = msg


class CusUrl:
    def __init__(self, host, path, schema):
        self.host = host
        self.path = path
        self.schema = schema
        pass


class CusLanguageTransOne(object):
    def __init__(self):
        # 应用ID（到控制台获取）
        self.APPID = "668c78a6"
        # 接口APISercet（到控制台机器翻译服务页面获取）
        self.Secret = "ZGVlMGM5OThmZDljYWVkYTFjNGZlYzll"
        # 接口APIKey（到控制台机器翻译服务页面获取）
        self.APIKey = "066246864d8bca0557c7b4fa9e128cde"
        # 术语资源唯一标识，请根据控制台定义的RES_ID替换具体值，如不需术语可以不用传递此参数
        self.RES_ID = "epktzy16ho"
        self.url = "https://itrans.xf-yun.com/v1/its"
        # 翻译原文本内容

    def assemble_ws_auth_url(self, requset_url, method="POST", api_key="", api_secret=""):
        u = self.parse_url(requset_url)
        host = u.host
        path = u.path
        now = datetime.now()
        date = format_date_time(mktime(now.timetuple()))
        # print(date)
        # date = "Thu, 12 Dec 2019 01:57:27 GMT"
        signature_origin = "host: {}\ndate: {}\n{} {} HTTP/1.1".format(host, date, method, path)
        # print(signature_origin)
        signature_sha = hmac.new(api_secret.encode('utf-8'), signature_origin.encode('utf-8'),
                                 digestmod=hashlib.sha256).digest()
        signature_sha = base64.b64encode(signature_sha).decode(encoding='utf-8')
        authorization_origin = "api_key=\"%s\", algorithm=\"%s\", headers=\"%s\", signature=\"%s\"" % (
            api_key, "hmac-sha256", "host date request-line", signature_sha)
        authorization = base64.b64encode(authorization_origin.encode('utf-8')).decode(encoding='utf-8')
        # print(authorization_origin)
        values = {
            "host": host,
            "date": date,
            "authorization": authorization
        }

        return requset_url + "?" + urlencode(values)

    def parse_url(self, requset_url):
        stidx = requset_url.index("://")
        host = requset_url[stidx + 3:]
        schema = requset_url[:stidx + 3]
        edidx = host.index("/")
        if edidx <= 0:
            raise CusAssembleHeaderException("invalid request url:" + requset_url)
        path = host[edidx:]
        host = host[:edidx]
        u = CusUrl(host, path, schema)
        return u

    def sha256base64(self, data):
        sha256 = hashlib.sha256()
        sha256.update(data)
        digest = base64.b64encode(sha256.digest()).decode(encoding='utf-8')
        return digest

    def def_trans(self, text):
        if text == '':
            return ''
        request_url = self.assemble_ws_auth_url(self.url, "POST", self.APIKey, self.Secret)
        headers = {'content-type': "application/json", 'host': 'itrans.xf-yun.com', 'app_id': self.APPID}
        # print(request_url)
        # print(text)
        v_int_ifcut = int(len(text) / 5000)
        v_list_text = text.split('\n#| ')
        v_list_cut_list = []
        v_list_cn = []
        v_int_cut_num = int(len(v_list_text) / (v_int_ifcut + 1))
        # print(v_int_ifcut)
        # print(v_int_cut_num)
        # print(len(v_list_text))
        for i in range(len(v_list_text)):
            v_list_cut_list.append(v_list_text[i])
            if i == v_int_cut_num - 1:
                text = "\n#| ".join(v_list_cut_list)
                # print(text)
                body = {
                    "header": {
                        "app_id": self.APPID,
                        "status": 3,
                        "res_id": self.RES_ID
                    },
                    "parameter": {
                        "its": {
                            "from": "en",
                            "to": "cn",
                            "result": {}
                        }
                    },
                    "payload": {
                        "input_data": {
                            "encoding": "utf8",
                            "status": 3,
                            "text": base64.b64encode(text.encode("utf-8")).decode('utf-8')
                        }
                    }
                }
                response = requests.post(request_url, data=json.dumps(body), headers=headers)
                tempResult = json.loads(response.content.decode())
                v_json = json.loads(base64.b64decode(tempResult['payload']['result']['text']).decode())
                # print(base64.b64decode(tempResult['payload']['result']['text']).decode())
                v_list_cn.append(v_json['trans_result']['dst'].upper())
                v_list_cut_list.clear()
                v_int_cut_num = (v_int_ifcut + 1) * v_int_cut_num + 1
        return "\n#| ".join(v_list_cn)


class CusLanguageTransTwo(object):
    def __init__(self, host="itrans.xfyun.cn"):
        # 应用ID（到控制台获取）
        self.APPID = "668c78a6"
        # 接口APISercet（到控制台机器翻译服务页面获取）
        self.Secret = "ZGVlMGM5OThmZDljYWVkYTFjNGZlYzll"
        # 接口APIKey（到控制台机器翻译服务页面获取）
        self.APIKey = "066246864d8bca0557c7b4fa9e128cde"

        # 以下为POST请求
        self.Host = host
        self.RequestUri = "/v2/its"
        # 设置url
        # print(host)
        self.url = "https://" + host + self.RequestUri
        self.HttpMethod = "POST"
        self.Algorithm = "hmac-sha256"
        self.HttpProto = "HTTP/1.1"

        # 设置当前时间
        curTime_utc = datetime.datetime.utcnow()
        self.Date = self.httpdate(curTime_utc)
        # 设置业务参数
        # 语种列表参数值请参照接口文档：https://www.xfyun.cn/doc/nlp/xftrans/API.html
        self.Text = "你好吗"
        self.BusinessArgs = {
            "from": "en",
            "to": "cn",
        }

    def hashlib_256(self, res):
        m = hashlib.sha256(bytes(res.encode(encoding='utf-8'))).digest()
        result = "SHA-256=" + base64.b64encode(m).decode(encoding='utf-8')
        return result

    def httpdate(self, dt):
        """
        Return a string representation of a date according to RFC 1123
        (HTTP/1.1).

        The supplied date must be in UTC.

        """
        weekday = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][dt.weekday()]
        month = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep",
                 "Oct", "Nov", "Dec"][dt.month - 1]
        return "%s, %02d %s %04d %02d:%02d:%02d GMT" % (weekday, dt.day, month,
                                                        dt.year, dt.hour, dt.minute, dt.second)

    def generateSignature(self, digest):
        signatureStr = "host: " + self.Host + "\n"
        signatureStr += "date: " + self.Date + "\n"
        signatureStr += self.HttpMethod + " " + self.RequestUri \
                        + " " + self.HttpProto + "\n"
        signatureStr += "digest: " + digest
        signature = hmac.new(bytes(self.Secret.encode(encoding='utf-8')),
                             bytes(signatureStr.encode(encoding='utf-8')),
                             digestmod=hashlib.sha256).digest()
        result = base64.b64encode(signature)
        return result.decode(encoding='utf-8')

    def init_header(self, data):
        digest = self.hashlib_256(data)
        # print(digest)
        sign = self.generateSignature(digest)
        authHeader = 'api_key="%s", algorithm="%s", ' \
                     'headers="host date request-line digest", ' \
                     'signature="%s"' \
                     % (self.APIKey, self.Algorithm, sign)
        # print(authHeader)
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "Method": "POST",
            "Host": self.Host,
            "Date": self.Date,
            "Digest": digest,
            "Authorization": authHeader
        }
        return headers

    def get_body(self, text):
        content = str(base64.b64encode(text.encode('utf-8')), 'utf-8')
        postdata = {
            "common": {"app_id": self.APPID},
            "business": self.BusinessArgs,
            "data": {
                "text": content,
            }
        }
        body = json.dumps(postdata)
        # print(body)
        return body

    def def_trans(self, text):
        if self.APPID == '' or self.APIKey == '' or self.Secret == '':
            print('Appid 或APIKey 或APISecret 为空！请打开demo代码，填写相关信息。')
        else:
            code = 0
            body = self.get_body(text)
            headers = self.init_header(body)
            # print(self.url)
            if text == '':
                return ''
            response = requests.post(self.url, data=body, headers=headers, timeout=8)
            status_code = response.status_code
            # print(response.content)
            if status_code != 200:
                # 鉴权失败
                print("Http请求失败，状态码：" + str(status_code) + "，错误信息：" + response.text)
                print("请根据错误信息检查代码，接口文档：https://www.xfyun.cn/doc/nlp/xftrans/API.html")
            else:
                # 鉴权成功
                # respData = json.loads(response.text)
                respData = response.json()
                if respData.get('data', ''):
                    # print(respData['data']['result']['trans_result']['dst'])
                    return respData['data']['result']['trans_result']['dst']
                else:
                    return respData['desc']
                # 以下仅用于调试
                # code = str(respData["code"])
                # if code != '0':
                #     print("请前往https://www.xfyun.cn/document/error-code?code=" + code + "查询解决办法")


class CusLanguageTrans(object):
    def __init__(self):
        # 应用ID（到控制台获取）
        self.APPID = "668c78a6"
        # 接口APISercet（到控制台机器翻译服务页面获取）
        self.Secret = "ZGVlMGM5OThmZDljYWVkYTFjNGZlYzll"
        # 接口APIKey（到控制台机器翻译服务页面获取）
        self.APIKey = "066246864d8bca0557c7b4fa9e128cde"
        # 术语资源唯一标识，请根据控制台定义的RES_ID替换具体值，如不需术语可以不用传递此参数
        self.RES_ID = "epktzy16ho"
        self.url = "https://fanyi.xfyun.cn/api-tran/trans/its"
        self.cookies = {
            '_wafuid': '42179584',
            'JSESSIONID': '5C64510B0B5D7E4C6F8ADF6E3D9EEF19',
            '_gcl_au': '1.1.534731207.1666665336',
            'Hm_lvt_fe740601c79b0c00b6d5458d146aa5ef': '1666665336',
            'gr_user_id': '4d219912-8592-41f1-9405-a3f3b77165db',
            '8473744dbcf62d60_gr_session_id': '20b029c9-36ef-4685-9cbb-ac525a8d18e3',
            '8473744dbcf62d60_gr_session_id_20b029c9-36ef-4685-9cbb-ac525a8d18e3': 'true',
            '_ga': 'GA1.2.75942607.1666665337',
            '_gid': 'GA1.2.1536677545.1666665337',
            'Hm_lpvt_fe740601c79b0c00b6d5458d146aa5ef': '1666665358',
            'di_c_mti': 'f634cf0f-f6c0-697e-a6c2-b130d51ed2ea',
            'Hm_lvt_46f7583efda1cc689658545dca371747': '1666665365',
            'ssoSessionId': '06c679b6-ee78-4763-aa91-2b4ace6a0810',
            'account_id': '14795786860',
            'Hm_lpvt_46f7583efda1cc689658545dca371747': '1666665395',
        }

        self.headers = {
            'Host': 'fanyi.xfyun.cn',
            'Accept': 'application/json, text/plain, */*',
            'X-Requested-With': 'XMLHttpRequest',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36',
            'Origin': 'https://fanyi.xfyun.cn',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': 'https://fanyi.xfyun.cn/console/trans/text',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            # Requests sorts cookies= alphabetically
            # 'Cookie': '_wafuid=42179584; JSESSIONID=5C64510B0B5D7E4C6F8ADF6E3D9EEF19; _gcl_au=1.1.534731207.1666665336; Hm_lvt_fe740601c79b0c00b6d5458d146aa5ef=1666665336; gr_user_id=4d219912-8592-41f1-9405-a3f3b77165db; 8473744dbcf62d60_gr_session_id=20b029c9-36ef-4685-9cbb-ac525a8d18e3; 8473744dbcf62d60_gr_session_id_20b029c9-36ef-4685-9cbb-ac525a8d18e3=true; _ga=GA1.2.75942607.1666665337; _gid=GA1.2.1536677545.1666665337; Hm_lpvt_fe740601c79b0c00b6d5458d146aa5ef=1666665358; di_c_mti=f634cf0f-f6c0-697e-a6c2-b130d51ed2ea; Hm_lvt_46f7583efda1cc689658545dca371747=1666665365; ssoSessionId=06c679b6-ee78-4763-aa91-2b4ace6a0810; account_id=14795786860; Hm_lpvt_46f7583efda1cc689658545dca371747=1666665395',
        }
        self.session = requests.Session()
        self.session.mount(self.url, requests.adapters.HTTPAdapter(max_retries=3))
        # 翻译原文本内容

    def def_trans(self, text):
        return text.replace('\n', '')

        if text == '':
            return ''

        v_int_ifcut = int(len(text) / 2000)
        v_list_text = text.split('\n#| ')
        v_list_cut_list = []
        v_list_cn = []
        v_int_cut_num = int(len(v_list_text) / (v_int_ifcut + 1))
        # print(v_int_ifcut)
        # print(v_int_cut_num)
        # print(len(v_list_text))
        for i in range(len(v_list_text)):
            v_list_cut_list.append(v_list_text[i])
            if i == v_int_cut_num - 1:
                text = "\n#| ".join(v_list_cut_list)
                # print(text)
                data = {
                    'from': 'en',
                    'to': 'cn',
                    'text': text,
                }
                time.sleep(1)
                # print(text)
                request = self.session.post(url=self.url, cookies=self.cookies, headers=self.headers, data=data)
                response = request.json()
                if response.get('data', ''):
                    v_json = json.loads(response['data'])['trans_result']['dst'].upper()
                    if v_json == '':
                        GV_ERROR_MESS = u'{0}'.format('翻译错误未执行完毕')
                    v_list_cn.append(v_json)
                    v_list_cut_list.clear()
                    v_int_cut_num = (v_int_ifcut + 1) * v_int_cut_num + 1
        return "\n#| ".join(v_list_cn)


class CusMyThreadSendDir(threading.Thread):
    def __init__(self, cur_num, total_num, ssh_ip, port, pwd):
        super(CusMyThreadSendDir, self).__init__()
        self.cur_num = cur_num
        self.total_num = total_num
        self.ssh_ip = ssh_ip
        self.port = int(port)
        self.username = "root"
        self.pwd = pwd
        self.transport = paramiko.Transport(sock=(self.ssh_ip, self.port))
        self.transport.connect(username=self.username, password=self.pwd)
        self.sftp = paramiko.SFTPClient.from_transport(self.transport)

    def run(self):
        def create_remote_dir(dir):
            for item in dir:
                try:
                    self.sftp.stat(item)
                    pass
                except IOError:
                    # print("Create a new directory: ", item)
                    self.sftp.mkdir(item)

        def for_dir():
            for res in path:
                if os.path.isdir(res):
                    local_dir_path.append(res)
            remote_dir_path.append(des)

        def for_zdir():
            des_src_dir.append(remote_dir_path[1])
            des_src_dir_list = des_src_dir[0].split("/")
            des_dir_list = des_src_dir_list[1:]
            c = ""
            remote_des_src_path = []
            for item in des_dir_list:
                c += "/" + item
                remote_des_src_path.append(c)
            create_remote_dir(remote_des_src_path)
            create_remote_dir(remote_dir_path)
            for res in path:
                if os.path.isfile(res):
                    local_file_path.append(res)

        src = "./senddir/02_sjzx/"
        des = "/tmp/zabbix_dir/"
        sep = "/"
        path = []
        local_dir_path = []
        local_file_path = []
        remote_dir_path = []
        remote_file_path = []
        des_src_dir = []
        for i in os.listdir(src):
            path.append(src + sep + i)
        for n in path:
            if os.path.isdir(n) and os.listdir(n):
                for i in os.listdir(n):
                    path.append(n + sep + i)
        local_dir_path.append(src)
        local_dir = src.split("/")
        local_dir_first = local_dir[0:-1]
        global a
        if len(local_dir_first) == 0:
            for_dir()
            for res in local_dir_path:
                remote_dir_path.append(des + "/" + res)
            for_zdir()
            for res in local_file_path:
                remote_file_path.append(des + "/" + res)
        else:
            if len(local_dir_first) == 1:
                dir_join = "/".join(local_dir_first)
                a = dir_join
            else:
                dir_join = "/".join(local_dir_first)
                a = dir_join + "/"
            for res in path:
                if os.path.isdir(res):
                    local_dir_path.append(res)
            remote_dir_path.append(des)
            b = [item.split(a)[-1] for item in local_dir_path]
            for res in b:
                if len(local_dir_first) == 1:
                    remote_dir_path.append(des + res)
                else:
                    remote_dir_path.append(des + "/" + res)
            for_zdir()
            d = [item.split(a)[-1] for item in local_file_path]
            for res in d:
                if len(local_dir_first) == 1:
                    remote_file_path.append(des + res)
                else:
                    remote_file_path.append(des + "/" + res)
        time_start = time.time()
        local_file_num = len(local_file_path)
        try:
            for i in range(local_file_num):
                self.sftp.put(local_file_path[i], remote_file_path[i])
            total_time = time.time() - time_start
            self.transport.close()
            print("""Send Successful, total time: %s""" % str(total_time))
        except Exception as e:
            print("""Send failed, %s""" % e, end='')


class CusMyThreadCfgZabbixAgent(threading.Thread):
    def __init__(self, cur_num, total_num, ssh_ip, port, pwd, zabbix_server_ip):
        super(CusMyThreadCfgZabbixAgent, self).__init__()
        self.cur_num = cur_num
        self.total_num = total_num
        self.ssh_ip = ssh_ip
        self.port = int(port)
        self.username = "root"
        self.pwd = pwd
        self.zabbix_server_ip = zabbix_server_ip
        self.ssh = paramiko.SSHClient()
        self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        # self.cmd = """
        # rm -rf /tmp/zabbix_dir
        # """

        self.cmd1 = """
        sh /tmp/zabbix_dir/install.sh --cleardir
        """

        self.cmd2 = """
        sh /tmp/zabbix_dir/install.sh --senddir
        """

        self.cmd3 = """
        sh /tmp/zabbix_dir/install.sh --install
        """

    def run(self):
        try:
            # self.ssh.connect(hostname=self.ssh_ip, port=self.port, username=self.username, password=self.pwd)
            # stdin, stdout, stderr = self.ssh.exec_command(self.cmd)
            # res, err = stdout.read(), stderr.read()
            # result = res if err else res
            # self.ssh.close()

            self.ssh.connect(hostname=self.ssh_ip, port=self.port, username=self.username, password=self.pwd)
            stdin, stdout, stderr = self.ssh.exec_command(self.cmd1)
            res, err = stdout.read(), stderr.read()
            result = res if err else res
            self.ssh.close()
            print("""%s/%s: %s host zabbix-agent ok, err: %s""" % (self.cur_num, self.total_num, self.ssh_ip, err))
            self.ssh.connect(hostname=self.ssh_ip, port=self.port, username=self.username, password=self.pwd)
            stdin, stdout, stderr = self.ssh.exec_command(self.cmd2)
            res, err = stdout.read(), stderr.read()
            result = res if err else res
            self.ssh.close()
            print("""%s/%s: %s host zabbix-agent ok, err: %s""" % (self.cur_num, self.total_num, self.ssh_ip, err))
            self.ssh.connect(hostname=self.ssh_ip, port=self.port, username=self.username, password=self.pwd)
            stdin, stdout, stderr = self.ssh.exec_command(self.cmd3)
            res, err = stdout.read(), stderr.read()
            result = res if err else res
            self.ssh.close()
            print("""%s/%s: %s host zabbix-agent ok, err: %s""" % (self.cur_num, self.total_num, self.ssh_ip, err))
        except Exception as e:
            print("""%s/%s: %s""" % (self.cur_num, self.total_num, e))


class CusMyThreadCfgSj(threading.Thread):
    def __init__(self, cur_num, total_num, s_dic):
        self.session = requests.Session()
        super(CusMyThreadCfgSj, self).__init__()
        self.cur_num = cur_num
        self.total_num = total_num
        self.s_dic = {}
        self.s_dic.update(s_dic)
        self.username = "root"
        self.ssh = paramiko.SSHClient()
        self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        # 删除文件
        self.cmd1 = """
        rm -rf /tmp/zabbix_dir 1>&2
        """
        # 配置网卡
        self.cmd2 = """
        sh /tmp/zabbix_dir/cfgip.sh {0} {1} {2} {3} {4} {5} {6} {7} 1>&2
        """.format(self.s_dic['s_eth'], self.s_dic['s_ip'], self.s_dic['s_newmask'], self.s_dic['s_gateway'],
                   self.s_dic['s_dns1'], self.s_dic['s_dns2'], self.s_dic['s_host'], self.s_dic['s_app'])
        # 配置主机名并重启
        self.cmd3 = """
        sh /tmp/zabbix_dir/cfghost.sh {0}
        """.format(self.s_dic['s_host'])
        # 配置免密登录需要修改cfgssh.sh密码
        self.cmd4 = """
        sh /tmp/zabbix_dir/cfgssh.sh 1>&2
        """
        # 配置JDK
        self.cmd5 = """
        sh /tmp/zabbix_dir/cfgjdk.sh
        """
        # 配置mysql将需要配置mysql的excel行放到第一行，第二行置空
        self.cmd6 = """
        sh /tmp/zabbix_dir/cfgmysql.sh
        """
        # 配置redis
        self.cmd7 = """
        sh /tmp/zabbix_dir/cfgredis.sh
        """
        # 配置zookeeper将需要配置zookeeper的excel行放到第一行，第二行置空，需要配置JDK
        self.cmd8 = """
        sh /tmp/zabbix_dir/cfgzookeeper.sh
        """
        # 配置kafka将需要配置kafka的excel行放到第一行，第二行置空，需要配置JDK
        self.cmd9 = """
        sh /tmp/zabbix_dir/cfgkafka.sh
        """
        # 配置elasticsearch将需要配置elasticsearch的excel行放到第一行，第二行置空，需要配置JDK11
        self.cmd10 = """
        sh /tmp/zabbix_dir/cfgelasticsearch.sh
        """
        # 配置kibana将需要配置kibana的excel行放到第一行，第二行置空，需要配置elasticsearch
        self.cmd11 = """
        sh /tmp/zabbix_dir/cfgkibana.sh
        """
        # 导入es数据
        # list_a_ = []
        # with open('senddir/ip/kibana.txt', 'r') as a_:
        #     [list_a_.append(line_) for line_ in a_]
        # # 导入es数据库
        # self.def_create_index(list_a_[0].split(' ')[1])
        # 配置naco
        self.cmd12 = """
        sh /tmp/zabbix_dir/cfgnacos.sh
        """
        # self.authID = ""
        # # naco的web页面操作，无需此操作
        # list_a_ = []
        # with open('senddir/ip/nacos.txt', 'r') as a_:
        #     [list_a_.append(line_) for line_ in a_]
        # # 创建命名空间，无需此操作
        # self.def_login(list_a_[0].split(' ')[0])
        # self.def_nacos_conf(list_a_[0].split(' ')[0])
        self.cmd13 = """
        sh /tmp/zabbix_dir/cfgopenresty.sh
        """
        self.cmd14 = """
        sh /tmp/zabbix_dir/cfgspark.sh
        """
        self.cmd15 = """
        sh /tmp/zabbix_dir/install.sh
        """

    def run(self):
        try:
            self.ssh.connect(hostname=self.s_dic['ssh_ip'], port=self.s_dic['port'], username=self.username, password=self.s_dic['pwd'])
            # 根据需要修改执行的命令 cmd2 timeout=30
            stdin, stdout, stderr = self.ssh.exec_command(self.cmd15)
            res, err = stdout.read(), stderr.read()
            result = res if err else res
            print("""%s/%s: %s host ok, err: %s""" % (self.cur_num, self.total_num, self.s_dic['ssh_ip'], err))
            self.ssh.close()

        except Exception as e:
            print("""%s/%s: %s""" % (self.cur_num, self.total_num, e))

    def get_filelist(self, dir, Filelist, suffix):
        if os.path.isfile(dir):
            if dir.endswith(suffix):
                Filelist.append(dir)
        elif os.path.isdir(dir):
            for s in os.listdir(dir):
                newDir = os.path.join(dir, s)
                self.get_filelist(newDir, Filelist, suffix)
        return Filelist

    def def_create_index(self, host_ip_):
        headers = {
            'Host': '{0}:5601'.format(host_ip_),
            'Accept': 'text/plain, */*; q=0.01',
            'kbn-version': '7.2.0',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36',
            # Already added when you pass json= but not when you pass data=
            # 'Content-Type': 'application/json',
            'Origin': 'http://{0}:5601'.format(host_ip_),
            'Referer': 'http://{0}:5601/app/kibana'.format(host_ip_),
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }

        file_list = []
        self.get_filelist('senddir\\patch\\json', file_list, '')
        for i_ in range(len(file_list)):
            line_data = ""
            with open(file_list[i_], 'r') as a_:
                next(a_)
                for line_ in a_:
                    line_data = line_data + ''.join(line_)
                json_data = json.loads(line_data)
                fileName_ = os.path.basename(file_list[i_])
                params = {
                    'path': fileName_,
                    'method': 'PUT',
                }
            self.session.mount('http://{0}:5601/api/console/proxy'.format(host_ip_), requests.adapters.HTTPAdapter(max_retries=3))
            request = self.session.post(url='http://{0}:5601/api/console/proxy'.format(host_ip_), params=params, headers=headers, json=json_data, verify=False)
            response = request.json()
            if response.get('acknowledged', '') != '':
                print(u"创建索引: \033[;32m%s\033[0m 成功! 返回值为: \033[;32m%s\033[0m" % (
                    fileName_, response['acknowledged']))
            elif response.get('error', '') != '':
                print(u"创建索引: \033[;31m%s\033[0m 失败! 原因: \033[;31m%s\033[0m" % (
                    fileName_, response['error']['reason']))

    def def_login(self, host_ip_):
        headers = {
            'Host': '{0}:38848'.format(host_ip_),
            'Accept': 'application/json, text/plain, */*',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36',
            'Origin': 'http://{0}:38848'.format(host_ip_),
            'Referer': 'http://{0}:38848/nacos/'.format(host_ip_),
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }

        data = {
            'username': 'nacos',
            'password': 'nacos',
        }

        try:
            request = self.session.post(url='http://{0}:38848/nacos/v1/auth/users/login'.format(host_ip_), headers=headers, data=data, verify=False)
            response = request.json()
            if response.get('accessToken', '') != '':
                self.authID = response['accessToken']
            elif response.get('error', '') != '':
                print(u"用户认证失败请检查! 原因: \033[;31m%s\033[0m" % (response))
                sys.exit(1)
        except Exception as ee:
            print(u"地址请求失败请检查! 原因: \033[;31m%s\033[0m" % ee)
            sys.exit(1)

    def def_nacos_conf(self, host_ip_):
        headers = {
            'Host': '{0}:38848'.format(host_ip_),
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'X-Requested-With': 'XMLHttpRequest',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36',
            'Origin': 'http://{0}:38848'.format(host_ip_),
            'Referer': 'http://{0}:38848/nacos/'.format(host_ip_),
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }

        data = {
            'customNamespaceId': 'sjzx',
            'namespaceName': 'sjzx',
            'namespaceDesc': 'sjzx',
            'namespaceId': '',
        }
        self.session.mount('http://{0}:38848/nacos/v1/console/namespaces?&accessToken={1}'.format(host_ip_, self.authID), requests.adapters.HTTPAdapter(max_retries=3))
        request = self.session.post(url='http://{0}:38848/nacos/v1/console/namespaces?&accessToken={1}'.format(host_ip_, self.authID), headers=headers, data=data, verify=False)
        response = request.json()
        if response != False:
            print(u"创建命名空间: \033[;32m%s\033[0m 成功! 返回值为: \033[;32m%s\033[0m" % (
                "sjzx", response))
        elif response == False:
            print(u"创建命名空间: \033[;31m%s\033[0m 失败! 原因: \033[;31m%s\033[0m" % (
                "sjzx", response))
        exit(1)


class CusPoEdit(object):
    def __init__(self):
        self.excel_file = None
        self.lv_list_msgcomment = []
        self.lv_list_msgid = []
        self.lv_list_msgid_plural = []
        self.lv_list_msgstr = []
        self.lv_list_msgstr0 = []
        self.lv_list_msgctxt = []
        self.po = polib.POFile()

    def def_create_pofile(self, po_file, occurrences, msgid, msgid_plural, msgstr,msgstr_plural, msgctxt):
        with open(po_file, 'a', encoding='utf-8') as s_hosts:
            if occurrences != None:
                if occurrences != "":
                    s_hosts.write(str(occurrences) + '\r')
            if msgctxt != None:
                if msgctxt != "":
                    s_hosts.write('msgctxt ' + str(msgctxt) + '\r')
            if msgid != None:
                if msgid != "":
                    s_hosts.write('msgid ' + str(msgid) + '\r')
            if msgid_plural != None:
                if msgid_plural != "":
                    s_hosts.write('msgid_plural ' + str(msgid_plural) + '\r')
            if msgstr != None:
                if msgstr != "":
                    s_hosts.write('msgstr ' + str(msgstr) + '\r')
            if msgstr_plural != None:
                if msgstr_plural != "":
                    s_hosts.write('msgstr[0] ' + str(msgstr_plural) + '\r')
        s_hosts.close()

    def def_translate(self, po_file):
        po = polib.pofile(po_file)
        self.lv_list_msgcomment = []
        self.lv_list_msgid = []
        self.lv_list_msgid_plural = []
        self.lv_list_msgstr = []
        self.lv_list_msgstr0 = []
        self.lv_list_msgctxt = []
        i = 1
        for entry in po:
            pattern = re.compile(r'(?:#.*)\n(?:".*\n.*)|(?:#.*)\n')
            res = re.findall(pattern, str(entry).replace('"\n"',''))
            lv_str = ''
            result = []
            if len(res) == 0:
                self.lv_list_msgcomment.append("")
            for lv_tuple in res:
                result.append("".join(lv_tuple))
            for value in result:
                lv_str = lv_str + value
            self.lv_list_msgcomment.append(lv_str)
            # if i == 37:
            #     # print(self.lv_list_msgstr0)
            #     print(str(entry))
            #     print(res)
            #     exit(1)

            pattern = None
            res = None
            lv_tuple = None
            pattern = re.compile(r'[^# ]msgid (?:.*)\n(?:".*\n.*)|[^# ]msgid (?:.*)\n')
            res = re.findall(pattern, str(entry).replace('"\n"', ''))
            if len(res) == 0:
                self.lv_list_msgid.append("")
            for lv_tuple in res:
                self.lv_list_msgid.append("".join(str(lv_tuple).replace('msgid ', "").replace('\n', '')))
            # if i == 1:
            #     print(self.lv_list_msgid)
            #     print(str(entry))
            #     print(res)
            #     exit(1)

            pattern = None
            res = None
            lv_tuple = None
            pattern = re.compile(r'msgid_plural (?:.*)\n(?:".*\n.*)|msgid_plural (?:.*)\n')
            res = re.findall(pattern, str(entry).replace('"\n"',''))
            if len(res) == 0:
                self.lv_list_msgid_plural.append("")
            for lv_tuple in res:
                self.lv_list_msgid_plural.append("".join(str(lv_tuple).replace('msgid_plural ', "").replace('\n', '')))
            # if i == 37:
            #     # print(self.lv_list_msgstr0)
            #     print(str(entry))
            #     print(res)
            #     exit(1)

            pattern = None
            res = None
            lv_tuple = None
            pattern = re.compile(r'msgstr (?:.*)\n(?:".*\n.*)|msgstr (?:.*)\n')
            res = re.findall(pattern, str(entry).replace('"\n"',''))
            if len(res) == 0:
                self.lv_list_msgstr.append("")
            for lv_tuple in res:
                self.lv_list_msgstr.append("".join(str(lv_tuple).replace('msgstr ', "").replace('\n', '')))

            pattern = None
            res = None
            lv_tuple = None
            pattern = re.compile(r'msgstr\[0\] (?:.*)\n(?:".*\n.*)|msgstr\[0\] (?:.*)\n')
            res = re.findall(pattern, str(entry).replace('"\n"', ''))
            if len(res) == 0:
                self.lv_list_msgstr0.append("")
            for lv_tuple in res:
                self.lv_list_msgstr0.append("".join(str(lv_tuple).replace('msgstr[0] ', "").replace('\n', '')))
            # if i == 37:
            #     # print(self.lv_list_msgstr0)
            #     print(str(entry))
            #     print(res)
            #     exit(1)

            pattern = None
            res = None
            lv_tuple = None
            pattern = re.compile(r'msgctxt (?:.*)\n(?:".*\n.*)|msgctxt (?:.*)\n')
            res = re.findall(pattern, str(entry).replace('"\n"',''))
            if len(res) == 0:
                self.lv_list_msgctxt.append("")
            for lv_tuple in res:
                self.lv_list_msgctxt.append("".join(str(lv_tuple).replace('msgctxt ', "").replace('\n', '')))
            i = i + 1
        print(len(self.lv_list_msgcomment), len(self.lv_list_msgid), len(self.lv_list_msgid_plural), len(self.lv_list_msgstr),
              len(self.lv_list_msgstr0), len(self.lv_list_msgctxt))
        return {'msgcomment': self.lv_list_msgcomment, 'msgid': self.lv_list_msgid, 'msgid_plural': self.lv_list_msgid_plural,
                'msgstr': self.lv_list_msgstr, 'msgstr0': self.lv_list_msgstr0, 'msgctxt': self.lv_list_msgctxt}
